from __future__ import annotations

import csv
import contextlib
import io
import json
import logging
import os
import re
import runpy
import sqlite3
import subprocess
import sys
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from fnmatch import fnmatch
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import Body, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

APP_TITLE = "Operator Workbench"
REPO_ROOT = Path(os.environ.get("OPERATOR_UI_REPO_ROOT", "/srv/operator-stack-clean")).resolve()
SCRIPTS_DIR = REPO_ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from operator_workflows.excel_financial_extractor import (  # noqa: E402
    extract_financial_snapshot_from_workbook,
    snapshot_to_markdown,
    write_snapshot_json,
)
from operator_workflows.financial_mom_219128 import (  # noqa: E402
    build_219128_feb_mar_mom_report,
    default_feb_mar_paths,
)
from operator_workflows.project_id_utils import (  # noqa: E402
    extract_project_id_from_rel_and_name,
    is_valid_operator_project_id,
)

RUNTIME_ROOT = Path(os.environ.get("OPERATOR_UI_RUNTIME_ROOT", str(REPO_ROOT / "runtime" / "operator_ui"))).resolve()
SESSION_ROOT = RUNTIME_ROOT / "sessions"
RUN_HISTORY_FILE = RUNTIME_ROOT / "run_history.json"
RUNS_ROOT = Path(os.environ.get("OPERATOR_UI_RUNS_ROOT", str(REPO_ROOT / "runs"))).resolve()
REGISTRY_PATH = Path(
    os.environ.get(
        "OPERATOR_UI_WORKFLOW_REGISTRY",
        str(REPO_ROOT / "build_control" / "operator_local" / "09_WORKFLOW_REGISTRY.json"),
    )
).resolve()

SCRIPT_INIT = REPO_ROOT / "scripts" / "init_operator_run.py"
SCRIPT_RUN = REPO_ROOT / "scripts" / "run_workflow.py"
SCRIPT_VALIDATE = REPO_ROOT / "scripts" / "validate_operator_run.py"

WORKFLOW_MAP = {
    "compare": "wf_compare_markdown",
    "risk": "wf_extract_risk_lines",
    "financial": "wf_financial_markdown_delta",
}
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TEXT_SUFFIXES = {".md", ".txt", ".csv"}
MAX_UPLOAD_BYTES = int(os.environ.get("OPERATOR_UI_MAX_UPLOAD_BYTES", str(100 * 1024 * 1024)))
MAX_EXCEL_ROWS_PER_SHEET = 80
MAX_EXCEL_COLS = 16
WORKBOOK_PREVIEW_MAX_ROWS = int(os.environ.get("OPERATOR_UI_WORKBOOK_PREVIEW_MAX_ROWS", "40"))
WORKBOOK_PREVIEW_MAX_COLS = int(os.environ.get("OPERATOR_UI_WORKBOOK_PREVIEW_MAX_COLS", "12"))
WORKBOOK_HEADER_SCAN_ROWS = int(os.environ.get("OPERATOR_UI_WORKBOOK_HEADER_SCAN_ROWS", "12"))
DEFAULT_LOCAL_WORKSPACE = Path(
    os.environ.get("OPERATOR_UI_DEFAULT_LOCAL_WORKSPACE", str(REPO_ROOT))
).resolve()
ALLOWED_LOCAL_WORKSPACES = [
    Path(p).resolve()
    for p in os.environ.get(
        "OPERATOR_UI_ALLOWED_WORKSPACES",
        f"{REPO_ROOT}{os.pathsep}{RUNS_ROOT}",
    ).split(os.pathsep)
    if p.strip()
]
READABLE_LOCAL_SUFFIXES = {
    ".md",
    ".txt",
    ".json",
    ".csv",
    ".yaml",
    ".yml",
    ".log",
    ".diff",
}
MAX_LOCAL_FILE_READ_BYTES = int(os.environ.get("OPERATOR_UI_MAX_LOCAL_FILE_READ_BYTES", str(300_000)))
PAIRING_LOW_CONFIDENCE_THRESHOLD = float(os.environ.get("OPERATOR_UI_PAIRING_LOW_CONFIDENCE_THRESHOLD", "0.66"))
# UI driver table in templates/index.html: primary = top N by |Δ| or |Δ| ≥ threshold (all remaining → audit). Display-only.
OPERATOR_UI_DRIVER_TABLE_PRIMARY_N = int(os.environ.get("OPERATOR_UI_DRIVER_TABLE_PRIMARY_N", "5"))
OPERATOR_UI_DRIVER_TABLE_IMPACT_USD = float(os.environ.get("OPERATOR_UI_DRIVER_TABLE_IMPACT_USD", "10000"))

STATE_DIR = (RUNTIME_ROOT / "state").resolve()
WORKSPACE_CONFIG_PATH = STATE_DIR / "workspace_config.json"
WORKSPACE_INDEX_DB = STATE_DIR / "workspace_index.db"
_INDEX_SCAN_MAX_FILES = int(os.environ.get("OPERATOR_UI_INDEX_MAX_FILES", "4000"))
# _count_workbooks_under: cap applies to matched Excel files; walk bound applies to file names seen.
_COUNT_WORKBOOKS_MAX = int(os.environ.get("OPERATOR_UI_COUNT_WORKBOOKS_MAX", "5000"))
_COUNT_WORKBOOKS_WALK_MAX_FILES = int(
    os.environ.get("OPERATOR_UI_COUNT_WORKBOOKS_WALK_MAX_FILES", "2000000")
)
# Internal token for unlabeled / no-project-id workbook groups (UI + index filters).
UNLABELED_GROUP_ID = "__unlabeled__"

# Log via uvicorn's logger so INFO/exception lines show in the uvicorn process terminal.
_uvicorn_log = logging.getLogger("uvicorn.error")

app = FastAPI(title=APP_TITLE, version="1.0.0")
# Local operator UI: avoid opaque browser "Failed to fetch" when any proxy/preview uses a
# different browser-visible origin than the API (CORS). Not for internet-facing deployment.
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.environ.get("OPERATOR_UI_CORS_ORIGINS", "*").split(","),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
templates = Jinja2Templates(directory=str(Path(__file__).resolve().parent / "templates"))
app.mount("/static", StaticFiles(directory=str(Path(__file__).resolve().parent / "static")), name="static")


@app.middleware("http")
async def _request_access_log(request: Request, call_next):
    response = await call_next(request)
    _uvicorn_log.info(
        "[operator_local_ui] %s %s -> %s",
        request.method,
        request.url.path,
        response.status_code,
    )
    return response


@dataclass
class SessionState:
    run_id: str
    status: str
    workflow_mode: str
    workflow_name: str | None
    uploaded_files: list[dict[str, Any]]
    created_at: str
    finished_at: str | None = None
    error: str | None = None
    run_dir: str | None = None
    envelope: dict[str, Any] | None = None
    structured_output: dict[str, Any] | None = None
    outputs_dir: str | None = None

    def to_json(self) -> dict[str, Any]:
        return {
            "run_id": self.run_id,
            "status": self.status,
            "workflow_mode": self.workflow_mode,
            "workflow_name": self.workflow_name,
            "uploaded_files": self.uploaded_files,
            "created_at": self.created_at,
            "finished_at": self.finished_at,
            "error": self.error,
            "run_dir": self.run_dir,
            "envelope": self.envelope,
            "structured_output": self.structured_output,
            "outputs_dir": self.outputs_dir,
        }


@dataclass
class LocalActionSpec:
    key: str
    description: str
    requires_confirmation: bool = False


@dataclass
class LocalWorkspaceContext:
    root: Path


SUGGESTED_FIRST_COMPARE = "compare latest report"

# Contracts that need at least one .xlsx/.xlsm/… in the workspace tree to proceed meaningfully.
CONTRACTS_NEEDING_WORKBOOK_FILES: frozenset[str] = frozenset(
    {
        "compare_latest_report",
        "compare_latest_prior_reports",
        "compare_and_show_labor_deltas",
        "run_weekly_review",
        "find_latest_prior_reports",
        "inspect_workbook",
        "preview_report_sheet",
        "find_report_sheets",
        "generate_financial_signals",
    }
)

# Commands that need a completed compare (outputs on disk) before they can run.
CONTRACTS_NEEDING_LATEST_PASS_RUN: frozenset[str] = frozenset(
    {
        "summarize_for_owner",
        "export_top_changes",
        "assess_cost_vs_revenue",
        "list_current_run_artifacts",
    }
)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _run_id() -> str:
    return f"ui_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"


def _safe_name(name: str) -> str:
    base = Path(name).name
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._")
    return cleaned or "upload.bin"


def _ensure_dirs() -> None:
    SESSION_ROOT.mkdir(parents=True, exist_ok=True)
    RUNTIME_ROOT.mkdir(parents=True, exist_ok=True)
    RUNS_ROOT.mkdir(parents=True, exist_ok=True)
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    _index_ensure_schema()


def _session_dir(run_id: str) -> Path:
    return SESSION_ROOT / run_id


def _session_state_path(run_id: str) -> Path:
    return _session_dir(run_id) / "state.json"


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def _default_workspace_config() -> dict[str, Any]:
    return {
        "version": 1,
        "default_workspace_root": "",
        "allowed_workspace_roots": [str(p) for p in ALLOWED_LOCAL_WORKSPACES],
        "last_selected_project": "",
        "last_confirmed_pair": {},
        "indexed_at": "",
        "last_compare_at": "",
        "last_compare_project_id": "",
        "preferred_report_family": "",
    }


def _load_workspace_config() -> dict[str, Any]:
    base = _default_workspace_config()
    if not WORKSPACE_CONFIG_PATH.is_file():
        return base
    try:
        data = _load_json(WORKSPACE_CONFIG_PATH)
    except Exception:  # noqa: BLE001
        return base
    if not isinstance(data, dict):
        return base
    merged: dict[str, Any] = {**base, **data}
    return merged


def _save_workspace_config(updates: dict[str, Any]) -> None:
    cur = _load_workspace_config()
    cur.update({k: v for k, v in updates.items() if v is not None})
    _save_json(WORKSPACE_CONFIG_PATH, cur)


def _persist_compare_memory_to_config(selected_pair: dict[str, Any], plan: dict[str, Any]) -> None:
    if not plan.get("prior_path") or not plan.get("current_path"):
        return
    curp = str((selected_pair.get("current") or {}).get("project_id") or "").strip()
    prvp = str((selected_pair.get("prior") or {}).get("project_id") or "").strip()
    pid = curp or prvp
    _save_workspace_config(
        {
            "last_confirmed_pair": {
                "prior_path": str(plan.get("prior_path") or ""),
                "current_path": str(plan.get("current_path") or ""),
                "pair_id": str(selected_pair.get("pair_id") or ""),
            },
            "last_selected_project": pid,
            "last_compare_at": _utc_now_iso(),
            "last_compare_project_id": pid,
            "preferred_report_family": str(
                (selected_pair.get("current") or {}).get("report_family")
                or (selected_pair.get("prior") or {}).get("report_family")
                or ""
            ),
        }
    )


def _validate_allowed_root_path(raw: str) -> Path:
    p = Path(raw).resolve()
    if not p.exists() or not p.is_dir():
        raise HTTPException(status_code=400, detail=f"Path is not an existing directory: {p}")
    if not _is_allowed_workspace_root(p):
        raise HTTPException(
            status_code=403,
            detail="Workspace path is not under an approved root for this Operator Local instance.",
        )
    return p


def _index_db_connect() -> sqlite3.Connection:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(str(WORKSPACE_INDEX_DB), timeout=30.0)
    con.row_factory = sqlite3.Row
    return con


def _index_ensure_schema() -> None:
    con = _index_db_connect()
    try:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS workbooks (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              root TEXT NOT NULL,
              rel_path TEXT NOT NULL,
              filename TEXT NOT NULL,
              extension TEXT NOT NULL,
              size_bytes INTEGER,
              mtime_utc TEXT,
              project_id TEXT,
              report_family TEXT,
              version_date TEXT,
              version_sort INTEGER,
              source_type TEXT,
              confidence REAL,
              rank_score REAL,
              sheet_count INTEGER,
              report_sheets_json TEXT,
              updated_at TEXT,
              UNIQUE(root, rel_path)
            )
            """
        )
        con.execute("CREATE INDEX IF NOT EXISTS idx_wb_root ON workbooks(root)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_wb_proj ON workbooks(root, project_id)")
        con.commit()
    finally:
        con.close()


def _index_delete_root(con: sqlite3.Connection, root: str) -> None:
    con.execute("DELETE FROM workbooks WHERE root = ?", (root,))


def _row_to_model(d: dict[str, Any] | sqlite3.Row) -> dict[str, Any]:
    if isinstance(d, sqlite3.Row):
        d = {k: d[k] for k in d.keys()}
    return {
        "path": str(d.get("rel_path", "")),
        "name": str(d.get("filename", "")),
        "modified_at": str(d.get("mtime_utc", "")),
        "rank_score": float(d.get("rank_score") or 0),
        "project_id": str(d.get("project_id") or ""),
        "report_family": str(d.get("report_family") or ""),
        "version_date": str(d.get("version_date") or ""),
        "version_sort": int(d.get("version_sort") or 0),
        "source_type": str(d.get("source_type") or "workspace"),
        "confidence": float(d.get("confidence") or 0.0),
    }


def _index_workbook_count_for_root(root: str) -> int:
    con = _index_db_connect()
    try:
        row = con.execute("SELECT COUNT(1) AS c FROM workbooks WHERE root = ?", (root,)).fetchone()
        return int(row[0]) if row else 0
    except Exception:  # noqa: BLE001
        return 0
    finally:
        con.close()


def _enrich_excel_for_index(
    workspace_root: Path, rel: str, _name: str, _rel_for_score: str, rank_score: float
) -> tuple[int, str, float]:
    """One read_only workbook open: sheet count + top report-sheets heuristic."""
    try:
        target, _ = _safe_workspace_relative_path(workspace_root, rel)
    except HTTPException:
        return 0, "[]", float(rank_score)
    if target.suffix.lower() not in EXCEL_SUFFIXES:
        return 0, "[]", float(rank_score)
    try:
        wb = load_workbook(filename=str(target), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return 0, "[]", float(rank_score)
    try:
        n = len(wb.worksheets)
        likely: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            name_l = (ws.title or "").lower()
            sc = 0
            if any(k in name_l for k in ("profit", "p&l", "income", "summary", "financial")):
                sc += 5
            if sc > 0 or any(k in name_l for k in ("p&l", "profit", "income", "result")):
                likely.append({"sheet_name": ws.title, "score": sc + 2, "reasons": ["heuristic"]})
        likely = sorted(likely, key=lambda x: int(x.get("score", 0)), reverse=True)[:12]
        rj = json.dumps(likely) if likely else "[]"
        if n == 0:
            n = 1
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return n, rj, float(rank_score)


def _scan_one_workbook_for_index(
    workspace_root: Path, rel: str, con: sqlite3.Connection, now: str
) -> bool:
    p = workspace_root / rel
    if not p.is_file() or p.suffix.lower() not in EXCEL_SUFFIXES:
        return False
    try:
        fr = _file_row(workspace_root, p)
    except Exception:  # noqa: BLE001
        return False
    name = p.name
    rscore = float(_score_workbook_candidate(p, workspace_root, ""))
    fam = _canonical_report_family(name)
    vdate, vsort = _extract_version_date(name)
    project_id = _extract_project_id(rel, name)
    conf = min(1.0, rscore / 220.0)
    if fam:
        conf = min(1.0, conf + 0.08)
    if vsort:
        conf = min(1.0, conf + 0.07)
    if project_id:
        conf = min(1.0, conf + 0.05)
    conf = round(min(conf, 1.0), 3)
    n_sheet, rj, _r2 = _enrich_excel_for_index(workspace_root, rel, name, rel, rscore)  # noqa: ARG001
    stype = _source_type_for_path(rel)
    con.execute(
        """
        INSERT OR REPLACE INTO workbooks (
          root, rel_path, filename, extension, size_bytes, mtime_utc,
          project_id, report_family, version_date, version_sort, source_type,
          confidence, rank_score, sheet_count, report_sheets_json, updated_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            str(workspace_root.resolve()),
            rel.replace("\\", "/"),
            name,
            p.suffix.lower().lstrip("."),
            int(fr.get("size_bytes") or 0),
            str(fr.get("modified_at") or ""),
            project_id,
            fam,
            vdate,
            int(vsort or 0),
            stype,
            conf,
            rscore,
            int(n_sheet or 0),
            rj,
            now,
        ),
    )
    return True


def _scan_workspace_path(root: Path, *, cap_files: int = _INDEX_SCAN_MAX_FILES) -> dict[str, Any]:
    if not root.exists() or not root.is_dir():
        return {"ok": False, "error": f"not a directory: {root}", "rows_indexed": 0, "root": str(root)}
    con = _index_db_connect()
    n_ok = 0
    t0 = datetime.now(timezone.utc)
    rroot = str(root.resolve())
    now = t0.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    try:
        _index_delete_root(con, rroot)
        n_wx = 0
        for p in root.rglob("*"):
            if n_wx >= cap_files:
                break
            if not p.is_file():
                continue
            if p.suffix.lower() not in EXCEL_SUFFIXES:
                continue
            rel = str(p.relative_to(root)).replace("\\", "/")
            if _scan_one_workbook_for_index(root, rel, con, now):
                n_ok += 1
            n_wx += 1
        con.commit()
    except Exception as exc:  # noqa: BLE001
        con.rollback()
        return {
            "ok": False,
            "error": str(exc),
            "rows_indexed": 0,
            "root": rroot,
        }
    finally:
        con.close()
    _save_workspace_config(
        {
            "indexed_at": now,
            "default_workspace_root": rroot,
        }
    )
    return {
        "ok": True,
        "rows_indexed": n_ok,
        "root": rroot,
        "indexed_at": now,
        "seconds": (datetime.now(timezone.utc) - t0).total_seconds(),
        "files_examined": n_wx,
        "cap_files": cap_files,
        "cap_reached": n_wx >= cap_files,
    }


def _index_candidate_models(
    workspace_root: Path, query: str, project_filter: str
) -> list[dict[str, Any]]:
    root = str(workspace_root.resolve())
    con = _index_db_connect()
    out: list[dict[str, Any]] = []
    try:
        if project_filter in (UNLABELED_GROUP_ID, "(unlabeled)"):
            qsql = (
                "SELECT * FROM workbooks WHERE root = ? AND (project_id IS NULL OR TRIM(project_id) = '') "
                "ORDER BY version_sort DESC, mtime_utc DESC"
            )
            rows = con.execute(qsql, (root,)).fetchall()
        elif project_filter:
            qsql = "SELECT * FROM workbooks WHERE root = ? AND project_id = ? ORDER BY version_sort DESC, mtime_utc DESC"
            rows = con.execute(qsql, (root, project_filter)).fetchall()
        else:
            rows = con.execute(
                "SELECT * FROM workbooks WHERE root = ? ORDER BY rank_score DESC, mtime_utc DESC",
                (root,),
            ).fetchall()
        for row in rows:
            m = _row_to_model(row)
            out.append(m)
    except Exception:  # noqa: BLE001
        return []
    finally:
        con.close()
    # Full natural-language commands (e.g. "compare latest report for 219128") are not filename
    # substrings; when SQL already scoped by project_id, do not filter them away.
    if query.strip() and not project_filter:
        qt = query.strip().lower()
        out = [m for m in out if qt in (m.get("name") or "").lower() or qt in (m.get("path") or "").lower()]
    return out


def _trend_sort_key(row: dict[str, Any]) -> tuple[int, int, str, str]:
    version_sort = int(row.get("version_sort", 0) or 0)
    modified_at = str(row.get("modified_at") or "")
    path = str(row.get("path") or "")
    if version_sort:
        return (0, version_sort, modified_at, path)
    fallback_sort = 99991231
    date_match = re.search(r"\b(20\d{2})-(\d{2})-(\d{2})\b", modified_at)
    if date_match:
        y, mo, d = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
        fallback_sort = y * 10_000 + mo * 100 + d
    return (1, fallback_sort, modified_at, path)


def _report_period_label(row: dict[str, Any]) -> str:
    version_date = str(row.get("version_date") or "").strip()
    if version_date:
        return version_date
    modified_at = str(row.get("modified_at") or "").strip()
    if modified_at:
        return modified_at
    return str(row.get("path") or "")


def _trend_report_ref(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(row, dict):
        return None
    return {
        "sequence": int(row.get("sequence", 0) or 0),
        "path": str(row.get("path") or ""),
        "name": str(row.get("name") or ""),
        "report_family": str(row.get("report_family") or "unknown"),
        "period": _report_period_label(row),
        "version_date": str(row.get("version_date") or ""),
        "modified_at": str(row.get("modified_at") or ""),
        "order_basis": str(row.get("order_basis") or ""),
    }


def _parse_report_date(row: dict[str, Any]) -> datetime | None:
    raw = str(row.get("version_date") or row.get("modified_at") or "").strip()
    if not raw:
        return None
    token = raw[:10]
    try:
        return datetime.strptime(token, "%Y-%m-%d")
    except ValueError:
        return None


def _cadence_observation(reports_used: list[dict[str, Any]]) -> str:
    dated = [_parse_report_date(row) for row in reports_used]
    dated = [d for d in dated if d is not None]
    report_count = len(reports_used)
    if report_count == 0:
        return "No indexed reports are available for this scoped project."
    if len(dated) < 2:
        return f"{len(dated)} of {report_count} report(s) have usable dates; cadence cannot be described from index metadata."
    span_days = (dated[-1] - dated[0]).days
    if len(dated) == report_count:
        return f"All {report_count} report(s) have usable dates spanning {span_days} day(s)."
    return f"{len(dated)} of {report_count} report(s) have usable dates spanning {span_days} day(s); undated reports use stable index fallback ordering."


def _build_owner_trend_summary(
    *,
    project_id: str,
    reports_used: list[dict[str, Any]],
    period_start: str,
    period_end: str,
    family_counts: dict[str, int],
) -> dict[str, Any]:
    report_count = len(reports_used)
    report_families = {family: family_counts[family] for family in sorted(family_counts)}
    latest = _trend_report_ref(reports_used[-1] if reports_used else None)
    prior = _trend_report_ref(reports_used[-2] if len(reports_used) >= 2 else None)
    oldest = _trend_report_ref(reports_used[0] if reports_used else None)
    family_line = ", ".join(f"{family}: {count}" for family, count in report_families.items()) or "none"
    cadence = _cadence_observation(reports_used)
    owner_lines = [
        f"Project {project_id} has {report_count} indexed report(s) in this scoped trend window.",
        f"The indexed period runs from {period_start or 'unknown'} to {period_end or 'unknown'}.",
    ]
    if latest and prior:
        owner_lines.append(
            f"The latest indexed report is {latest['name'] or latest['path']}; the prior indexed report is {prior['name'] or prior['path']}."
        )
    elif latest:
        owner_lines.append(f"The only indexed report is {latest['name'] or latest['path']}.")
    if oldest:
        owner_lines.append(f"The oldest indexed report is {oldest['name'] or oldest['path']}.")
    owner_lines.append(f"Report families in this window: {family_line}. {cadence}")
    return {
        "project_filter": project_id,
        "report_count": report_count,
        "period_start": period_start,
        "period_end": period_end,
        "report_families": report_families,
        "latest_report": latest,
        "prior_report": prior,
        "oldest_report": oldest,
        "cadence_observation": cadence,
        "owner_lines": owner_lines[:5],
    }


def _build_project_trend_artifact(workspace_root: Path, project_id: str) -> dict[str, Any]:
    rows = _index_candidate_models(workspace_root, "", project_id)
    ordered = sorted(rows, key=_trend_sort_key)
    reports_used: list[dict[str, Any]] = []
    family_counts: dict[str, int] = {}
    for idx, row in enumerate(ordered, start=1):
        family = str(row.get("report_family") or "unknown")
        family_counts[family] = family_counts.get(family, 0) + 1
        reports_used.append(
            {
                "sequence": idx,
                "path": str(row.get("path") or ""),
                "name": str(row.get("name") or ""),
                "report_family": family,
                "version_date": str(row.get("version_date") or ""),
                "version_sort": int(row.get("version_sort", 0) or 0),
                "modified_at": str(row.get("modified_at") or ""),
                "order_basis": "version_date" if int(row.get("version_sort", 0) or 0) else "modified_at_or_path",
                "source_type": str(row.get("source_type") or "workspace"),
            }
        )
    source_line = _workspace_source_line_for_transcript(
        index_backed=_index_workbook_count_for_root(str(workspace_root)) > 0
    )
    period_start = _report_period_label(ordered[0]) if ordered else ""
    period_end = _report_period_label(ordered[-1]) if ordered else ""
    status = "completed" if len(ordered) >= 3 else "insufficient_data"
    trend_summary = {
        "status": status,
        "basis": "workspace_index_metadata",
        "ordering": "oldest-to-newest by version_date/version_sort, then modified_at, then path",
        "family_counts": family_counts,
        "message": (
            f"Project {project_id} has {len(ordered)} indexed report(s); trend requires at least 3."
            if status == "insufficient_data"
            else f"Project {project_id} has {len(ordered)} indexed report(s) ordered for deterministic trend review."
        ),
    }
    owner_trend_summary = _build_owner_trend_summary(
        project_id=project_id,
        reports_used=reports_used,
        period_start=period_start,
        period_end=period_end,
        family_counts=family_counts,
    )
    latest_workbook_profit_summary: dict[str, Any] | None = None
    if reports_used:
        last_path = workspace_root / str(reports_used[-1].get("path") or "")
        if last_path.suffix.lower() in EXCEL_SUFFIXES and last_path.is_file():
            try:
                snap = extract_financial_snapshot_from_workbook(last_path)
                wps = snap.get("workbook_profit_summary")
                latest_workbook_profit_summary = wps if isinstance(wps, dict) else None
            except OSError:
                latest_workbook_profit_summary = None
            except Exception:  # noqa: BLE001
                latest_workbook_profit_summary = None
    return {
        "project_filter": project_id,
        "report_count": len(ordered),
        "reports_used": reports_used,
        "period_start": period_start,
        "period_end": period_end,
        "trend_summary": trend_summary,
        "owner_trend_summary": owner_trend_summary,
        "latest_workbook_profit_summary": latest_workbook_profit_summary,
        "source_line": source_line,
        "provenance": source_line,
    }


def _build_multi_report_compare_artifact(
    workspace_root: Path,
    project_id: str,
    requested_report_count: int,
) -> dict[str, Any]:
    rows = _index_candidate_models(workspace_root, "", project_id)
    ordered = sorted(rows, key=_trend_sort_key)
    available_count = len(ordered)
    selected = ordered[-requested_report_count:] if available_count >= requested_report_count else ordered
    reports_used: list[dict[str, Any]] = []
    family_counts: dict[str, int] = {}
    for idx, row in enumerate(selected, start=1):
        family = str(row.get("report_family") or "unknown")
        family_counts[family] = family_counts.get(family, 0) + 1
        reports_used.append(
            {
                "sequence": idx,
                "path": str(row.get("path") or ""),
                "name": str(row.get("name") or ""),
                "report_family": family,
                "version_date": str(row.get("version_date") or ""),
                "version_sort": int(row.get("version_sort", 0) or 0),
                "modified_at": str(row.get("modified_at") or ""),
                "order_basis": "version_date" if int(row.get("version_sort", 0) or 0) else "modified_at_or_path",
                "source_type": str(row.get("source_type") or "workspace"),
            }
        )
    comparison_pairs: list[dict[str, Any]] = []
    for idx in range(max(0, len(reports_used) - 1)):
        prior = reports_used[idx]
        current = reports_used[idx + 1]
        comparison_pairs.append(
            {
                "pair_sequence": idx + 1,
                "from_report": _trend_report_ref(prior),
                "to_report": _trend_report_ref(current),
                "summary": (
                    f"Compare report {idx + 1} to report {idx + 2}: "
                    f"{prior.get('name') or prior.get('path')} -> {current.get('name') or current.get('path')}."
                ),
            }
        )
    source_line = _workspace_source_line_for_transcript(
        index_backed=_index_workbook_count_for_root(str(workspace_root)) > 0
    )
    period_start = _report_period_label(reports_used[0]) if reports_used else ""
    period_end = _report_period_label(reports_used[-1]) if reports_used else ""
    status = "completed" if available_count >= requested_report_count else "insufficient_data"
    family_line = ", ".join(f"{family}: {family_counts[family]}" for family in sorted(family_counts)) or "none"
    if status == "completed":
        owner_lines = [
            f"Project {project_id} has a {requested_report_count}-report indexed comparison window.",
            f"The selected window runs from {period_start or 'unknown'} to {period_end or 'unknown'}.",
            f"The window creates {len(comparison_pairs)} adjacent metadata pair(s) in oldest-to-newest order.",
            f"Report families in the selected window: {family_line}.",
        ]
    else:
        owner_lines = [
            f"Project {project_id} has {available_count} indexed report(s); this command requires {requested_report_count}.",
            "No global workspace fallback was used.",
        ]
    m_latest_wps: dict[str, Any] | None = None
    if reports_used:
        m_last = workspace_root / str(reports_used[-1].get("path") or "")
        if m_last.suffix.lower() in EXCEL_SUFFIXES and m_last.is_file():
            try:
                m_snap = extract_financial_snapshot_from_workbook(m_last)
                m_w = m_snap.get("workbook_profit_summary")
                m_latest_wps = m_w if isinstance(m_w, dict) else None
            except OSError:
                m_latest_wps = None
            except Exception:  # noqa: BLE001
                m_latest_wps = None
    return {
        "status": status,
        "project_filter": project_id,
        "requested_report_count": requested_report_count,
        "report_count": len(reports_used),
        "available_report_count": available_count,
        "reports_used": reports_used,
        "pair_count": len(comparison_pairs),
        "comparison_pairs": comparison_pairs,
        "period_start": period_start,
        "period_end": period_end,
        "latest_workbook_profit_summary": m_latest_wps,
        "source_line": source_line,
        "provenance": source_line,
        "owner_lines": owner_lines,
    }


def _workbook_model_for_relpath(workspace_root: Path, rel: str) -> dict[str, Any] | None:
    rel_n = str(rel or "").replace("\\", "/").strip()
    for m in _index_candidate_models(workspace_root, "", ""):
        if str(m.get("path") or "").replace("\\", "/") == rel_n:
            return m
    try:
        t, r = _safe_workspace_relative_path(workspace_root, rel_n)
    except HTTPException:
        return None
    if t.suffix.lower() not in EXCEL_SUFFIXES or not t.is_file():
        return None
    name = t.name
    fam = _canonical_report_family(name)
    vd, vs = _extract_version_date(name)
    mtime = t.stat().st_mtime
    mta = datetime.fromtimestamp(mtime, tz=timezone.utc).isoformat()
    return {
        "path": r,
        "name": name,
        "project_id": _extract_project_id(r, name),
        "report_family": fam,
        "version_date": vd,
        "version_sort": vs,
        "modified_at": mta,
        "source_type": "workspace",
    }


def _models_for_project_paths(
    workspace_root: Path, project_id: str, selected_paths: list[str]
) -> tuple[list[dict[str, Any]], str | None]:
    out: list[dict[str, Any]] = []
    for p in selected_paths:
        m = _workbook_model_for_relpath(workspace_root, p)
        if m is None:
            return [], f"Workbook not found or not under workspace: {p}"
        pid_m = str(m.get("project_id") or "").strip()
        if pid_m and str(project_id).strip() and pid_m != str(project_id).strip():
            return [], f"Path {p} is indexed as project {pid_m}, not {project_id}."
        out.append(m)
    return out, None


def _build_multi_artifact_from_selected_paths(
    workspace_root: Path, project_id: str, selected_paths: list[str]
) -> dict[str, Any] | str:
    models, err = _models_for_project_paths(workspace_root, project_id, selected_paths)
    if err:
        return err
    if len(models) < 3:
        return "At least three selected reports are required for trend / multi-report analysis."
    ordered = sorted(models, key=_trend_sort_key)
    selected = ordered
    available_count = len(ordered)
    reports_used: list[dict[str, Any]] = []
    family_counts: dict[str, int] = {}
    for idx, row in enumerate(selected, start=1):
        family = str(row.get("report_family") or "unknown")
        family_counts[family] = family_counts.get(family, 0) + 1
        reports_used.append(
            {
                "sequence": idx,
                "path": str(row.get("path") or ""),
                "name": str(row.get("name") or ""),
                "report_family": family,
                "version_date": str(row.get("version_date") or ""),
                "version_sort": int(row.get("version_sort", 0) or 0),
                "modified_at": str(row.get("modified_at") or ""),
                "order_basis": "version_date" if int(row.get("version_sort", 0) or 0) else "modified_at_or_path",
                "source_type": str(row.get("source_type") or "workspace"),
            }
        )
    comparison_pairs: list[dict[str, Any]] = []
    for idx in range(max(0, len(reports_used) - 1)):
        prior = reports_used[idx]
        current = reports_used[idx + 1]
        comparison_pairs.append(
            {
                "pair_sequence": idx + 1,
                "from_report": _trend_report_ref(prior),
                "to_report": _trend_report_ref(current),
                "summary": (
                    f"Compare report {idx + 1} to report {idx + 2}: "
                    f"{prior.get('name') or prior.get('path')} -> {current.get('name') or current.get('path')}."
                ),
            }
        )
    source_line = _workspace_source_line_for_transcript(
        index_backed=_index_workbook_count_for_root(str(workspace_root)) > 0
    )
    period_start = _report_period_label(reports_used[0]) if reports_used else ""
    period_end = _report_period_label(reports_used[-1]) if reports_used else ""
    m_latest_wps: dict[str, Any] | None = None
    if reports_used:
        m_last = workspace_root / str(reports_used[-1].get("path") or "")
        if m_last.suffix.lower() in EXCEL_SUFFIXES and m_last.is_file():
            try:
                m_snap = extract_financial_snapshot_from_workbook(m_last)
                m_w = m_snap.get("workbook_profit_summary")
                m_latest_wps = m_w if isinstance(m_w, dict) else None
            except OSError:
                m_latest_wps = None
            except Exception:  # noqa: BLE001
                m_latest_wps = None
    family_line = ", ".join(f"{family}: {family_counts[family]}" for family in sorted(family_counts)) or "none"
    owner_lines = [
        f"Project {project_id}: {len(reports_used)} report(s) selected for Financial Signals (oldest-to-newest).",
        f"Window from {period_start or 'unknown'} to {period_end or 'unknown'}.",
        f"Adjacent pair count: {len(comparison_pairs)}. Report families: {family_line}.",
    ]
    return {
        "status": "completed",
        "project_filter": project_id,
        "requested_report_count": len(selected_paths),
        "report_count": len(reports_used),
        "available_report_count": available_count,
        "reports_used": reports_used,
        "pair_count": len(comparison_pairs),
        "comparison_pairs": comparison_pairs,
        "period_start": period_start,
        "period_end": period_end,
        "latest_workbook_profit_summary": m_latest_wps,
        "source_line": source_line,
        "provenance": source_line,
        "owner_lines": owner_lines,
    }


def _mom_219128_feb_mar_payload() -> dict[str, Any]:
    """Deterministic Feb→Mar 219128 comparison (validated workbooks on disk), no browser math."""
    try:
        feb, mar = default_feb_mar_paths(REPO_ROOT)
        return build_219128_feb_mar_mom_report(feb, mar)
    except (OSError, FileNotFoundError, ValueError) as e:
        logging.getLogger("operator_local_ui.app").info("mom_219128_feb_mar skipped: %s", e)
        return {
            "available": False,
            "schema": "mom_219128_feb_mar_v1",
            "error": str(e),
        }


def _run_generate_financial_signals(
    workspace: LocalWorkspaceContext, payload: dict[str, Any]
) -> JSONResponse:
    task = payload.get("task_payload") if isinstance(payload.get("task_payload"), dict) else {}
    analysis = str(task.get("analysis_type") or "").strip()
    project_id = str(task.get("project_id") or "").strip()
    paths = [str(p) for p in list(task.get("selected_paths") or []) if str(p).strip()]
    q = "Report Builder: Run analysis"
    trace: list[dict[str, Any]] = [
        {
            "action": "report_builder.generate_financial_signals",
            "details": f"analysis={analysis!r} project_id={project_id!r} paths={len(paths)}",
        }
    ]
    if not project_id or not paths:
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="failed",
            answer="Select a project, one or more monthly reports, and an analysis type.",
            result={"financial_signals": None, "next_steps": []},
            trace=trace,
        )
    single_types = {
        "current_profit_snapshot",
        "projected_profit_breakdown",
        "labor_rate_profit_analysis",
    }
    pair_types = {"compare_two_reports", "cost_movement_signals"}
    n = len(paths)
    if analysis not in single_types and analysis not in pair_types and analysis != "trend_across_reports":
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="failed",
            answer="Unknown analysis type for Report Builder.",
            result={"next_steps": ["Choose a valid analysis type from the list."]},
            trace=trace,
        )
    if analysis in single_types and n != 1:
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="failed",
            answer="Select exactly one report for this analysis.",
            result={},
            trace=trace,
        )
    if analysis in pair_types and n != 2:
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="failed",
            answer="Select exactly two reports for this analysis.",
            result={},
            trace=trace,
        )
    if analysis == "trend_across_reports" and n < 3:
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="insufficient_data",
            answer="Trend across reports needs at least three selected monthly reports.",
            result={},
            trace=trace,
        )

    if n == 1:
        models, perr = _models_for_project_paths(workspace.root, project_id, paths)
        if perr or not models:
            return _assistant_task_envelope(
                query=q,
                workspace_root=str(workspace.root),
                contract="generate_financial_signals",
                status="failed",
                answer=perr or "Workbook not found.",
                result={},
                trace=trace,
            )
        rel0 = str(models[0].get("path") or paths[0])
        try:
            t0, _r0 = _safe_workspace_relative_path(workspace.root, rel0)
        except HTTPException as e:
            return _assistant_task_envelope(
                query=q,
                workspace_root=str(workspace.root),
                contract="generate_financial_signals",
                status="failed",
                answer=str(e.detail),
                result={},
                trace=trace,
            )
        try:
            snap = extract_financial_snapshot_from_workbook(t0)
        except OSError as e:
            return _assistant_task_envelope(
                query=q,
                workspace_root=str(workspace.root),
                contract="generate_financial_signals",
                status="failed",
                answer=f"Could not read workbook: {e}",
                result={},
                trace=trace,
            )
        wps = snap.get("workbook_profit_summary") if isinstance(snap.get("workbook_profit_summary"), dict) else {}
        fb = snap.get("financial_workbench") if isinstance(snap.get("financial_workbench"), dict) else None
        out_block: dict[str, Any] = {
            "analysis_type": analysis,
            "project_id": project_id,
            "selected_paths": paths,
            "workbook_profit_summary": wps,
            "snapshot": snap,
            "financial_workbench": fb,
        }
        if project_id == "219128":
            out_block["mom_219128_feb_mar"] = _mom_219128_feb_mar_payload()
        if analysis == "labor_rate_profit_analysis":
            out_block["focus"] = "labor_rate_profit"
        elif analysis == "current_profit_snapshot":
            out_block["focus"] = "summary_profit"
        else:
            out_block["focus"] = "projected_profit_breakdown"
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="completed",
            answer="Analysis complete for the selected report (deterministic extraction).",
            result={
                "financial_signals": out_block,
                "source_line": _workspace_source_line_for_transcript(
                    index_backed=_index_workbook_count_for_root(str(workspace.root)) > 0
                ),
            },
            trace=trace,
        )

    if n == 2:
        models, perr = _models_for_project_paths(workspace.root, project_id, paths)
        if perr or len(models) != 2:
            return _assistant_task_envelope(
                query=q,
                workspace_root=str(workspace.root),
                contract="generate_financial_signals",
                status="failed",
                answer=perr or "Could not resolve both workbooks for the project.",
                result={},
                trace=trace,
            )
        ordered = sorted(models, key=_trend_sort_key)
        prior, current = ordered[0], ordered[1]
        plan = _build_compare_plan(
            workspace.root,
            prior_path=str(prior.get("path") or ""),
            current_path=str(current.get("path") or ""),
            requested_mode="financial",
            selection_reason="Report Builder: explicit two-report selection (oldest to newest by index metadata).",
        )
        compare = _execute_local_compare(
            workspace.root,
            prior_path=str(plan["prior_path"]),
            current_path=str(plan["current_path"]),
            workflow_mode="financial",
        )
        so = compare.get("structured_output", {}) if isinstance(compare.get("structured_output"), dict) else {}
        movers = _top_changes_from_structured_output(so, limit=20)
        out_block: dict[str, Any] = {
            "analysis_type": analysis,
            "project_id": project_id,
            "selected_paths": paths,
            "compare_run_id": str(compare.get("run_id", "") or ""),
            "pair": {"prior": plan["prior_path"], "current": plan["current_path"]},
            "structured_output": so,
            "largest_movers": movers[:10],
            "compare_summary": list(compare.get("summary") or []),
        }
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="completed",
            answer="Analysis complete (deterministic two-report compare).",
            result={
                "financial_signals": out_block,
                "run_payload": compare,
            },
            trace=trace,
        )

    # 3+ trend
    multi = _build_multi_artifact_from_selected_paths(workspace.root, project_id, paths)
    if isinstance(multi, str):
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="failed",
            answer=multi,
            result={},
            trace=trace,
        )
    if len(paths) < 3 or int(multi.get("report_count") or 0) < 3:
        return _assistant_task_envelope(
            query=q,
            workspace_root=str(workspace.root),
            contract="generate_financial_signals",
            status="insufficient_data",
            answer="Select at least three reports for trend analysis.",
            result={"found": multi},
            trace=trace,
        )
    multi_period_delta = _build_multi_period_delta(workspace.root, multi)
    out_block = {
        "analysis_type": "trend_across_reports",
        "project_id": project_id,
        "selected_paths": paths,
        "multi_artifact": multi,
        "multi_period_delta": multi_period_delta,
    }
    return _assistant_task_envelope(
        query=q,
        workspace_root=str(workspace.root),
        contract="generate_financial_signals",
        status="completed",
        answer="Analysis complete (trend and cost impact from selected reports).",
        result={"financial_signals": out_block},
        trace=trace,
    )


def _movement_key(row: dict[str, Any]) -> str:
    raw = str(row.get("category_label") or row.get("category") or "").strip().lower()
    return re.sub(r"\s+", " ", raw)


def _movement_label(row: dict[str, Any]) -> str:
    raw = str(row.get("category_label") or row.get("category") or "").strip()
    return re.sub(r"\s+", " ", raw) if raw else "Line item"


def _movement_delta(row: dict[str, Any]) -> float:
    try:
        return float(row.get("delta", 0) or 0)
    except (TypeError, ValueError):
        return 0.0


_ALLOWED_MOVEMENT_CATEGORIES: tuple[str, ...] = (
    "budget_change",
    "commitment_change",
    "actual_cost_change",
    "projected_cost_change",
    "profit_or_margin_change",
    "uncategorized",
)


def _normalize_movement_category(value: Any) -> str:
    raw = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    aliases = {
        "budget": "budget_change",
        "budget_change": "budget_change",
        "commitment": "commitment_change",
        "commitment_change": "commitment_change",
        "committed_cost": "commitment_change",
        "actual": "actual_cost_change",
        "actual_cost": "actual_cost_change",
        "actual_cost_change": "actual_cost_change",
        "projected": "projected_cost_change",
        "projection": "projected_cost_change",
        "projected_cost": "projected_cost_change",
        "projected_cost_change": "projected_cost_change",
        "forecast": "projected_cost_change",
        "profit": "profit_or_margin_change",
        "margin": "profit_or_margin_change",
        "profit_or_margin": "profit_or_margin_change",
        "profit_or_margin_change": "profit_or_margin_change",
        "uncategorized": "uncategorized",
    }
    return aliases.get(raw, "")


def _classification_basis(category: str, field: str, value: str, rule: str, confidence: str) -> dict[str, str]:
    return {
        "category": category,
        "matched_field": field,
        "matched_value": value,
        "matched_rule": rule,
        "confidence": confidence,
    }


def _cost_type_basis(cost_type: str, field: str, value: str, rule: str, confidence: str) -> dict[str, str]:
    return {
        "cost_type": cost_type,
        "matched_field": field,
        "matched_value": value,
        "matched_rule": rule,
        "confidence": confidence,
    }


def _normalize_cost_type(value: Any) -> str:
    raw = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    aliases = {
        "labor": "labor",
        "lab": "labor",
        "payroll": "labor",
        "material": "material",
        "materials": "material",
        "mat": "material",
        "sub": "subcontract",
        "subcontract": "subcontract",
        "subcontractor": "subcontract",
        "subcontractors": "subcontract",
        "supplier_sub": "subcontract",
        "uncategorized": "uncategorized",
    }
    return aliases.get(raw, "")


def _classify_cost_type(row: dict[str, Any]) -> dict[str, str]:
    structured_fields = (
        "cost_type",
        "cost_category",
        "cost_class",
        "phase_type",
        "account_type",
        "normalized_cost_type",
        "category",
    )
    for field in structured_fields:
        if field not in row:
            continue
        raw = str(row.get(field) or "").strip()
        ctype = _normalize_cost_type(raw)
        if ctype and ctype != "uncategorized":
            return _cost_type_basis(ctype, field, raw, "explicit_structured_field", "structured")

    keyword_fields = ("label", "key", "category_label", "description", "category", "text", "account", "notes")
    keyword_rules: list[tuple[str, tuple[str, ...]]] = [
        ("labor", ("| lab", " labor", "lab ", "payroll", "wage", "foreman", "project manager")),
        ("material", ("| mat", " material", "materials", "supplier", "blocking", "toilet", "sealant")),
        ("subcontract", ("| sub", " subcontract", "subcontractor", "subcontractors", " sub ")),
    ]
    for field in keyword_fields:
        raw = str(row.get(field) or "")
        text = f" {raw.lower()} "
        if not raw:
            continue
        for ctype, tokens in keyword_rules:
            for token in tokens:
                if token in text:
                    return _cost_type_basis(ctype, field, raw, f"keyword:{token.strip()}", "keyword")
    return _cost_type_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized")


def _classify_movement(row: dict[str, Any]) -> dict[str, str]:
    structured_fields = (
        "movement_category",
        "movement_type",
        "change_type",
        "line_type",
        "account_type",
        "cost_type",
        "section",
    )
    for field in structured_fields:
        if field not in row:
            continue
        raw = str(row.get(field) or "").strip()
        cat = _normalize_movement_category(raw)
        if cat and cat != "uncategorized":
            return _classification_basis(cat, field, raw, "explicit_structured_field", "structured")

    keyword_fields = ("key", "label", "category_label", "category", "tier", "account", "notes", "text")
    keyword_rules: list[tuple[str, tuple[str, ...]]] = [
        ("budget_change", ("budget", "original budget", "revised budget", "approved budget")),
        ("commitment_change", ("commitment", "committed", "subcontract", "contract", "purchase order", " po ")),
        ("actual_cost_change", ("actual", "actual cost", "cost to date", "incurred")),
        ("projected_cost_change", ("projected", "forecast", "eac", "estimate at completion", "cost to complete", "etc")),
        ("profit_or_margin_change", ("profit", "margin", "fee", "gross margin", "p/l", "pl ")),
    ]
    for field in keyword_fields:
        raw = str(row.get(field) or "")
        text = raw.lower()
        if not text:
            continue
        for cat, tokens in keyword_rules:
            for token in tokens:
                if token in text:
                    return _classification_basis(cat, field, raw, f"keyword:{token.strip()}", "keyword")
    return _classification_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized")


def _top_changes_from_structured_output(structured_output: dict[str, Any], limit: int = 10) -> list[dict[str, Any]]:
    split = merge_and_split_driver_rows(structured_output)
    rows = list(split.get("primary") or [])
    if not rows:
        rows = list(split.get("audit") or [])
    out: list[dict[str, Any]] = []
    for idx, row in enumerate(rows[:limit], start=1):
        if not isinstance(row, dict):
            continue
        delta = _movement_delta(row)
        basis = _classify_movement(row)
        cost_basis = _classify_cost_type(row)
        out.append(
            {
                "rank": idx,
                "key": _movement_key(row),
                "label": _movement_label(row),
                "delta": delta,
                "abs_delta": abs(delta),
                "prior_value": row.get("prior_value"),
                "current_value": row.get("current_value"),
                "tier": str(row.get("tier") or ""),
                "movement_category": basis["category"],
                "classification_basis": basis,
                "cost_type": cost_basis["cost_type"],
                "cost_type_basis": cost_basis,
            }
        )
    return out


def _movement_category_summary(movers: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    categories = [
        "budget_change",
        "commitment_change",
        "actual_cost_change",
        "projected_cost_change",
        "profit_or_margin_change",
        "uncategorized",
    ]
    out: dict[str, dict[str, Any]] = {
        cat: {"count": 0, "cumulative_abs_delta": 0.0}
        for cat in categories
    }
    for mover in movers:
        cat = str(mover.get("movement_category") or "uncategorized")
        if cat not in out:
            cat = "uncategorized"
        out[cat]["count"] = int(out[cat]["count"]) + 1
        out[cat]["cumulative_abs_delta"] = round(
            float(out[cat]["cumulative_abs_delta"]) + float(mover.get("cumulative_abs_delta") or mover.get("abs_delta") or 0),
            2,
        )
    return out


def _classification_summary(movers: list[dict[str, Any]]) -> dict[str, Any]:
    confidence_counts = {"structured": 0, "keyword": 0, "uncategorized": 0}
    category_counts = {cat: 0 for cat in _ALLOWED_MOVEMENT_CATEGORIES}
    for mover in movers:
        cat = str(mover.get("movement_category") or "uncategorized")
        if cat not in category_counts:
            cat = "uncategorized"
        category_counts[cat] += 1
        basis = mover.get("classification_basis") if isinstance(mover.get("classification_basis"), dict) else {}
        conf = str(basis.get("confidence") or ("uncategorized" if cat == "uncategorized" else "keyword"))
        if conf not in confidence_counts:
            conf = "uncategorized"
        confidence_counts[conf] += 1
    return {
        "category_counts": category_counts,
        "confidence_counts": confidence_counts,
        "classified_count": confidence_counts["structured"] + confidence_counts["keyword"],
        "uncategorized_count": category_counts["uncategorized"],
        "total_mover_count": len(movers),
    }


def _action_item_from_mover(prefix: str, mover: dict[str, Any], amount_key: str = "cumulative_delta") -> str:
    label = str(mover.get("label") or mover.get("key") or "Line item")
    cat = str(mover.get("movement_category") or "uncategorized")
    amount = _format_usd_signed(float(mover.get(amount_key) or mover.get("delta") or 0))
    return f"{prefix}: review {label} ({cat}, {amount}) against the source workbook lines."


def _action_view_item(source: str, mover: dict[str, Any], amount_key: str) -> dict[str, Any]:
    amount = float(mover.get(amount_key) or mover.get("delta") or 0)
    label = str(mover.get("label") or mover.get("key") or "Line item")
    category = str(mover.get("movement_category") or "uncategorized")
    pair_sequences = list(mover.get("pair_sequences") or [])
    if not pair_sequences and mover.get("pair_sequence"):
        pair_sequences = [int(mover.get("pair_sequence") or 0)]
    return {
        "key": str(mover.get("key") or label.lower()),
        "label": label,
        "movement_category": category,
        "classification_basis": mover.get("classification_basis")
        if isinstance(mover.get("classification_basis"), dict)
        else _classification_basis(category, "", "", "propagated_category", "uncategorized" if category == "uncategorized" else "keyword"),
        "amount": amount,
        "abs_amount": abs(amount),
        "source": source,
        "pair_sequences": pair_sequences,
        "line": f"{label} ({category}, {_format_usd_signed(amount)})",
    }


def _build_action_view(
    *,
    largest_cumulative_movers: list[dict[str, Any]],
    latest_period_watchlist: list[dict[str, Any]],
    repeated_risk_items: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    used: set[str] = set()

    def add_unique(target: list[dict[str, Any]], source: str, rows: list[dict[str, Any]], amount_key: str, cap: int) -> None:
        for row in rows:
            item = _action_view_item(source, row, amount_key)
            key = item["key"]
            if key in used:
                continue
            target.append(item)
            used.add(key)
            if len(target) >= cap:
                break

    top_issues: list[dict[str, Any]] = []
    add_unique(top_issues, "largest_cumulative", largest_cumulative_movers, "cumulative_delta", 2)
    add_unique(top_issues, "latest_period", latest_period_watchlist, "delta", 3)
    add_unique(top_issues, "repeated", repeated_risk_items, "cumulative_delta", 3)

    new_this_period_source = [
        row
        for row in latest_period_watchlist
        if int(row.get("pair_count") or len(row.get("pair_sequences") or [])) <= 1
    ]
    new_this_period: list[dict[str, Any]] = []
    add_unique(new_this_period, "latest_period_only", new_this_period_source, "delta", 5)

    ongoing_risks: list[dict[str, Any]] = []
    add_unique(ongoing_risks, "repeated", repeated_risk_items, "cumulative_delta", 5)

    remaining = [*largest_cumulative_movers, *latest_period_watchlist, *repeated_risk_items]
    watchlist: list[dict[str, Any]] = []
    add_unique(watchlist, "remaining", remaining, "cumulative_delta", 5)
    return {
        "top_issues": top_issues[:3],
        "new_this_period": new_this_period[:5],
        "ongoing_risks": ongoing_risks[:5],
        "watchlist": watchlist[:5],
    }


def _build_cost_type_drilldown(movers: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    buckets: dict[str, dict[str, Any]] = {
        "labor": {"count": 0, "total_abs_movement": 0.0, "items": [], "pair_coverage": []},
        "material": {"count": 0, "total_abs_movement": 0.0, "items": [], "pair_coverage": []},
        "subcontract": {"count": 0, "total_abs_movement": 0.0, "items": [], "pair_coverage": []},
        "uncategorized": {"count": 0, "total_abs_movement": 0.0, "items": [], "pair_coverage": []},
    }
    pair_sets: dict[str, set[int]] = {k: set() for k in buckets}
    confidence_counts: dict[str, dict[str, int]] = {
        k: {"structured_count": 0, "keyword_count": 0, "uncategorized_count": 0}
        for k in buckets
    }
    sorted_movers = sorted(
        movers,
        key=lambda m: (float(m.get("cumulative_abs_delta") or m.get("abs_delta") or 0), str(m.get("label") or "")),
        reverse=True,
    )
    for mover in sorted_movers:
        bucket = str(mover.get("cost_type") or "uncategorized")
        if bucket not in buckets:
            bucket = "uncategorized"
        amount = float(mover.get("cumulative_abs_delta") or mover.get("abs_delta") or 0)
        pair_sequences = sorted(set(int(x) for x in list(mover.get("pair_sequences") or []) if int(x or 0)))
        basis = mover.get("cost_type_basis") if isinstance(mover.get("cost_type_basis"), dict) else {}
        confidence = str(basis.get("confidence") or ("uncategorized" if bucket == "uncategorized" else "keyword"))
        confidence_key = f"{confidence}_count" if confidence in {"structured", "keyword", "uncategorized"} else "uncategorized_count"
        buckets[bucket]["count"] = int(buckets[bucket]["count"]) + 1
        buckets[bucket]["total_abs_movement"] = round(float(buckets[bucket]["total_abs_movement"]) + amount, 2)
        confidence_counts[bucket][confidence_key] = int(confidence_counts[bucket][confidence_key]) + 1
        pair_sets[bucket].update(pair_sequences)
        if len(buckets[bucket]["items"]) < 10:
            buckets[bucket]["items"].append(
                {
                    "key": str(mover.get("key") or ""),
                    "label": str(mover.get("label") or mover.get("key") or "Line item"),
                    "movement_category": str(mover.get("movement_category") or "uncategorized"),
                    "cost_type": bucket,
                    "cost_type_basis": mover.get("cost_type_basis")
                    if isinstance(mover.get("cost_type_basis"), dict)
                    else _cost_type_basis(bucket, "", "", "propagated_cost_type", "uncategorized" if bucket == "uncategorized" else "keyword"),
                    "cumulative_delta": float(mover.get("cumulative_delta") or mover.get("delta") or 0),
                    "cumulative_abs_delta": amount,
                    "pair_sequences": pair_sequences,
                }
            )
    for bucket, pairs in pair_sets.items():
        buckets[bucket]["pair_coverage"] = sorted(pairs)
        buckets[bucket]["confidence_breakdown"] = confidence_counts[bucket]
    return buckets


def _build_multi_period_delta(workspace_root: Path, multi: dict[str, Any]) -> dict[str, Any]:
    pair_results: list[dict[str, Any]] = []
    limitations: list[str] = []
    movement_index: dict[str, dict[str, Any]] = {}
    comparison_pairs = list(multi.get("comparison_pairs") or [])
    for pair in comparison_pairs:
        prior_report = pair.get("from_report") if isinstance(pair.get("from_report"), dict) else {}
        current_report = pair.get("to_report") if isinstance(pair.get("to_report"), dict) else {}
        pair_result: dict[str, Any] = {
            "pair_sequence": int(pair.get("pair_sequence") or 0),
            "prior_report": prior_report,
            "current_report": current_report,
            "period_start": str(prior_report.get("period") or prior_report.get("version_date") or ""),
            "period_end": str(current_report.get("period") or current_report.get("version_date") or ""),
            "top_changes": [],
            "status": "not_run",
            "limitations": [],
        }
        try:
            compare_payload = _execute_local_compare(
                workspace_root,
                prior_path=str(prior_report.get("path") or ""),
                current_path=str(current_report.get("path") or ""),
                workflow_mode="financial",
            )
            structured_output = (
                compare_payload.get("structured_output", {}) if isinstance(compare_payload.get("structured_output"), dict) else {}
            )
            top_changes = _top_changes_from_structured_output(structured_output)
            pair_result["status"] = str(compare_payload.get("status") or "completed")
            pair_result["run_id"] = str(compare_payload.get("run_id") or "")
            pair_result["top_changes"] = top_changes
            if not top_changes:
                msg = f"Pair {pair_result['pair_sequence']} produced no deterministic top-change rows."
                pair_result["limitations"].append(msg)
                limitations.append(msg)
            for change in top_changes:
                key = str(change.get("key") or "")
                if not key:
                    continue
                entry = movement_index.setdefault(
                    key,
                    {
                        "key": key,
                        "label": str(change.get("label") or key),
                        "movement_category": str(change.get("movement_category") or "uncategorized"),
                        "classification_basis": change.get("classification_basis")
                        if isinstance(change.get("classification_basis"), dict)
                        else _classification_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized"),
                        "cost_type": str(change.get("cost_type") or "uncategorized"),
                        "cost_type_basis": change.get("cost_type_basis")
                        if isinstance(change.get("cost_type_basis"), dict)
                        else _cost_type_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized"),
                        "pair_sequences": [],
                        "pair_count": 0,
                        "cumulative_delta": 0.0,
                        "cumulative_abs_delta": 0.0,
                    },
                )
                entry["pair_sequences"].append(pair_result["pair_sequence"])
                entry["pair_count"] = int(entry["pair_count"]) + 1
                entry["cumulative_delta"] = float(entry["cumulative_delta"]) + float(change.get("delta") or 0)
                entry["cumulative_abs_delta"] = float(entry["cumulative_abs_delta"]) + float(change.get("abs_delta") or 0)
        except Exception as exc:  # noqa: BLE001
            msg = f"Pair {pair_result['pair_sequence']} compare could not run: {exc}"
            pair_result["status"] = "failed"
            pair_result["limitations"].append(msg)
            limitations.append(msg)
        pair_results.append(pair_result)

    movers = list(movement_index.values())
    for mover in movers:
        mover["pair_sequences"] = sorted(set(int(x) for x in mover.get("pair_sequences", [])))
        mover["cumulative_delta"] = round(float(mover.get("cumulative_delta") or 0), 2)
        mover["cumulative_abs_delta"] = round(float(mover.get("cumulative_abs_delta") or 0), 2)
    repeated_movers = sorted(
        [m for m in movers if int(m.get("pair_count") or 0) > 1],
        key=lambda m: (float(m.get("cumulative_abs_delta") or 0), str(m.get("label") or "")),
        reverse=True,
    )
    largest_cumulative_movers = sorted(
        movers,
        key=lambda m: (float(m.get("cumulative_abs_delta") or 0), str(m.get("label") or "")),
        reverse=True,
    )[:10]
    latest_period_movers = list(pair_results[-1].get("top_changes") or []) if pair_results else []
    movement_categories = _movement_category_summary(movers)
    classification_summary = _classification_summary(movers)
    uncategorized_count = int(classification_summary.get("uncategorized_count") or 0)
    action_items: list[str] = []
    action_keys: set[str] = set()
    for mover in largest_cumulative_movers[:3]:
        action_items.append(_action_item_from_mover("Cumulative mover", mover))
        action_keys.add(str(mover.get("key") or ""))
    repeated_added = 0
    for mover in repeated_movers:
        key = str(mover.get("key") or "")
        if key in action_keys:
            continue
        item = _action_item_from_mover("Repeated mover", mover)
        action_items.append(item)
        action_keys.add(key)
        repeated_added += 1
        if repeated_added >= 2:
            break
    latest_period_watchlist = [
        {
            "key": str(row.get("key") or ""),
            "label": str(row.get("label") or row.get("key") or "Line item"),
            "movement_category": str(row.get("movement_category") or "uncategorized"),
            "classification_basis": row.get("classification_basis")
            if isinstance(row.get("classification_basis"), dict)
            else _classification_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized"),
            "cost_type": str(row.get("cost_type") or "uncategorized"),
            "cost_type_basis": row.get("cost_type_basis")
            if isinstance(row.get("cost_type_basis"), dict)
            else _cost_type_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized"),
            "delta": float(row.get("delta") or 0),
            "pair_sequence": int(pair_results[-1].get("pair_sequence") or 0) if pair_results else 0,
        }
        for row in latest_period_movers[:5]
    ]
    repeated_risk_items = [
        {
            "key": str(row.get("key") or ""),
            "label": str(row.get("label") or row.get("key") or "Line item"),
            "movement_category": str(row.get("movement_category") or "uncategorized"),
            "classification_basis": row.get("classification_basis")
            if isinstance(row.get("classification_basis"), dict)
            else _classification_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized"),
            "cost_type": str(row.get("cost_type") or "uncategorized"),
            "cost_type_basis": row.get("cost_type_basis")
            if isinstance(row.get("cost_type_basis"), dict)
            else _cost_type_basis("uncategorized", "", "", "no_allowed_rule_matched", "uncategorized"),
            "pair_sequences": list(row.get("pair_sequences") or []),
            "pair_count": int(row.get("pair_count") or 0),
            "cumulative_delta": float(row.get("cumulative_delta") or 0),
            "cumulative_abs_delta": float(row.get("cumulative_abs_delta") or 0),
        }
        for row in repeated_movers[:5]
    ]
    action_view = _build_action_view(
        largest_cumulative_movers=largest_cumulative_movers,
        latest_period_watchlist=latest_period_watchlist,
        repeated_risk_items=repeated_risk_items,
    )
    cost_type_drilldown = _build_cost_type_drilldown(movers)
    successful_pairs = len([p for p in pair_results if str(p.get("status")) == "completed"])
    owner_lines = [
        f"Computed deterministic financial deltas for {successful_pairs} of {len(pair_results)} adjacent pair(s).",
    ]
    if largest_cumulative_movers:
        top = largest_cumulative_movers[0]
        owner_lines.append(
            f"Largest cumulative mover by absolute delta: {top.get('label')} ({_format_usd_signed(float(top.get('cumulative_delta') or 0))})."
        )
    if repeated_movers:
        owner_lines.append(f"{len(repeated_movers)} mover(s) appeared in more than one adjacent pair.")
    if latest_period_movers:
        owner_lines.append(
            f"Latest period leading mover: {latest_period_movers[0].get('label')} ({_format_usd_signed(float(latest_period_movers[0].get('delta') or 0))})."
        )
    if limitations:
        owner_lines.append("Review limitations before using this as an owner-ready trend readout.")
    return {
        "pair_results": pair_results,
        "repeated_movers": repeated_movers[:10],
        "largest_cumulative_movers": largest_cumulative_movers,
        "latest_period_movers": latest_period_movers[:10],
        "action_items": action_items[:5],
        "movement_categories": movement_categories,
        "classification_summary": classification_summary,
        "uncategorized_count": uncategorized_count,
        "latest_period_watchlist": latest_period_watchlist,
        "repeated_risk_items": repeated_risk_items,
        "action_view": action_view,
        "cost_type_drilldown": cost_type_drilldown,
        "owner_lines": owner_lines[:5],
        "limitations": limitations,
    }


def _project_index_summary_for_ui(workspace_root: Path) -> list[dict[str, Any]]:
    root = str(workspace_root.resolve())
    con = _index_db_connect()
    out: list[dict[str, Any]] = []
    try:
        rows = con.execute(
            """
            SELECT
              coalesce(nullif(project_id, ''), '(unlabeled)') AS pid,
              count(*) as n,
              max(version_date) as latest_vd,
              max(mtime_utc) as latest_m
            FROM workbooks
            WHERE root = ?
            GROUP BY coalesce(nullif(project_id, ''), '(unlabeled)')
            ORDER BY n DESC, latest_m DESC
            """,
            (root,),
        ).fetchall()
        cfg = _load_workspace_config()
        lproj = str(cfg.get("last_compare_project_id") or cfg.get("last_selected_project") or "")
        lcmp = str(cfg.get("last_compare_at") or "")
        for r in rows:
            raw_pid = str(r["pid"])
            is_unlabeled = raw_pid == "(unlabeled)"
            token = UNLABELED_GROUP_ID if is_unlabeled else raw_pid
            display_name = (
                "Unlabeled (no project ID in paths)" if is_unlabeled else raw_pid
            )
            if is_unlabeled:
                match = bool(lcmp) and lproj in ("", "(unlabeled)", UNLABELED_GROUP_ID)
            else:
                match = bool(lproj) and lproj == token
            out.append(
                {
                    "project_id": token,
                    "display_name": display_name,
                    "is_unlabeled": is_unlabeled,
                    "report_count": int(r["n"] or 0),
                    "latest_version_date": str(r["latest_vd"] or ""),
                    "latest_modified_at": str(r["latest_m"] or ""),
                    "last_compared": lcmp if match else "",
                    "is_last_project": match,
                }
            )
    except Exception:  # noqa: BLE001
        return []
    finally:
        con.close()
    return out


def _extract_project_id_from_query(q: str) -> str:
    s = (q or "").lower()
    for pat in (
        r"\b(?:for|project)\s+(\d{5,})\b",
        r"\breports (?:on|for) (?:project )?(\d{5,})\b",
        r"\bshow reports for (?:project )?(\d{5,})\b",
    ):
        m = re.search(pat, s, re.IGNORECASE)
        if m and is_valid_operator_project_id(m.group(1)):
            return m.group(1)
    qstrip = (q or "").strip()
    m2 = re.match(r"^\s*(\d{5,})\s*$", qstrip)
    if m2 and is_valid_operator_project_id(m2.group(1)):
        return m2.group(1)
    return ""


def _workbook_looked_paths_message(workspace: LocalWorkspaceContext, indexed: bool) -> str:
    cfg = _load_workspace_config()
    roots = [str(workspace.root)]
    extra = [str(p) for p in cfg.get("allowed_workspace_roots") or [] if isinstance(p, str)]
    idx = "workspace index" if indexed else "filesystem walk (index empty)"
    return (
        f"Looked for Excel under primary workspace: {roots[0]} "
        f"({idx}). Allowed policy roots: {', '.join(sorted(set(roots + extra))[:4])}."
    )


def _workspace_source_line_for_transcript(
    *, index_backed: bool, indexed_at: str = ""
) -> str:
    """
    One-line source label for the operator transcript: index vs live rglob, with timestamp
    from persisted config when the index is backed by SQLite.
    """
    ts = (indexed_at or str(_load_workspace_config().get("indexed_at") or "")).strip()
    if index_backed and ts:
        return f"Source: workspace index (indexed {ts})."
    if index_backed and not ts:
        return "Source: workspace index (not yet timestamped; run scan workspace)."
    return "Source: live scan (not indexed)."


def _scoping_project_id_from_query(q: str) -> str:
    """Return numeric project_id when the user scoped the command (e.g. '... for 219128')."""
    return _extract_project_id_from_query((q or "").strip())


def _no_pairing_for_scoped_project(project_id: str, pair_selection: dict[str, Any] | None) -> bool:
    if not (project_id or "").strip():
        return False
    if not isinstance(pair_selection, dict):
        return True
    sp = pair_selection.get("selected_pair")
    return not isinstance(sp, dict)


def _load_history() -> list[dict[str, Any]]:
    if not RUN_HISTORY_FILE.exists():
        return []
    try:
        data = _load_json(RUN_HISTORY_FILE)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _append_history(row: dict[str, Any]) -> None:
    rows = _load_history()
    rows.append(row)
    rows = rows[-30:]
    _save_json(RUN_HISTORY_FILE, rows)


def _save_state(state: SessionState) -> None:
    _save_json(_session_state_path(state.run_id), state.to_json())


def _load_state(run_id: str) -> SessionState:
    path = _session_state_path(run_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Run not found: {run_id}")
    payload = _load_json(path)
    return SessionState(
        run_id=payload["run_id"],
        status=payload["status"],
        workflow_mode=payload.get("workflow_mode", "auto"),
        workflow_name=payload.get("workflow_name"),
        uploaded_files=list(payload.get("uploaded_files", [])),
        created_at=payload.get("created_at", _utc_now_iso()),
        finished_at=payload.get("finished_at"),
        error=payload.get("error"),
        run_dir=payload.get("run_dir"),
        envelope=payload.get("envelope"),
        structured_output=payload.get("structured_output"),
        outputs_dir=payload.get("outputs_dir"),
    )


def _is_allowed_workspace_root(path: Path) -> bool:
    for allowed in {DEFAULT_LOCAL_WORKSPACE, *ALLOWED_LOCAL_WORKSPACES}:
        if path == allowed or allowed in path.parents:
            return True
    return False


def _workspace_context(requested_root: str | None) -> LocalWorkspaceContext:
    root = Path(requested_root).resolve() if requested_root else DEFAULT_LOCAL_WORKSPACE
    if not root.exists() or not root.is_dir():
        raise HTTPException(status_code=400, detail=f"Workspace root does not exist: {root}")
    if not _is_allowed_workspace_root(root):
        raise HTTPException(
            status_code=403,
            detail=(
                "Workspace root is not approved. "
                f"Allowed roots: {', '.join(str(x) for x in ALLOWED_LOCAL_WORKSPACES)}"
            ),
        )
    return LocalWorkspaceContext(root=root)


def _count_workbooks_under(
    root: Path,
    *,
    max_workbooks: int | None = None,
    max_file_names_seen: int | None = None,
) -> int:
    """
    Count Excel workbooks under root. The primary cap limits Excel matches (not rglob path entries),
    so trees with many non-Excel files cannot yield a false zero. Stops when either cap is hit.
    """
    if not root.exists() or not root.is_dir():
        return 0
    cap_wb = _COUNT_WORKBOOKS_MAX if max_workbooks is None else max(0, int(max_workbooks))
    cap_files = _COUNT_WORKBOOKS_WALK_MAX_FILES if max_file_names_seen is None else max(0, int(max_file_names_seen))
    n_wb = 0
    seen = 0
    for _dirpath, _dirnames, filenames in os.walk(root, topdown=True, followlinks=False):
        for name in filenames:
            seen += 1
            if seen > cap_files:
                return n_wb
            if Path(name).suffix.lower() not in EXCEL_SUFFIXES:
                continue
            n_wb += 1
            if n_wb >= cap_wb:
                return n_wb
    return n_wb


def _readiness_snapshot(path: Path) -> dict[str, Any]:
    exists = path.exists() and path.is_dir()
    return {
        "exists": exists,
        "allowed": _is_allowed_workspace_root(path) if exists else False,
        "workbook_count": _count_workbooks_under(path) if exists else 0,
    }


def _open_workspace_for_assistant(
    requested_root: str | None,
) -> tuple[LocalWorkspaceContext | None, int, dict[str, Any] | None]:
    """
    Open workspace for the assistant, never raising HTTPException.
    Returns (context, workbook_count, error_fragment). If error_fragment is set, context is None.
    """
    raw = (str(requested_root).strip() if requested_root is not None else "")
    if not raw:
        raw = str(_load_workspace_config().get("default_workspace_root") or "").strip()
    root = Path(raw).resolve() if raw else DEFAULT_LOCAL_WORKSPACE
    trace = [{"action": "preflight.workspace", "path": str(root), "configured_explicit": bool(raw)}]
    if not root.exists() or not root.is_dir():
        return (
            None,
            0,
            {
                "status": "needs_setup",
                "message": f"This workspace path is missing or not a folder: {root}",
                "readiness": {"exists": False, "allowed": False, "workbook_count": 0, "resolved_root": str(root)},
                "trace": trace
                + [{"action": "preflight.workspace", "outcome": "missing_or_not_directory"}],
            },
        )
    if not _is_allowed_workspace_root(root):
        return (
            None,
            0,
            {
                "status": "needs_setup",
                "message": "This Operator Local instance is not allowed to use that path as a workspace. Pick a path under the approved project folders.",
                "readiness": {
                    "exists": True,
                    "allowed": False,
                    "workbook_count": 0,
                    "resolved_root": str(root),
                },
                "trace": trace + [{"action": "preflight.workspace", "outcome": "forbidden"}],
            },
        )
    wb = _count_workbooks_under(root)
    trace.append({"action": "preflight.workspace", "outcome": "ok", "workbook_count": wb})
    return LocalWorkspaceContext(root=root), wb, None


def _run_has_operator_outputs(run_id: str) -> bool:
    rid = (run_id or "").strip()
    if not rid or not re.match(r"^[A-Za-z0-9._-]+$", rid):
        return False
    return (RUNS_ROOT / rid / "outputs" / "operator_envelope.json").is_file()


def _safe_workspace_relative_path(workspace_root: Path, raw_path: str) -> tuple[Path, str]:
    candidate = Path(raw_path)
    target = (workspace_root / candidate).resolve() if not candidate.is_absolute() else candidate.resolve()
    if workspace_root not in target.parents and target != workspace_root:
        raise HTTPException(status_code=400, detail="Path escapes current workspace root.")
    rel = str(target.relative_to(workspace_root))
    return target, rel


def _list_workspace_files(workspace_root: Path, query: str = "", limit: int = 60) -> list[dict[str, Any]]:
    q = query.strip().lower()
    rows: list[dict[str, Any]] = []
    for p in sorted(workspace_root.rglob("*"), key=lambda x: x.stat().st_mtime, reverse=True):
        if not p.is_file():
            continue
        rel = str(p.relative_to(workspace_root))
        if q and q not in rel.lower():
            continue
        stat = p.stat()
        rows.append(
            {
                "path": rel,
                "name": p.name,
                "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
                .replace(microsecond=0)
                .isoformat()
                .replace("+00:00", "Z"),
            }
        )
        if len(rows) >= max(1, min(limit, 200)):
            break
    return rows


def _read_workspace_file(workspace_root: Path, raw_path: str) -> dict[str, Any]:
    target, rel = _safe_workspace_relative_path(workspace_root, raw_path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {raw_path}")
    suffix = target.suffix.lower()
    if suffix not in READABLE_LOCAL_SUFFIXES:
        raise HTTPException(status_code=400, detail=f"File type is not approved for read: {suffix or '(none)'}")
    if target.stat().st_size > MAX_LOCAL_FILE_READ_BYTES:
        raise HTTPException(
            status_code=400,
            detail=(
                f"File too large for inline read ({target.stat().st_size} bytes). "
                f"Limit is {MAX_LOCAL_FILE_READ_BYTES} bytes."
            ),
        )
    return {
        "path": rel,
        "content": target.read_text(encoding="utf-8", errors="replace"),
        "size_bytes": target.stat().st_size,
    }


def _file_row(workspace_root: Path, file_path: Path) -> dict[str, Any]:
    stat = file_path.stat()
    return {
        "path": str(file_path.relative_to(workspace_root)),
        "name": file_path.name,
        "size_bytes": stat.st_size,
        "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z"),
    }


def _canonical_report_family(name: str) -> str:
    x = name.lower()
    x = re.sub(r"\.(xlsx|xlsm|xltx|xltm|md|txt|csv|json)$", "", x)
    x = re.sub(r"[^a-z0-9]+", " ", x)
    x = re.sub(r"\b(20\d{2}[-_./]?\d{1,2}[-_./]?\d{1,2}|\d{1,2}[-_./]\d{1,2}[-_./]20\d{2})\b", " ", x)
    x = re.sub(r"\b(v?\d+|prior|current|before|after|snapshot|update)\b", " ", x)
    x = re.sub(r"[^a-z0-9]+", " ", x).strip()
    return re.sub(r"\s+", " ", x)


def _score_workbook_candidate(file_path: Path, workspace_root: Path, query: str) -> float:
    rel = str(file_path.relative_to(workspace_root)).lower()
    name = file_path.name.lower()
    ext = file_path.suffix.lower()
    score = 0.0
    if ext in EXCEL_SUFFIXES:
        score += 120.0
    elif ext == ".md":
        score += 45.0
    elif ext in {".csv", ".json"}:
        score += 30.0
    if any(k in name for k in ("profit", "p&l", "financial", "income")):
        score += 30.0
    if "report" in name:
        score += 20.0
    if "update" in name:
        score += 10.0
    if re.search(r"\d{6,}", name):
        score += 8.0
    if re.search(r"(20\d{2}[-_./]?\d{1,2}[-_./]?\d{1,2}|\d{1,2}[-_./]\d{1,2}[-_./]20\d{2})", name):
        score += 8.0
    stop_tokens = {
        "find",
        "latest",
        "prior",
        "previous",
        "version",
        "compare",
        "show",
        "deltas",
        "delta",
        "report",
        "workbook",
        "files",
        "relevant",
        "tell",
        "whether",
        "cost",
        "revenue",
    }
    query_tokens = [t for t in re.findall(r"[a-z0-9]+", query.lower()) if len(t) >= 3 and t not in stop_tokens]
    for tok in query_tokens:
        if tok in rel:
            score += 3.0
    score += min(file_path.stat().st_mtime / 1_000_000_000.0, 5.0)
    return score


def _ranked_workbook_candidates(workspace_root: Path, query: str = "", limit: int = 20) -> list[dict[str, Any]]:
    candidates: list[tuple[float, float, Path]] = []
    for p in workspace_root.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() not in EXCEL_SUFFIXES | {".md", ".csv", ".json"}:
            continue
        rel = str(p.relative_to(workspace_root)).lower()
        if not any(k in rel for k in ("profit", "report", "financial", "income", "p&l", "snapshot", "workbook")):
            continue
        score = _score_workbook_candidate(p, workspace_root, query)
        candidates.append((score, p.stat().st_mtime, p))
    ranked = sorted(candidates, key=lambda row: (row[0], row[1]), reverse=True)
    out: list[dict[str, Any]] = []
    for score, _, p in ranked[: max(1, min(limit, 60))]:
        row = _file_row(workspace_root, p)
        row["rank_score"] = round(score, 2)
        out.append(row)
    return out


def _latest_profit_reports(workspace_root: Path, query: str = "", limit: int = 8) -> list[dict[str, Any]]:
    ranked = _ranked_workbook_candidates(workspace_root, query=query, limit=max(limit * 3, 12))
    return ranked[:limit]


def _recent_workbook_files(workspace_root: Path, query: str = "", limit: int = 12) -> list[dict[str, Any]]:
    return _ranked_workbook_candidates(workspace_root, query=query, limit=limit)


def _find_prior_version(workspace_root: Path, base_path: str, preferred_ext: str | None = None) -> dict[str, Any] | None:
    target, rel = _safe_workspace_relative_path(workspace_root, base_path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail=f"Base file not found: {base_path}")
    target_stat = target.stat()
    stem = target.stem.lower()
    suffix = target.suffix.lower()
    family = _canonical_report_family(target.name)
    candidates: list[Path] = []
    for p in workspace_root.rglob("*"):
        if not p.is_file() or p.resolve() == target:
            continue
        if preferred_ext and p.suffix.lower() != preferred_ext:
            continue
        if p.suffix.lower() != suffix:
            continue
        candidate_family = _canonical_report_family(p.name)
        if family and candidate_family and family != candidate_family and stem not in p.stem.lower():
            continue
        if p.stat().st_mtime >= target_stat.st_mtime:
            continue
        candidates.append(p)
    if not candidates:
        return None
    prior = sorted(candidates, key=lambda x: x.stat().st_mtime, reverse=True)[0]
    row = _file_row(workspace_root, prior)
    row["for_path"] = rel
    return row


def _workbook_metadata(workspace_root: Path, workbook_path: str) -> dict[str, Any]:
    target, rel = _safe_workspace_relative_path(workspace_root, workbook_path)
    if target.suffix.lower() not in EXCEL_SUFFIXES:
        raise HTTPException(status_code=400, detail="Workbook inspection supports .xlsx/.xlsm/.xltx/.xltm only.")
    wb = load_workbook(filename=str(target), data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        sheets.append(
            {
                "name": ws.title,
                "max_row": int(ws.max_row or 0),
                "max_column": int(ws.max_column or 0),
            }
        )
    return {
        "path": rel,
        "name": target.name,
        "size_bytes": target.stat().st_size,
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def _find_report_sheets(workspace_root: Path, workbook_path: str) -> dict[str, Any]:
    target, rel = _safe_workspace_relative_path(workspace_root, workbook_path)
    if target.suffix.lower() not in EXCEL_SUFFIXES:
        raise HTTPException(status_code=400, detail="Report-sheet detection supports workbook files only.")
    wb = load_workbook(filename=str(target), data_only=True, read_only=True)
    likely: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        name_l = ws.title.lower()
        score = 0
        reasons: list[str] = []
        if any(k in name_l for k in ("profit", "p&l", "income", "summary", "financial")):
            score += 5
            reasons.append("sheet-name keyword")
        sampled_values: list[str] = []
        rows = ws.iter_rows(min_row=1, max_row=WORKBOOK_HEADER_SCAN_ROWS, min_col=1, max_col=8, values_only=True)
        for row in rows:
            sampled_values.extend([str(c).lower() for c in row if c is not None])
        joined = " ".join(sampled_values)
        if any(k in joined for k in ("revenue", "cost", "profit", "gross", "operating")):
            score += 4
            reasons.append("header keyword match")
        if score > 0:
            likely.append({"sheet_name": ws.title, "score": score, "reasons": reasons})
    likely_sorted = sorted(likely, key=lambda x: x["score"], reverse=True)
    return {"path": rel, "report_sheets": likely_sorted}


def _preview_workbook_sheet(
    workspace_root: Path,
    workbook_path: str,
    sheet_name: str | None = None,
    max_rows: int = WORKBOOK_PREVIEW_MAX_ROWS,
    max_cols: int = WORKBOOK_PREVIEW_MAX_COLS,
) -> dict[str, Any]:
    target, rel = _safe_workspace_relative_path(workspace_root, workbook_path)
    if target.suffix.lower() not in EXCEL_SUFFIXES:
        raise HTTPException(status_code=400, detail="Workbook preview supports workbook files only.")
    wb = load_workbook(filename=str(target), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    rows_out: list[list[str]] = []
    rows = ws.iter_rows(
        min_row=1,
        max_row=max(1, min(max_rows, WORKBOOK_PREVIEW_MAX_ROWS)),
        min_col=1,
        max_col=max(1, min(max_cols, WORKBOOK_PREVIEW_MAX_COLS)),
        values_only=True,
    )
    for row in rows:
        row_vals = ["" if cell is None else str(cell) for cell in row]
        if any(v.strip() for v in row_vals):
            rows_out.append(row_vals)
    return {
        "path": rel,
        "sheet_name": ws.title,
        "preview_rows": rows_out,
        "preview_row_count": len(rows_out),
        "max_rows": max_rows,
        "max_cols": max_cols,
    }


def _select_latest_prior_workbooks(workspace_root: Path, query: str = "") -> dict[str, Any]:
    candidates = _ranked_workbook_candidates(workspace_root, query=query, limit=120)
    workbook_candidates = [c for c in candidates if Path(str(c["path"])).suffix.lower() in EXCEL_SUFFIXES]

    def role_for(name: str) -> str:
        n = name.lower()
        if "prior" in n or "before" in n:
            return "prior"
        if "current" in n or "after" in n:
            return "current"
        return "unknown"

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in workbook_candidates:
        family = _canonical_report_family(str(row["name"])) or str(row["name"]).lower()
        grouped.setdefault(family, []).append(row)
    best_pair: tuple[dict[str, Any], dict[str, Any], float] | None = None
    best_family = ""
    for family, rows in grouped.items():
        rows_sorted = sorted(rows, key=lambda r: (str(r["modified_at"]), float(r.get("rank_score", 0))), reverse=True)
        if len(rows_sorted) < 2:
            continue
        current_rows = [r for r in rows_sorted if role_for(str(r["name"])) == "current"]
        prior_rows = [r for r in rows_sorted if role_for(str(r["name"])) == "prior"]
        latest_candidate = current_rows[0] if current_rows else rows_sorted[0]
        prior_candidate = None
        # Prefer a directory sibling with prior/current counterpart naming.
        latest_path = str(latest_candidate["path"])
        latest_name = str(latest_candidate["name"])
        latest_dir = str(Path(latest_path).parent)
        if "current" in latest_name.lower():
            expected_prior = latest_name.lower().replace("current", "prior", 1)
            for row in rows_sorted:
                if str(Path(str(row["path"])).parent) == latest_dir and str(row["name"]).lower() == expected_prior:
                    prior_candidate = row
                    break
        if prior_rows:
            if not prior_candidate:
                prior_candidate = prior_rows[0]
        else:
            for row in rows_sorted[1:]:
                if row["path"] != latest_candidate["path"]:
                    prior_candidate = row
                    break
        if not prior_candidate:
            continue
        pair_score = float(latest_candidate.get("rank_score", 0)) + float(prior_candidate.get("rank_score", 0))
        if not best_pair or pair_score > best_pair[2]:
            best_pair = (latest_candidate, prior_candidate, pair_score)
            best_family = family
    if not best_pair and len(workbook_candidates) >= 2:
        best_pair = (workbook_candidates[0], workbook_candidates[1], float(workbook_candidates[0].get("rank_score", 0)))
        best_family = "fallback-top-ranked"
    if not best_pair:
        return {
            "latest": None,
            "prior": None,
            "family": "",
            "candidates": workbook_candidates[:10],
            "selection_reason": "Insufficient workbook candidates.",
        }
    latest, prior, _ = best_pair
    return {
        "latest": latest,
        "prior": prior,
        "family": best_family,
        "candidates": workbook_candidates[:10],
        "selection_reason": (
            "Selected highest-ranked workbook family and used latest two timestamps."
            if best_family != "fallback-top-ranked"
            else "Selected top-ranked workbooks by score and recency."
        ),
    }


def _source_type_for_path(rel_path: str) -> str:
    return "upload" if "runtime/operator_ui/sessions/" in rel_path.replace("\\", "/") else "workspace"


def _extract_project_id(rel_path: str, file_name: str) -> str:
    return extract_project_id_from_rel_and_name(rel_path, file_name)


def _extract_version_date(raw: str) -> tuple[str, int]:
    s = raw.lower()
    m = re.search(r"(?<!\d)(20\d{2})[-_./](\d{1,2})[-_./](\d{1,2})(?!\d)", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}", y * 10_000 + mo * 100 + d
    m = re.search(r"(?<!\d)(\d{1,2})[-_./](\d{1,2})[-_./](20\d{2})(?!\d)", s)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}", y * 10_000 + mo * 100 + d
    return "", 0


def _pairing_overrides(query: str, context: dict[str, Any]) -> dict[str, Any]:
    q = query.lower()
    selected_files = list(context.get("selected_files") or [])
    use_selected = "use these files instead" in q and len(selected_files) >= 2
    workspace_only = ("ignore uploads" in q) or ("use workspace files only" in q)
    return {
        "use_selected_files": use_selected,
        "workspace_only": workspace_only,
        "selected_pair_id": str(context.get("selected_pair_id") or "").strip(),
        "last_confirmed_pair": context.get("last_confirmed_pair")
        if isinstance(context.get("last_confirmed_pair"), dict)
        else {},
        "preferred_report_family": str(context.get("preferred_report_family") or "").strip().lower(),
    }


def _candidate_models_from_rglob(
    workspace_root: Path,
    query: str,
    overrides: dict[str, Any],
    context: dict[str, Any],
) -> list[dict[str, Any]]:
    rows = _ranked_workbook_candidates(workspace_root, query=query, limit=180)
    out: list[dict[str, Any]] = []
    for row in rows:
        rel = str(row.get("path", ""))
        p = Path(rel)
        if p.suffix.lower() not in EXCEL_SUFFIXES:
            continue
        name = str(row.get("name", p.name))
        family = _canonical_report_family(name)
        version_date, version_sort = _extract_version_date(name)
        rank_score = float(row.get("rank_score", 0) or 0)
        base_conf = min(1.0, rank_score / 220.0)
        if family:
            base_conf += 0.08
        if version_sort:
            base_conf += 0.07
        project_id = _extract_project_id(rel, name)
        if project_id:
            base_conf += 0.05
        model = {
            "path": rel,
            "name": name,
            "modified_at": row.get("modified_at", ""),
            "rank_score": rank_score,
            "project_id": project_id,
            "report_family": family,
            "version_date": version_date,
            "version_sort": version_sort,
            "source_type": _source_type_for_path(rel),
            "confidence": round(min(base_conf, 1.0), 3),
        }
        out.append(model)
    return out


def _build_workbook_candidate_models(
    workspace_root: Path,
    query: str,
    overrides: dict[str, Any],
    context: dict[str, Any],
) -> list[dict[str, Any]]:
    pfilter = _extract_project_id_from_query(query) or str(
        (context or {}).get("project_id_filter") or ""
    ).strip()
    root_s = str(workspace_root.resolve())
    index_nonempty = _index_workbook_count_for_root(root_s) > 0
    out: list[dict[str, Any]] = []
    if index_nonempty:
        out = _index_candidate_models(workspace_root, query, pfilter)
        if pfilter and not out:
            # Do not fall back to a full-workspace rglob when the index is authoritative
            # but has no rows for this project.
            return []
    if not out:
        out = _candidate_models_from_rglob(workspace_root, query, overrides, context)
        if pfilter in (UNLABELED_GROUP_ID, "(unlabeled)"):
            f2 = [m for m in out if not str(m.get("project_id") or "").strip()]
        elif pfilter:
            f2 = [m for m in out if str(m.get("project_id") or "") == pfilter]
        else:
            f2 = out
        out = f2
    if overrides.get("workspace_only"):
        out = [x for x in out if x.get("source_type") == "workspace"]
    if overrides.get("use_selected_files"):
        selected = []
        for raw in list(context.get("selected_files") or [])[:2]:
            try:
                _, rel = _safe_workspace_relative_path(workspace_root, str(raw))
            except HTTPException:
                continue
            match = next((x for x in out if x["path"] == rel), None)
            if match:
                selected.append(match)
        if len(selected) == 2:
            return selected
    return out


def _candidate_sort_key(candidate: dict[str, Any]) -> tuple[int, str, float]:
    return (
        int(candidate.get("version_sort", 0) or 0),
        str(candidate.get("modified_at", "")),
        float(candidate.get("rank_score", 0) or 0),
    )


def _build_pair_options(
    workspace_root: Path,
    query: str,
    context: dict[str, Any],
    *,
    top_n: int = 3,
) -> dict[str, Any]:
    overrides = _pairing_overrides(query, context)
    pfilter_resolved = _extract_project_id_from_query(query) or str(
        (context or {}).get("project_id_filter") or ""
    ).strip()
    candidates = _build_workbook_candidate_models(workspace_root, query, overrides, context)
    pair_options: list[dict[str, Any]] = []

    if overrides.get("use_selected_files") and len(candidates) == 2:
        current, prior = candidates[0], candidates[1]
        if _candidate_sort_key(prior) > _candidate_sort_key(current):
            current, prior = prior, current
        pair_options.append(
            {
                "pair_id": f"{prior['path']}|{current['path']}",
                "current": current,
                "prior": prior,
                "pair_confidence": 1.0,
                "ranking_factors": [
                    "operator override: use selected files",
                    "pair chosen directly from selected file inputs",
                ],
                "selection_reason": "Operator explicitly requested selected files for compare.",
                "score": 1.0,
            }
        )
    else:
        grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for c in candidates:
            key = (str(c.get("project_id") or "unknown"), str(c.get("report_family") or "unknown"))
            grouped.setdefault(key, []).append(c)
        for (project_id, family), rows in grouped.items():
            rows_sorted = sorted(rows, key=_candidate_sort_key, reverse=True)
            if len(rows_sorted) < 2:
                continue
            for idx in range(len(rows_sorted) - 1):
                current = rows_sorted[idx]
                prior = rows_sorted[idx + 1]
                if current["path"] == prior["path"]:
                    continue
                factors: list[str] = []
                score = 0.0
                if family != "unknown":
                    score += 0.36
                    factors.append("report family match")
                else:
                    score += 0.12
                    factors.append("weak family match (fallback)")
                if project_id != "unknown" and current.get("project_id") == prior.get("project_id"):
                    score += 0.2
                    factors.append("project id match")
                else:
                    score += 0.05
                    factors.append("project id missing or inconsistent")
                c_date, p_date = int(current.get("version_sort", 0) or 0), int(prior.get("version_sort", 0) or 0)
                if c_date and p_date:
                    diff_days = abs(c_date - p_date)
                    prox = max(0.04, 0.24 * (1.0 / (1.0 + (diff_days / 100.0))))
                    score += prox
                    factors.append(f"version-date proximity ({diff_days} date-distance)")
                else:
                    score += 0.08
                    factors.append("date proximity unavailable (modified timestamp fallback)")
                source_types = {str(current.get("source_type")), str(prior.get("source_type"))}
                if source_types == {"workspace"}:
                    score += 0.12
                    factors.append("source preference: workspace pair")
                elif "workspace" in source_types:
                    score += 0.08
                    factors.append("source preference: mixed workspace/upload")
                else:
                    score += 0.04
                    factors.append("source preference: upload pair")
                score += (float(current.get("confidence", 0)) + float(prior.get("confidence", 0))) * 0.06
                preferred_family = overrides.get("preferred_report_family")
                if preferred_family and preferred_family == str(current.get("report_family", "")):
                    score += 0.07
                    factors.append("session bias: preferred report family")
                last_pair = overrides.get("last_confirmed_pair") or {}
                if isinstance(last_pair, dict) and (
                    str(last_pair.get("prior_path", "")) == current["path"]
                    or str(last_pair.get("current_path", "")) == current["path"]
                    or str(last_pair.get("prior_path", "")) == prior["path"]
                    or str(last_pair.get("current_path", "")) == prior["path"]
                ):
                    score += 0.06
                    factors.append("session bias: aligns with last confirmed pair")
                pair_options.append(
                    {
                        "pair_id": f"{prior['path']}|{current['path']}",
                        "current": current,
                        "prior": prior,
                        "pair_confidence": round(min(score, 1.0), 3),
                        "ranking_factors": factors,
                        "selection_reason": "; ".join(factors[:3]),
                        "score": round(score, 4),
                    }
                )

    pair_options = sorted(
        pair_options,
        key=lambda p: (
            float(p.get("score", 0)),
            str((p.get("current") or {}).get("modified_at", "")),
            str((p.get("prior") or {}).get("modified_at", "")),
        ),
        reverse=True,
    )
    top_options = pair_options[: max(1, min(top_n, 5))]
    viable_count = len([p for p in top_options if float(p.get("pair_confidence", 0)) >= 0.58])
    selected_pair = top_options[0] if top_options else None
    forced_pair_id = str(overrides.get("selected_pair_id") or "")
    if forced_pair_id:
        forced = next((p for p in top_options if str(p.get("pair_id")) == forced_pair_id), None)
        if forced:
            selected_pair = forced
            selected_pair["selection_reason"] = (
                f"Operator selected this pair option. {selected_pair.get('selection_reason', '')}"
            ).strip()
    selected_conf = float(selected_pair.get("pair_confidence", 0)) if isinstance(selected_pair, dict) else 0.0
    disambiguation_required = (viable_count > 1) or (selected_conf < PAIRING_LOW_CONFIDENCE_THRESHOLD)
    requires_operator_selection = selected_conf < PAIRING_LOW_CONFIDENCE_THRESHOLD
    return {
        "pair_candidates": top_options,
        "selected_pair": selected_pair,
        "pairing_confidence": round(selected_conf, 3),
        "disambiguation_required": disambiguation_required,
        "requires_operator_selection": requires_operator_selection,
        "filtered_candidate_count": len(candidates),
        "project_filter": pfilter_resolved,
        "overrides_applied": {
            "use_selected_files": bool(overrides.get("use_selected_files")),
            "workspace_only": bool(overrides.get("workspace_only")),
            "preferred_report_family": str(overrides.get("preferred_report_family") or ""),
        },
    }


def _artifacts_for_run(run_id: str) -> list[str]:
    state = _load_state(run_id)
    if not state.run_dir:
        return []
    return _artifact_relative_paths(Path(state.run_dir))


def _run_cost_vs_revenue_signal(run_id: str) -> dict[str, Any]:
    state = _load_state(run_id)
    if state.status != "completed" or not isinstance(state.structured_output, dict):
        raise HTTPException(status_code=400, detail=f"Run is not a completed structured run: {run_id}")
    sd = state.structured_output.get("summary_deltas", {}) if isinstance(state.structured_output, dict) else {}
    revenue = float(sd.get("revenue", 0) or 0)
    cost = float(sd.get("cost", 0) or 0)
    if revenue == 0 and cost == 0:
        rows = state.structured_output.get("material_diff_items", [])
        if isinstance(rows, list):
            for row in rows:
                if not isinstance(row, dict):
                    continue
                label = str(row.get("category_label") or row.get("category") or "").lower()
                delta = float(row.get("delta", 0) or 0)
                if "revenue" in label or "sales" in label or "income" in label:
                    revenue += delta
                if "cost" in label or "expense" in label or "labor" in label:
                    cost += delta
    signal = "balanced"
    if abs(cost) > abs(revenue):
        signal = "cost-led"
    elif abs(revenue) > abs(cost):
        signal = "revenue-led"
    return {
        "run_id": run_id,
        "signal": signal,
        "revenue_delta": revenue,
        "cost_delta": cost,
        "profit_delta": float(sd.get("profit", 0) or 0),
    }


def _derive_labor_deltas(run_id: str, structured_output: dict[str, Any], limit: int = 8) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key in ("material_diff_items", "material_diff_audit_items"):
        for row in structured_output.get(key, []) if isinstance(structured_output, dict) else []:
            if not isinstance(row, dict):
                continue
            label = str(row.get("category_label") or row.get("category") or "")
            if "labor" in label.lower() or "| lab" in label.lower():
                rows.append(
                    {
                        "category_label": label,
                        "prior_value": float(row.get("prior_value", 0) or 0),
                        "current_value": float(row.get("current_value", 0) or 0),
                        "delta": float(row.get("delta", 0) or 0),
                        "tier": str(row.get("tier", "audit")),
                    }
                )
    if rows:
        return sorted(rows, key=lambda r: abs(float(r.get("delta", 0))), reverse=True)[:limit]
    state = _load_state(run_id)
    if not state.run_dir:
        return []
    run_dir = Path(state.run_dir)
    prior_snapshot = run_dir / "inputs" / "prior_financial_snapshot.json"
    current_snapshot = run_dir / "inputs" / "current_financial_snapshot.json"
    if not prior_snapshot.exists() or not current_snapshot.exists():
        return []
    prior_data = _load_json(prior_snapshot)
    current_data = _load_json(current_snapshot)
    prior_map = {
        str(row.get("name", "")): float(row.get("value", 0) or 0)
        for row in prior_data.get("categories", [])
        if isinstance(row, dict)
    }
    current_map = {
        str(row.get("name", "")): float(row.get("value", 0) or 0)
        for row in current_data.get("categories", [])
        if isinstance(row, dict)
    }
    all_keys = sorted(set(prior_map.keys()) | set(current_map.keys()))
    out: list[dict[str, Any]] = []
    for key in all_keys:
        lk = key.lower()
        if "labor" not in lk and "| lab" not in lk and " lab" not in lk:
            continue
        prior_val = float(prior_map.get(key, 0))
        current_val = float(current_map.get(key, 0))
        out.append(
            {
                "category_label": key,
                "prior_value": prior_val,
                "current_value": current_val,
                "delta": current_val - prior_val,
                "tier": "audit",
            }
        )
    return sorted(out, key=lambda r: abs(float(r.get("delta", 0))), reverse=True)[:limit]


def _workflow_phase(workflow_name: str) -> str:
    registry = _load_json(REGISTRY_PATH)
    for wf in registry.get("workflows", []):
        if wf.get("name") == workflow_name:
            return str(wf.get("phase", "Phase 5"))
    raise HTTPException(status_code=400, detail=f"Workflow missing from registry: {workflow_name}")


def _run_checked(command: list[str]) -> None:
    if os.environ.get("OPERATOR_UI_RUN_SCRIPTS_IN_PROCESS") == "1" and len(command) >= 2:
        _run_script_in_process(command)
        return
    proc = subprocess.run(command, cwd=str(REPO_ROOT), capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        detail = "\n".join(
            [
                f"command failed ({proc.returncode}): {' '.join(command)}",
                "--- stdout ---",
                proc.stdout.rstrip(),
                "--- stderr ---",
                proc.stderr.rstrip(),
            ]
        )
        raise RuntimeError(detail)


def _run_script_in_process(command: list[str]) -> None:
    script_path = Path(command[1]).resolve()
    if script_path.parent != SCRIPTS_DIR.resolve() or not script_path.is_file():
        raise RuntimeError(f"Refusing to run non-operator script in-process: {script_path}")

    old_argv = sys.argv[:]
    old_cwd = Path.cwd()
    stdout = io.StringIO()
    stderr = io.StringIO()
    exit_code = 0
    try:
        sys.argv = [str(script_path), *command[2:]]
        os.chdir(REPO_ROOT)
        with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
            try:
                runpy.run_path(str(script_path), run_name="__main__")
            except SystemExit as exc:
                code = exc.code
                if isinstance(code, int):
                    exit_code = code
                elif code:
                    exit_code = 1
    finally:
        os.chdir(old_cwd)
        sys.argv = old_argv

    if exit_code != 0:
        detail = "\n".join(
            [
                f"command failed ({exit_code}): {' '.join(command)}",
                "--- stdout ---",
                stdout.getvalue().rstrip(),
                "--- stderr ---",
                stderr.getvalue().rstrip(),
            ]
        )
        raise RuntimeError(detail)


def _excel_to_markdown(source: Path, target: Path) -> None:
    wb = load_workbook(filename=str(source), data_only=True, read_only=True)
    lines: list[str] = [f"# Excel Intake — {source.name}", ""]
    for ws in wb.worksheets[:6]:
        lines.append(f"## Sheet: {ws.title}")
        lines.append("")
        rows = ws.iter_rows(min_row=1, max_row=MAX_EXCEL_ROWS_PER_SHEET, max_col=MAX_EXCEL_COLS, values_only=True)
        table_rows: list[list[str]] = []
        for row in rows:
            cells = [("" if val is None else str(val).strip()) for val in row]
            if any(cells):
                table_rows.append(cells)
        if not table_rows:
            lines.append("- (no non-empty rows in sampled range)")
            lines.append("")
            continue
        header = table_rows[0]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join("---" for _ in header) + " |")
        for row in table_rows[1:]:
            lines.append("| " + " | ".join(row) + " |")
        lines.append("")
    target.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def _prepare_workflow_inputs(run_id: str, workflow_name: str, uploaded_files: list[dict[str, Any]]) -> list[Path]:
    run_dir = RUNS_ROOT / run_id
    prepared_dir = run_dir / "inputs"
    prepared_dir.mkdir(parents=True, exist_ok=True)
    prepared: list[Path] = []

    def convert_if_needed(src: Path, role: str) -> Path:
        if src.suffix.lower() in EXCEL_SUFFIXES:
            out = prepared_dir / f"{role}_{src.stem}.md"
            _excel_to_markdown(src, out)
            return out
        return src

    role_map: dict[str, Path] = {}
    for item in uploaded_files:
        role = str(item.get("role", ""))
        p = Path(str(item["stored_path"]))
        role_map[role] = p

    if workflow_name in {"wf_compare_markdown", "wf_financial_markdown_delta"}:
        if "prior" not in role_map or "current" not in role_map:
            raise HTTPException(status_code=400, detail="Compare/financial run requires prior and current files.")
        prior = role_map["prior"]
        current = role_map["current"]
        if (
            workflow_name == "wf_financial_markdown_delta"
            and prior.suffix.lower() in EXCEL_SUFFIXES
            and current.suffix.lower() in EXCEL_SUFFIXES
        ):
            prior_snapshot = extract_financial_snapshot_from_workbook(prior)
            current_snapshot = extract_financial_snapshot_from_workbook(current)
            prior_json = prepared_dir / "prior_financial_snapshot.json"
            current_json = prepared_dir / "current_financial_snapshot.json"
            write_snapshot_json(prior_json, prior_snapshot)
            write_snapshot_json(current_json, current_snapshot)
            (prepared_dir / "intake_debug_prior.md").write_text(
                snapshot_to_markdown(prior_snapshot), encoding="utf-8"
            )
            (prepared_dir / "intake_debug_current.md").write_text(
                snapshot_to_markdown(current_snapshot), encoding="utf-8"
            )
            prepared.append(prior_json)
            prepared.append(current_json)
            return prepared
        prepared.append(convert_if_needed(prior, "prior"))
        prepared.append(convert_if_needed(current, "current"))
    elif workflow_name == "wf_extract_risk_lines":
        source = role_map.get("source")
        if not source:
            raise HTTPException(status_code=400, detail="Risk run requires source file.")
        prepared.append(convert_if_needed(source, "source"))
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported workflow: {workflow_name}")
    return prepared


def _detect_workflow(mode: str, files: list[dict[str, Any]]) -> str:
    if mode in WORKFLOW_MAP:
        return WORKFLOW_MAP[mode]
    if len(files) == 1:
        return "wf_extract_risk_lines"
    if len(files) == 2:
        suffixes = {Path(f["original_name"]).suffix.lower() for f in files}
        if suffixes & EXCEL_SUFFIXES:
            return "wf_financial_markdown_delta"
        return "wf_compare_markdown"
    raise HTTPException(status_code=400, detail="Auto-detection supports 1 file (risk) or 2 files (compare/financial).")


def _store_upload(run_id: str, role: str, upload: UploadFile) -> dict[str, Any]:
    session_dir = _session_dir(run_id)
    upload_dir = session_dir / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe = _safe_name(upload.filename or f"{role}.bin")
    target = upload_dir / f"{role}_{safe}"
    size = 0
    with target.open("wb") as out:
        while True:
            chunk = upload.file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_UPLOAD_BYTES:
                out.close()
                target.unlink(missing_ok=True)
                raise HTTPException(status_code=400, detail=f"File too large: {upload.filename}")
            out.write(chunk)
    return {
        "role": role,
        "original_name": upload.filename or safe,
        "stored_path": str(target.resolve()),
        "size_bytes": size,
    }


def _artifact_relative_paths(run_dir: Path) -> list[str]:
    result: list[str] = []
    for sub in ("outputs", "inputs"):
        base = run_dir / sub
        if not base.exists():
            continue
        for p in sorted(base.rglob("*")):
            if p.is_file():
                result.append(str(p.relative_to(run_dir)))
    return result


def _uploaded_file_summary(uploaded_files: list[dict[str, Any]]) -> str:
    names: list[str] = []
    for f in uploaded_files:
        n = f.get("original_name")
        r = f.get("role", "")
        if n:
            names.append(f"{r}: {n}" if r else str(n))
    return " · ".join(names) if names else "(no files description)"


def _assistant_view(
    workflow_name: str | None,
    envelope: dict[str, Any],
    structured: dict[str, Any],
    uploaded_files: list[dict[str, Any]],
) -> dict[str, Any]:
    """Run-aware, evidence-backed strings for the assistant panel (no LLM)."""
    files_line = _uploaded_file_summary(uploaded_files)
    out: dict[str, Any] = {
        "lead": (
            f"Compared: {files_line}. "
            "The notes below are tied to this run’s engine output (deterministic), not a freeform chat model."
        ),
        "what_changed": list(envelope.get("what_i_found", [])),
        "what_to_review": list(envelope.get("what_needs_review", [])),
        "what_i_did": list(envelope.get("what_i_did", [])),
        "extraction_narrative": "",
    }
    notes_out: list[str] = []
    if workflow_name == "wf_financial_markdown_delta":
        ex = structured.get("extraction_confidence") if isinstance(structured, dict) else None
        if isinstance(ex, dict) and ex.get("rollup"):
            rollup = ex.get("rollup")
            p = ex.get("prior") if isinstance(ex.get("prior"), dict) else {}
            c = ex.get("current") if isinstance(ex.get("current"), dict) else {}
            pconf = p.get("confidence", "n/a")
            cconf = c.get("confidence", "n/a")
            ps = len(p.get("sheets_scanned", []) or [])
            cs = len(c.get("sheets_scanned", []) or [])
            pcat = p.get("category_count", 0)
            ccat = c.get("category_count", 0)
            out["extraction_narrative"] = (
                f"Extraction quality is {rollup} overall (prior {pconf}, current {cconf}). "
                f"Sheets read: {ps} (prior) / {cs} (current). "
                f"Category rows available: {pcat} (prior) / {ccat} (current). "
                "Gaps in extraction limit category drivers—treat the rollup accordingly."
            )
            for side in (p, c):
                for n in side.get("notes", []) or []:
                    if isinstance(n, str) and n and n not in notes_out:
                        notes_out.append(n)
        else:
            out["extraction_narrative"] = (
                "Structured Excel snapshots were not used for this run; comparison follows "
                "markdown/unified-diff financial heuristics (no per-category prior/current table)."
            )
    else:
        out["extraction_narrative"] = "—"
    out["extraction_notes"] = notes_out
    return out


def _financial_intake_artifact_entries(run_dir: Path) -> list[dict[str, str]]:
    """Stable labels for Excel structured-intake files (when present)."""
    entries: list[tuple[str, str]] = [
        ("inputs/prior_financial_snapshot.json", "Prior financial snapshot (JSON)"),
        ("inputs/current_financial_snapshot.json", "Current financial snapshot (JSON)"),
        ("inputs/intake_debug_prior.md", "Intake debug (prior)"),
        ("inputs/intake_debug_current.md", "Intake debug (current)"),
    ]
    return [{"path": rel, "label": lbl} for rel, lbl in entries if (run_dir / rel).is_file()]


LOCAL_ACTION_REGISTRY: dict[str, LocalActionSpec] = {
    "workspace.select_root": LocalActionSpec(
        key="workspace.select_root",
        description="Choose approved local workspace root.",
        requires_confirmation=False,
    ),
    "workspace.list_files": LocalActionSpec(
        key="workspace.list_files",
        description="List files in current workspace.",
        requires_confirmation=False,
    ),
    "workspace.search_files": LocalActionSpec(
        key="workspace.search_files",
        description="Search files by path tokens.",
        requires_confirmation=False,
    ),
    "workspace.read_file": LocalActionSpec(
        key="workspace.read_file",
        description="Read approved text-like file types.",
        requires_confirmation=False,
    ),
    "workspace.inspect_workbook": LocalActionSpec(
        key="workspace.inspect_workbook",
        description="Inspect workbook metadata and sheet list (read-only).",
        requires_confirmation=False,
    ),
    "workspace.preview_workbook_sheet": LocalActionSpec(
        key="workspace.preview_workbook_sheet",
        description="Preview bounded workbook sheet cells (read-only).",
        requires_confirmation=False,
    ),
    "workspace.find_report_sheets": LocalActionSpec(
        key="workspace.find_report_sheets",
        description="Detect likely report sheets in workbook (read-only).",
        requires_confirmation=False,
    ),
    "run.list_artifacts": LocalActionSpec(
        key="run.list_artifacts",
        description="List artifacts for a completed run.",
        requires_confirmation=False,
    ),
    "run.cost_vs_revenue": LocalActionSpec(
        key="run.cost_vs_revenue",
        description="Summarize whether current run is revenue-led or cost-led.",
        requires_confirmation=False,
    ),
    "compare.invoke_local": LocalActionSpec(
        key="compare.invoke_local",
        description="Invoke deterministic local compare wrapper.",
        requires_confirmation=True,
    ),
}


def _execute_local_compare(
    workspace_root: Path,
    *,
    prior_path: str,
    current_path: str,
    workflow_mode: str,
) -> dict[str, Any]:
    prior_target, prior_rel = _safe_workspace_relative_path(workspace_root, prior_path)
    current_target, current_rel = _safe_workspace_relative_path(workspace_root, current_path)
    if not prior_target.exists() or not current_target.exists():
        raise HTTPException(status_code=404, detail="Both selected files must exist in workspace.")
    selected_mode = workflow_mode
    selection_reason = "Explicit workflow mode from task payload."
    if workflow_mode in {"auto", "compare"} and (
        prior_target.suffix.lower() in EXCEL_SUFFIXES and current_target.suffix.lower() in EXCEL_SUFFIXES
    ):
        selected_mode = "financial"
        selection_reason = "Both inputs are workbooks; selected structured financial compare path."
    run_id = _run_id()
    state = SessionState(
        run_id=run_id,
        status="created",
        workflow_mode=selected_mode,
        workflow_name=None,
        uploaded_files=[
            {
                "role": "prior",
                "original_name": prior_target.name,
                "stored_path": str(prior_target),
                "size_bytes": prior_target.stat().st_size,
            },
            {
                "role": "current",
                "original_name": current_target.name,
                "stored_path": str(current_target),
                "size_bytes": current_target.stat().st_size,
            },
        ],
        created_at=_utc_now_iso(),
    )
    _save_state(state)
    executed = execute_run(run_id)
    payload = json.loads(executed.body.decode("utf-8"))
    payload["input_selection"] = {
        "prior": prior_rel,
        "current": current_rel,
    }
    payload["compare_path"] = payload.get("workflow")
    payload["compare_path_reason"] = selection_reason
    return payload


TASK_CONTRACTS: set[str] = {
    "find_latest_prior_reports",
    "compare_latest_report",
    "compare_latest_prior_reports",
    "compare_and_show_labor_deltas",
    "summarize_for_owner",
    "export_top_changes",
    "run_weekly_review",
    "assess_cost_vs_revenue",
    "list_current_run_artifacts",
    "inspect_workbook",
    "preview_report_sheet",
    "find_report_sheets",
    "scan_workspace",
    "list_projects",
    "show_project_reports",
    "trend_project_reports",
    "compare_multi_reports",
    "generate_financial_signals",
}


def _latest_completed_run_id() -> str | None:
    for row in reversed(_load_history()):
        if str(row.get("status", "")).upper() == "PASS" and row.get("run_id"):
            rid = str(row["run_id"])
            if _run_has_operator_outputs(rid):
                return rid
    return None


def _build_compare_plan(
    workspace_root: Path,
    *,
    prior_path: str,
    current_path: str,
    requested_mode: str,
    selection_reason: str,
) -> dict[str, Any]:
    prior_target, prior_rel = _safe_workspace_relative_path(workspace_root, prior_path)
    current_target, current_rel = _safe_workspace_relative_path(workspace_root, current_path)
    selected_mode = requested_mode
    mode_reason = selection_reason
    if requested_mode in {"auto", "compare"} and (
        prior_target.suffix.lower() in EXCEL_SUFFIXES and current_target.suffix.lower() in EXCEL_SUFFIXES
    ):
        selected_mode = "financial"
        mode_reason = "Both files are workbook inputs, so structured financial compare is selected."
    compare_path = WORKFLOW_MAP.get(selected_mode, "wf_financial_markdown_delta")
    expected_artifacts = [
        "outputs/operator_envelope.json",
        "outputs/structured_output.json",
    ]
    if compare_path == "wf_financial_markdown_delta":
        expected_artifacts.extend(
            [
                "inputs/prior_financial_snapshot.json",
                "inputs/current_financial_snapshot.json",
                "inputs/intake_debug_prior.md",
                "inputs/intake_debug_current.md",
            ]
        )
    return {
        "prior_path": prior_rel,
        "current_path": current_rel,
        "requested_mode": requested_mode,
        "selected_mode": selected_mode,
        "compare_path": compare_path,
        "compare_path_reason": mode_reason,
        "expected_artifacts": expected_artifacts,
    }


def _confidence_rollup(structured_output: dict[str, Any]) -> str:
    ex = structured_output.get("extraction_confidence", {}) if isinstance(structured_output, dict) else {}
    if isinstance(ex, dict) and ex.get("rollup"):
        return str(ex.get("rollup"))
    return "unknown"


def _format_usd_signed(value: float | int | None) -> str:
    n = float(value or 0)
    sign = "+" if n >= 0 else "-"
    return f"{sign}${abs(n):,.0f}"


def _driver_dedup_key(row: dict[str, Any]) -> str:
    return "|".join(
        [
            str(row.get("category_label", "")),
            str(row.get("category", "")),
            str(row.get("prior_value", "")),
            str(row.get("current_value", "")),
            str(row.get("delta", "")),
        ]
    )


def _driver_delta_abs(row: dict[str, Any]) -> float:
    try:
        return abs(float(row.get("delta", 0) or 0))
    except (TypeError, ValueError):
        return 0.0


def merge_and_split_driver_rows(
    structured_output: dict[str, Any] | None,
    *,
    primary_n: int | None = None,
    impact_usd: float | None = None,
) -> dict[str, Any]:
    """Match web UI mergeAndSplitDriverRows: merge primary+audit, dedupe, then primary vs smaller moves."""
    n_top = int(primary_n if primary_n is not None else OPERATOR_UI_DRIVER_TABLE_PRIMARY_N)
    threshold = float(impact_usd if impact_usd is not None else OPERATOR_UI_DRIVER_TABLE_IMPACT_USD)
    so = structured_output if isinstance(structured_output, dict) else {}
    raw: list[dict[str, Any]] = []
    for key in ("material_diff_items", "material_diff_audit_items"):
        for x in so.get(key, []) or []:
            if isinstance(x, dict):
                raw.append(x)
    seen: set[str] = set()
    deduped: list[dict[str, Any]] = []
    for row in raw:
        k = _driver_dedup_key(row)
        if k in seen:
            continue
        seen.add(k)
        deduped.append(row)
    total = len(deduped)
    if total == 0:
        return {
            "primary": [],
            "audit": [],
            "total": 0,
            "primaryCount": 0,
            "auditCount": 0,
        }
    scored = [(r, _driver_delta_abs(r)) for r in deduped]
    scored.sort(key=lambda t: t[1], reverse=True)
    in_primary: set[str] = set()
    for _row, ad in scored[:n_top]:
        in_primary.add(_driver_dedup_key(_row))
    for row, ad in scored:
        if ad >= threshold:
            in_primary.add(_driver_dedup_key(row))
    primary = [r for r in deduped if _driver_dedup_key(r) in in_primary]
    primary.sort(key=_driver_delta_abs, reverse=True)
    audit = [r for r in deduped if _driver_dedup_key(r) not in in_primary]
    audit.sort(key=_driver_delta_abs, reverse=True)
    return {
        "primary": primary,
        "audit": audit,
        "total": total,
        "primaryCount": len(primary),
        "auditCount": len(audit),
    }


def _format_owner_driver_label(row: dict[str, Any]) -> str:
    raw = str(row.get("category_label") or row.get("category") or "").strip()
    raw = re.sub(r"\s+", " ", raw) if raw else ""
    if not raw:
        raw = "Line item"
    low = raw.lower()
    is_labor = "labor" in low or "| lab" in low or re.search(r"\blab\b", low)
    if is_labor and "labor" not in low and "| lab" not in low:
        return f"Labor — {raw}"
    return raw


def _top_driver_rows(structured_output: dict[str, Any], limit: int = 5, labor_first: bool = False) -> list[dict[str, Any]]:
    rows = [x for x in list(structured_output.get("material_diff_items") or []) if isinstance(x, dict)]
    if not rows:
        return []
    if labor_first:
        labor_rows = [r for r in rows if "labor" in str(r.get("category_label") or r.get("category") or "").lower()]
        non_labor = [r for r in rows if r not in labor_rows]
        labor_sorted = sorted(labor_rows, key=lambda r: abs(float(r.get("delta", 0) or 0)), reverse=True)
        non_sorted = sorted(non_labor, key=lambda r: abs(float(r.get("delta", 0) or 0)), reverse=True)
        return (labor_sorted + non_sorted)[:limit]
    return sorted(rows, key=lambda r: abs(float(r.get("delta", 0) or 0)), reverse=True)[:limit]


def _risk_signals_for_run(compare_payload: dict[str, Any]) -> list[str]:
    structured_output = compare_payload.get("structured_output", {}) if isinstance(compare_payload, dict) else {}
    workflow = str(compare_payload.get("workflow") or "")
    risks: list[str] = []
    confidence = _confidence_rollup(structured_output)
    if confidence.lower() == "low":
        risks.append("Extraction confidence is low; treat line-level amounts as directional.")
    line_count = int(structured_output.get("material_diff_line_count", 0) or 0)
    if workflow == "wf_financial_markdown_delta" and line_count == 0:
        risks.append("No material lines crossed the threshold; review audit lines and workbook structure.")
    cost_vs_rev = None
    run_id = str(compare_payload.get("run_id") or "")
    if run_id:
        try:
            cost_vs_rev = _run_cost_vs_revenue_signal(run_id)
        except Exception:  # noqa: BLE001
            cost_vs_rev = None
    if isinstance(cost_vs_rev, dict) and str(cost_vs_rev.get("signal")) == "cost-led":
        risks.append("Cost-led movement is dominant; labor and expense controls need verification.")
    if not risks:
        risks.append("No high-severity deterministic risk was flagged for this run.")
    return risks[:4]


def _preview_link_for_workbook(workspace_root: Path, workbook_path: str, sheet_name: str | None = None) -> str:
    base = f"/api/local/workbook/preview?workspace_root={quote(str(workspace_root))}&path={quote(workbook_path)}"
    if sheet_name:
        base += f"&sheet_name={quote(sheet_name)}"
    return base


def _workbook_preview_links(workspace_root: Path, prior_path: str, current_path: str) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for label, path in (("prior", prior_path), ("current", current_path)):
        top_sheet = ""
        try:
            rs = _find_report_sheets(workspace_root, path)
            items = list(rs.get("report_sheets") or [])
            if items:
                top_sheet = str(items[0].get("sheet_name") or "")
        except Exception:  # noqa: BLE001
            top_sheet = ""
        out.append(
            {
                "label": label,
                "path": path,
                "sheet_name": top_sheet,
                "preview_url": _preview_link_for_workbook(workspace_root, path, top_sheet or None),
                "inspect_url": f"/api/local/workbook/inspect?workspace_root={quote(str(workspace_root))}&path={quote(path)}",
            }
        )
    return out


def _operator_compare_output(compare_payload: dict[str, Any], workspace_root: Path, plan: dict[str, Any]) -> dict[str, Any]:
    run_id = str(compare_payload.get("run_id") or "")
    structured_output = compare_payload.get("structured_output", {}) if isinstance(compare_payload, dict) else {}
    summary_deltas = structured_output.get("summary_deltas", {}) if isinstance(structured_output, dict) else {}
    cost_vs_revenue = None
    if run_id:
        try:
            cost_vs_revenue = _run_cost_vs_revenue_signal(run_id)
        except Exception:  # noqa: BLE001
            cost_vs_revenue = None
    top_drivers = _top_driver_rows(structured_output, limit=6, labor_first=True)
    labor_highlight = [r for r in top_drivers if "labor" in str(r.get("category_label") or r.get("category") or "").lower()][:3]
    artifacts = [str(p) for p in list(compare_payload.get("artifacts") or [])]
    artifact_links = [
        {"path": p, "url": f"/runs/{run_id}/artifacts/{quote(p)}"}
        for p in artifacts[:12]
        if run_id and p
    ]
    return {
        "profit_delta": float(summary_deltas.get("profit", 0) or 0),
        "cost_vs_revenue": cost_vs_revenue,
        "confidence": _confidence_rollup(structured_output),
        "top_drivers": top_drivers,
        "labor_emphasis": labor_highlight,
        "risk_signals": _risk_signals_for_run(compare_payload),
        "artifact_links": artifact_links,
        "workbook_preview_links": _workbook_preview_links(
            workspace_root,
            str(plan.get("prior_path") or ""),
            str(plan.get("current_path") or ""),
        ),
    }


def _summarize_owner_from_run(run_id: str) -> dict[str, Any]:
    state = _load_state(run_id)
    if state.status != "completed" or not isinstance(state.structured_output, dict):
        raise HTTPException(status_code=400, detail=f"Run is not ready for owner summary: {run_id}")
    so = state.structured_output
    sd = so.get("summary_deltas", {}) if isinstance(so, dict) else {}
    profit = float(sd.get("profit", 0) or 0)
    revenue = float(sd.get("revenue", 0) or 0)
    cost = float(sd.get("cost", 0) or 0)

    split = merge_and_split_driver_rows(so)
    primary = list(split.get("primary") or [])
    audit = list(split.get("audit") or [])
    total_m = int(split.get("total") or 0)
    audit_n = int(split.get("auditCount") or 0)
    n_primary = len(primary)

    cost_rev = _run_cost_vs_revenue_signal(run_id)
    sig = str(cost_rev.get("signal") or "balanced")
    ex_conf = _confidence_rollup(so)
    line_note = int(so.get("material_diff_line_count", 0) or 0)

    from_primary = bool(primary)
    source_rows = primary[:3] if from_primary else audit[:3]
    for_audit_fallback = from_primary is False and bool(audit)

    driver_lines: list[str] = []
    for r in source_rows:
        lab = _format_owner_driver_label(r if isinstance(r, dict) else {})
        d = float((r or {}).get("delta", 0) or 0) if isinstance(r, dict) else 0.0
        driver_lines.append(f"- {lab}: {_format_usd_signed(d)} impact")
    if not driver_lines and total_m == 0:
        driver_lines = ["- Not enough line-level movement in the material diff to name specific drivers."]

    if not from_primary and total_m > 0:
        main_header = "Main drivers (from Smaller moves; no line met the Top drivers bar by |Δ| or threshold):"
    else:
        main_header = "Main drivers:"

    main_drivers_block = f"{main_header}\n" + "\n".join(driver_lines)

    if profit > 0:
        out_k = "reflects an improvement in profitability versus the prior view"
    elif profit < 0:
        if cost > 0 and (abs(cost) >= abs(revenue) or revenue <= 0):
            out_k = "reflects margin compression (cost and profit deltas moving against the prior view)"
        else:
            out_k = "reflects a decline in profitability versus the prior view"
    else:
        out_k = "shows flat profit (no net change) net of the revenue and cost deltas"
    overall_sentence = (
        f"Profit moved {_format_usd_signed(profit)}, with revenue at {_format_usd_signed(revenue)} "
        f"and cost at {_format_usd_signed(cost)}; that {out_k}."
    )
    if from_primary and n_primary == 1:
        overall_sentence += (
            " The summary is based on a limited driver set (one line in Top drivers by impact)."
        )

    top0 = source_rows[0] if source_rows and isinstance(source_rows[0], dict) else None
    top_lbl = _format_owner_driver_label(top0) if top0 else "the largest line-level change"
    d0 = float(top0.get("delta", 0) or 0) if top0 else 0.0

    if total_m == 0:
        kc = (
            "The material diff is empty; work from the profit, revenue, and cost rollups in the run "
            "and confirm tie-out to the workbook because line-level support is not present."
        )
    elif for_audit_fallback and top0:
        kc = (
            f"No line reached the Top drivers bar, but the largest |Δ| among the {audit_n} smaller move(s) is "
            f"{top_lbl} ({_format_usd_signed(d0)}), and that is the first line to reconcile in source data."
        )
    elif from_primary and top0 and sig == "cost-led" and (profit < 0 or cost > 0) and (abs(cost) > abs(revenue) or (profit < 0 and d0 < 0)):
        kc = (
            f"The P&L rollups are cost-led versus revenue; the top driver line in this split is {top_lbl} "
            f"({_format_usd_signed(d0)}), so validate that category (and any labor tagged there) in source data first."
        )
    elif from_primary and top0 and sig == "revenue-led" and (abs(revenue) > abs(cost) or d0 < 0):
        kc = (
            f"The P&L rollups are revenue-led; the line-level lead in Top drivers is {top_lbl} "
            f"({_format_usd_signed(d0)}), so confirm that revenue category before messaging costs."
        )
    elif from_primary and top0 and d0 < 0:
        kc = (
            f"The largest |Δ| in Top drivers is a headwind in {top_lbl} ({_format_usd_signed(d0)}), "
            "and that is the first place to look in source workbooks."
        )
    elif from_primary and top0:
        kc = (
            f"The top driver in this split is {top_lbl} ({_format_usd_signed(d0)}); validate that line against "
            "the workbook before rolling the story up for an owner read."
        )
    else:
        kc = (
            f"Reconcile the rolled revenue ({_format_usd_signed(revenue)}), cost ({_format_usd_signed(cost)}), "
            f"and profit ({_format_usd_signed(profit)}) in the run outputs, then re-check line detail if needed."
        )

    if from_primary and top0:
        focus0 = top_lbl
    elif for_audit_fallback and top0:
        focus0 = top_lbl
    else:
        focus0 = "the P&L rollups in the run outputs"
    if from_primary:
        rfocus = (
            f"Start by validating {focus0} in the workbooks, using the same Top driver lines the table used by |Δ| and threshold."
        )
    elif for_audit_fallback and top0:
        rfocus = f"Reconcile the Smaller moves list first, beginning with {focus0}, to see whether any line should have been promoted to Top drivers."
    else:
        rfocus = "Re-open the run artifacts, confirm the extract, and re-run a compare when the workbook is stable."

    c_bits = [f"Extraction confidence: {ex_conf}."]
    c_bits.append(
        f"Line-level context: {total_m} material line(s) in this diff"
        + (f" (engine count {line_note})" if line_note and line_note != total_m else "")
        + "."
    )
    if from_primary and audit_n:
        c_bits.append(f"Smaller moves: {audit_n} line(s) (supporting context only, not the main driver bullets).")
    elif for_audit_fallback and total_m > 0:
        c_bits.append("Main driver bullets are drawn from Smaller moves (no line reached the Top drivers bar).")
    c_line = " ".join(c_bits)

    owner_summary_text = (
        f"Overall:\n{overall_sentence}\n\n{main_drivers_block}\n\n"
        f"Key concern:\n{kc}\n\nRecommended focus:\n{rfocus}\n\nConfidence:\n{c_line}"
    )

    top_for_json = [dict(p) for p in primary[:3]] if from_primary else [dict(a) for a in audit[:3]]

    return {
        "run_id": run_id,
        "profit_delta": profit,
        "revenue_delta": revenue,
        "cost_delta": cost,
        "cost_vs_revenue_signal": sig,
        "top_drivers": top_for_json,
        "driver_table_total_lines": total_m,
        "driver_table_primary_count": n_primary,
        "driver_table_smaller_moves_count": audit_n,
        "key_concern": kc,
        "recommended_focus": rfocus,
        "confidence_line": c_line,
        "owner_summary_text": owner_summary_text,
        "limited_one_primary_driver": bool(from_primary and n_primary == 1),
    }


def _export_row_notes_for_csv(row: dict[str, Any]) -> str:
    for key in ("text", "line", "source_line", "notes", "rationale"):
        v = row.get(key)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _export_top_changes_csv(run_id: str, *, max_rows: int = 2000) -> dict[str, Any]:
    state = _load_state(run_id)
    if state.status != "completed" or not isinstance(state.structured_output, dict):
        raise HTTPException(status_code=400, detail=f"Run is not ready for export: {run_id}")
    if not state.run_dir:
        raise HTTPException(status_code=400, detail=f"Run has no directory for export: {run_id}")
    split = merge_and_split_driver_rows(state.structured_output)
    primary = list(split.get("primary") or [])
    audit = list(split.get("audit") or [])
    top_n = int(split.get("primaryCount") or 0)
    sm_n = int(split.get("auditCount") or 0)
    total = int(split.get("total") or 0)
    if total == 0 or (not primary and not audit):
        return {
            "run_id": run_id,
            "path": None,
            "download_url": None,
            "row_count": 0,
            "top_driver_count": 0,
            "smaller_move_count": 0,
            "rows_written": 0,
            "truncated": False,
            "skipped": True,
            "message": "No merged material diff lines; nothing to export. Compare outputs may have no line-level material rows.",
            "grounding": "Top drivers and Smaller moves are derived the same way as the driver table (|Δ| rank + impact threshold). "
            "Top drivers use the same ranked driver table as the owner summary and export.",
        }
    out_rows: list[dict[str, Any]] = []
    for i, row in enumerate(primary, start=1):
        if len(out_rows) >= max_rows:
            break
        out_rows.append({"section": "Top Driver", "rank": i, "row": row})
    for i, row in enumerate(audit, start=1):
        if len(out_rows) >= max_rows:
            break
        out_rows.append({"section": "Smaller Move", "rank": i, "row": row})
    truncated = top_n + sm_n > len(out_rows)
    run_dir = Path(state.run_dir)
    export_rel = "outputs/top_changes_driver_table.csv"
    export_path = run_dir / export_rel
    export_path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "section",
        "rank",
        "category_label",
        "prior",
        "current",
        "delta",
        "abs_delta",
        "tier",
        "notes",
    ]
    with export_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for item in out_rows:
            row = item["row"]
            if not isinstance(row, dict):
                continue
            label = str(row.get("category_label") or row.get("category") or "")
            pr = float(row.get("prior_value", 0) or 0)
            cr = float(row.get("current_value", 0) or 0)
            dv = float(row.get("delta", 0) or 0)
            ad = _driver_delta_abs(row)
            writer.writerow(
                [
                    item["section"],
                    item["rank"],
                    label,
                    pr,
                    cr,
                    dv,
                    ad,
                    str(row.get("tier") or ""),
                    _export_row_notes_for_csv(row),
                ]
            )
    return {
        "run_id": run_id,
        "path": export_rel,
        "download_url": f"/runs/{run_id}/artifacts/{quote(export_rel)}",
        "row_count": top_n + sm_n,
        "rows_written": len(out_rows),
        "top_driver_count": top_n,
        "smaller_move_count": sm_n,
        "truncated": truncated,
        "skipped": False,
        "grounding": "Grounded in Top drivers by impact. Top drivers use the same ranked driver table as the owner summary and export.",
    }


def _resolve_task_contract(
    query: str,
    context: dict[str, Any],
    workspace: LocalWorkspaceContext,
) -> dict[str, Any]:
    q = query.strip().lower()
    qraw = query.strip()
    selected_files = list(context.get("selected_files") or [])
    trace: list[dict[str, Any]] = []
    workbook_path = str(context.get("target_file") or (selected_files[0] if selected_files else ""))
    ctx = dict(context) if isinstance(context, dict) else {}
    pid_from_q = _scoping_project_id_from_query(qraw)
    if pid_from_q:
        ctx["project_id_filter"] = pid_from_q
    pair_selection = _build_pair_options(workspace.root, qraw, ctx, top_n=3)
    auto_export = "export" in q or bool(context.get("auto_export_top_changes"))

    if re.match(r"^(scan|refresh)\s+workspace\s*!?\s*$", qraw, re.IGNORECASE):
        trace.append({"action": "workspace.index", "details": "scan/refresh workspace command"})
        return {"contract": "scan_workspace", "payload": {}, "trace": trace}
    if re.match(r"^list projects\s*$", qraw, re.IGNORECASE) or re.match(
        r"^list project library\s*$", qraw, re.IGNORECASE
    ):
        trace.append({"action": "workspace.projects", "details": "list projects command"})
        return {"contract": "list_projects", "payload": {}, "trace": trace}
    if re.match(r"^show reports (?:for|on) (?:project )?unlabeled\s*$", qraw, re.IGNORECASE):
        trace.append({"action": "workspace.reports", "details": "show reports for unlabeled group"})
        return {
            "contract": "show_project_reports",
            "payload": {"project_id": UNLABELED_GROUP_ID},
            "trace": trace,
        }
    mrep = re.match(r"^show reports (?:for|on) (?:project )?(\d{5,})\s*$", qraw, re.IGNORECASE)
    if not mrep and "show reports" in q and re.search(r"\b(\d{5,})\b", qraw):
        mrep = re.search(r"\b(\d{5,})\b", qraw)
    if mrep:
        pid = mrep.group(1)
        if is_valid_operator_project_id(pid):
            trace.append({"action": "workspace.reports", "details": f"show reports for {pid}"})
            return {"contract": "show_project_reports", "payload": {"project_id": pid}, "trace": trace}

    mtrend = re.match(r"^trend\s+project\s+(\d{5,})\s*$", qraw, re.IGNORECASE)
    if mtrend:
        pid = mtrend.group(1)
        if is_valid_operator_project_id(pid):
            trace.append({"action": "workspace.trend", "details": f"trend project {pid}"})
            return {"contract": "trend_project_reports", "payload": {"project_id": pid}, "trace": trace}

    mmulti = re.match(r"^compare\s+last\s+(\d{1,2})\s+reports\s+for\s+(\d{5,})\s*$", qraw, re.IGNORECASE)
    if mmulti:
        n_reports = int(mmulti.group(1))
        pid = mmulti.group(2)
        if is_valid_operator_project_id(pid):
            trace.append({"action": "workspace.compare_multi", "details": f"compare last {n_reports} reports for {pid}"})
            return {
                "contract": "compare_multi_reports",
                "payload": {"project_id": pid, "requested_report_count": n_reports},
                "trace": trace,
            }

    if "run weekly review" in q:
        if pid_from_q and _no_pairing_for_scoped_project(pid_from_q, pair_selection):
            trace.append(
                {
                    "action": "pair.resolve",
                    "outcome": "no_workbooks_for_project",
                    "project_id": pid_from_q,
                    "candidates": int((pair_selection or {}).get("filtered_candidate_count", -1)),
                }
            )
            return {
                "contract": "run_weekly_review",
                "payload": {
                    "no_reports_for_project": pid_from_q,
                    "pairing": pair_selection,
                },
                "trace": trace,
            }
        selected_pair = pair_selection.get("selected_pair") if isinstance(pair_selection, dict) else None
        if isinstance(selected_pair, dict):
            current = selected_pair.get("current", {})
            prior = selected_pair.get("prior", {})
            plan = _build_compare_plan(
                workspace.root,
                prior_path=str(prior["path"]),
                current_path=str(current["path"]),
                requested_mode="financial",
                selection_reason=str(selected_pair.get("selection_reason", "")),
            )
            trace.append({"action": "workspace.list_files", "details": "resolved weekly review report pair"})
            trace.append({"action": "compare.invoke_local", "details": "prepared weekly review compare plan"})
            return {
                "contract": "run_weekly_review",
                "payload": {
                    "plan": plan,
                    "pairing": pair_selection,
                    "auto_export_top_changes": True,
                },
                "trace": trace,
            }

    if "compare latest report" in q:
        if pid_from_q and _no_pairing_for_scoped_project(pid_from_q, pair_selection):
            trace.append(
                {
                    "action": "pair.resolve",
                    "outcome": "no_workbooks_for_project",
                    "project_id": pid_from_q,
                    "candidates": int((pair_selection or {}).get("filtered_candidate_count", -1)),
                }
            )
            return {
                "contract": "compare_latest_report",
                "payload": {
                    "no_reports_for_project": pid_from_q,
                    "pairing": pair_selection,
                },
                "trace": trace,
            }
        selected_pair = pair_selection.get("selected_pair") if isinstance(pair_selection, dict) else None
        if isinstance(selected_pair, dict):
            current = selected_pair.get("current", {})
            prior = selected_pair.get("prior", {})
            plan = _build_compare_plan(
                workspace.root,
                prior_path=str(prior["path"]),
                current_path=str(current["path"]),
                requested_mode="financial",
                selection_reason=str(selected_pair.get("selection_reason", "")),
            )
            trace.append({"action": "workspace.list_files", "details": "resolved latest report pair"})
            trace.append({"action": "compare.invoke_local", "details": "prepared latest-report compare plan"})
            return {
                "contract": "compare_latest_report",
                "payload": {"plan": plan, "pairing": pair_selection},
                "trace": trace,
            }

    if ("compare" in q and "labor" in q) and ("latest" in q and ("prior" in q or "previous" in q)):
        if pid_from_q and _no_pairing_for_scoped_project(pid_from_q, pair_selection):
            trace.append(
                {
                    "action": "pair.resolve",
                    "outcome": "no_workbooks_for_project",
                    "project_id": pid_from_q,
                }
            )
            return {
                "contract": "compare_and_show_labor_deltas",
                "payload": {
                    "no_reports_for_project": pid_from_q,
                    "pairing": pair_selection,
                },
                "trace": trace,
            }
        selected_pair = pair_selection.get("selected_pair") if isinstance(pair_selection, dict) else None
        if isinstance(selected_pair, dict):
            current = selected_pair.get("current", {})
            prior = selected_pair.get("prior", {})
            plan = _build_compare_plan(
                workspace.root,
                prior_path=str(prior["path"]),
                current_path=str(current["path"]),
                requested_mode="financial",
                selection_reason=str(selected_pair.get("selection_reason", "")),
            )
            trace.append({"action": "workspace.list_files", "details": "selected latest/prior pair for labor compare"})
            trace.append({"action": "compare.invoke_local", "details": "prepared approved compare plan"})
            return {
                "contract": "compare_and_show_labor_deltas",
                "payload": {"plan": plan, "pairing": pair_selection},
                "trace": trace,
            }

    if ("compare latest vs prior" in q) or ("compare" in q and "latest" in q and ("prior" in q or "previous" in q)):
        if pid_from_q and _no_pairing_for_scoped_project(pid_from_q, pair_selection):
            trace.append(
                {
                    "action": "pair.resolve",
                    "outcome": "no_workbooks_for_project",
                    "project_id": pid_from_q,
                }
            )
            return {
                "contract": "compare_latest_prior_reports",
                "payload": {
                    "no_reports_for_project": pid_from_q,
                    "pairing": pair_selection,
                },
                "trace": trace,
            }
        selected_pair = pair_selection.get("selected_pair") if isinstance(pair_selection, dict) else None
        if isinstance(selected_pair, dict):
            current = selected_pair.get("current", {})
            prior = selected_pair.get("prior", {})
            plan = _build_compare_plan(
                workspace.root,
                prior_path=str(prior["path"]),
                current_path=str(current["path"]),
                requested_mode="financial",
                selection_reason=str(selected_pair.get("selection_reason", "")),
            )
            trace.append({"action": "workspace.list_files", "details": "selected latest/prior pair for compare"})
            trace.append({"action": "compare.invoke_local", "details": "prepared approved compare plan"})
            return {
                "contract": "compare_latest_prior_reports",
                "payload": {"plan": plan, "pairing": pair_selection},
                "trace": trace,
            }

    if (
        "find latest + prior reports" in q
        or "find latest prior reports" in q
        or ("find" in q and "latest" in q and ("prior" in q or "previous" in q) and "report" in q)
        or ("latest vs prior" in q and "compare" not in q)
    ):
        selected_pair = pair_selection.get("selected_pair") if isinstance(pair_selection, dict) else None
        trace.append({"action": "workspace.list_files", "details": "resolved latest/prior report pair"})
        return {
            "contract": "find_latest_prior_reports",
            "payload": {
                "pairing": pair_selection,
                "pair": {
                    "latest": selected_pair.get("current") if isinstance(selected_pair, dict) else None,
                    "prior": selected_pair.get("prior") if isinstance(selected_pair, dict) else None,
                    "selection_reason": selected_pair.get("selection_reason", "") if isinstance(selected_pair, dict) else "",
                    "candidates": pair_selection.get("pair_candidates", []) if isinstance(pair_selection, dict) else [],
                },
            },
            "trace": trace,
        }

    if "cost vs revenue" in q or "cost or revenue" in q:
        run_id = str(context.get("run_id") or _latest_completed_run_id() or "")
        trace.append({"action": "run.cost_vs_revenue", "details": f"assess cost vs revenue for run {run_id or 'none'}"})
        return {"contract": "assess_cost_vs_revenue", "payload": {"run_id": run_id}, "trace": trace}

    if "summarize for owner" in q:
        run_id = str(context.get("run_id") or _latest_completed_run_id() or "")
        trace.append({"action": "run.owner_summary", "details": f"summarize latest run for owner ({run_id or 'none'})"})
        return {"contract": "summarize_for_owner", "payload": {"run_id": run_id}, "trace": trace}

    if "export top changes" in q:
        run_id = str(context.get("run_id") or _latest_completed_run_id() or "")
        trace.append({"action": "run.export_top_changes", "details": f"export top changes for run {run_id or 'none'}"})
        return {"contract": "export_top_changes", "payload": {"run_id": run_id}, "trace": trace}

    if "show run artifacts" in q or "list current run artifacts" in q or "open artifact files" in q:
        run_id = str(context.get("run_id") or _latest_completed_run_id() or "")
        trace.append({"action": "run.list_artifacts", "details": f"list artifacts for run {run_id or 'none'}"})
        return {"contract": "list_current_run_artifacts", "payload": {"run_id": run_id}, "trace": trace}

    if "inspect workbook" in q:
        if not workbook_path:
            selection = _select_latest_prior_workbooks(workspace.root, query=q)
            latest = selection.get("latest")
            workbook_path = str(latest["path"]) if isinstance(latest, dict) else ""
        trace.append({"action": "workspace.inspect_workbook", "details": f"inspect workbook {workbook_path or '(none)'}"})
        return {"contract": "inspect_workbook", "payload": {"workbook_path": workbook_path}, "trace": trace}

    if "preview report sheet" in q or "preview workbook sheet" in q or "preview workbook" in q:
        if not workbook_path:
            selection = _select_latest_prior_workbooks(workspace.root, query=q)
            latest = selection.get("latest")
            workbook_path = str(latest["path"]) if isinstance(latest, dict) else ""
        trace.append({"action": "workspace.preview_workbook_sheet", "details": f"preview workbook sheet for {workbook_path or '(none)'}"})
        return {
            "contract": "preview_report_sheet",
            "payload": {"workbook_path": workbook_path, "sheet_name": str(context.get("sheet_name") or "")},
            "trace": trace,
        }

    if "find report sheets" in q:
        if not workbook_path:
            selection = _select_latest_prior_workbooks(workspace.root, query=q)
            latest = selection.get("latest")
            workbook_path = str(latest["path"]) if isinstance(latest, dict) else ""
        trace.append({"action": "workspace.find_report_sheets", "details": f"find likely report sheets for {workbook_path or '(none)'}"})
        return {"contract": "find_report_sheets", "payload": {"workbook_path": workbook_path}, "trace": trace}

    # Fallback keeps deterministic report navigation behavior and avoids drifting into freeform chat.
    selected_pair = pair_selection.get("selected_pair") if isinstance(pair_selection, dict) else None
    trace.append({"action": "workspace.list_files", "details": "fallback to latest/prior report finder"})
    return {
        "contract": "find_latest_prior_reports",
        "payload": {
            "pairing": pair_selection,
            "pair": {
                "latest": selected_pair.get("current") if isinstance(selected_pair, dict) else None,
                "prior": selected_pair.get("prior") if isinstance(selected_pair, dict) else None,
                "selection_reason": selected_pair.get("selection_reason", "") if isinstance(selected_pair, dict) else "",
                "candidates": pair_selection.get("pair_candidates", []) if isinstance(pair_selection, dict) else [],
            },
        },
        "trace": trace,
    }


@app.on_event("startup")
def _startup() -> None:
    _ensure_dirs()


@app.get("/health")
def health() -> JSONResponse:
    return JSONResponse(
        {
            "status": "ok",
            "repo_root": str(REPO_ROOT),
            "runs_root": str(RUNS_ROOT),
            "registry_path": str(REGISTRY_PATH),
        }
    )


@app.get("/api/local/policy")
def local_policy() -> JSONResponse:
    return JSONResponse(
        {
            "workspace_policy": {
                "default_root": str(DEFAULT_LOCAL_WORKSPACE),
                "allowed_roots": [str(x) for x in ALLOWED_LOCAL_WORKSPACES],
            },
            "read_access": {
                "allowed": True,
                "suffixes": sorted(READABLE_LOCAL_SUFFIXES),
                "max_inline_bytes": MAX_LOCAL_FILE_READ_BYTES,
                "workbook_inspection": {"enabled": True, "suffixes": sorted(EXCEL_SUFFIXES)},
            },
            "write_actions": {"allowed_without_confirmation": False},
            "tool_execution": {"allowed": "approved wrappers only", "arbitrary_shell": False},
            "registered_actions": [
                {
                    "key": spec.key,
                    "description": spec.description,
                    "requires_confirmation": spec.requires_confirmation,
                }
                for spec in LOCAL_ACTION_REGISTRY.values()
            ],
        }
    )


@app.get("/api/local/workspace")
def local_workspace(root: str | None = None) -> JSONResponse:
    raw = (root or "").strip()
    if not raw:
        raw = str(_load_workspace_config().get("default_workspace_root") or "").strip()
    ctx, _wb, werr = _open_workspace_for_assistant(raw or None)
    cfg = _load_workspace_config()
    if ctx is not None:
        snap = _readiness_snapshot(ctx.root)
        nidx = _index_workbook_count_for_root(str(ctx.root))
        live_wb = int(snap.get("workbook_count", 0) or 0)
        maxc = max(live_wb, nidx)
        projects = _project_index_summary_for_ui(ctx.root)
        needs_scan = bool(
            nidx == 0
            and int(snap.get("workbook_count", 0) or 0) > 0
        ) or (str(cfg.get("indexed_at") or "").strip() == "" and maxc > 0)
        index_may_be_stale = bool(nidx > 0 and live_wb > nidx)
        return JSONResponse(
            {
                "workspace_root": str(ctx.root),
                "default_workspace": str(DEFAULT_LOCAL_WORKSPACE),
                "resolvable": True,
                "readiness": snap,
                "persisted": {
                    "default_workspace_root": str(cfg.get("default_workspace_root") or ""),
                    "indexed_at": str(cfg.get("indexed_at") or ""),
                    "last_compare_at": str(cfg.get("last_compare_at") or ""),
                    "last_compare_project_id": str(cfg.get("last_compare_project_id") or ""),
                    "allowed_workspace_roots": list(cfg.get("allowed_workspace_roots") or []),
                },
                "project_index": {
                    "indexed_workbooks": nidx,
                    "indexed_at": str(cfg.get("indexed_at") or ""),
                    "needs_scan": needs_scan,
                    "index_may_be_stale": index_may_be_stale,
                    "live_workbook_count": live_wb,
                    "scan_file_cap": _INDEX_SCAN_MAX_FILES,
                    "projects": projects,
                },
            }
        )
    rroot = str(werr.get("readiness", {}).get("resolved_root", DEFAULT_LOCAL_WORKSPACE) if werr else DEFAULT_LOCAL_WORKSPACE)
    return JSONResponse(
        {
            "workspace_root": rroot,
            "default_workspace": str(DEFAULT_LOCAL_WORKSPACE),
            "resolvable": False,
            "readiness": werr.get("readiness", {}) if werr else {},
            "message": werr.get("message", "") if werr else "Workspace could not be opened.",
            "persisted": {
                "default_workspace_root": str(cfg.get("default_workspace_root") or ""),
                "indexed_at": str(cfg.get("indexed_at") or ""),
            },
            "project_index": {
                "indexed_workbooks": 0,
                "projects": [],
                "needs_scan": True,
                "indexed_at": "",
                "index_may_be_stale": False,
                "live_workbook_count": 0,
                "scan_file_cap": _INDEX_SCAN_MAX_FILES,
            },
        }
    )


@app.post("/api/local/workspace")
def set_local_workspace(payload: dict[str, Any] = Body(...)) -> JSONResponse:
    root = str(payload.get("workspace_root") or "").strip()
    if not root:
        raise HTTPException(status_code=400, detail="workspace_root is required.")
    workspace = _workspace_context(root)
    allowed = payload.get("allowed_workspace_roots")
    upd: dict[str, Any] = {
        "default_workspace_root": str(workspace.root),
    }
    if isinstance(allowed, list):
        cleaned: list[str] = []
        for x in allowed:
            s = str(x or "").strip()
            if not s:
                continue
            try:
                p = _validate_allowed_root_path(s)
                cleaned.append(str(p))
            except HTTPException:
                continue
        if cleaned:
            upd["allowed_workspace_roots"] = list(dict.fromkeys(cleaned + [str(p) for p in ALLOWED_LOCAL_WORKSPACES]))
    _save_workspace_config(upd)
    return JSONResponse({"workspace_root": str(workspace.root), "persisted": upd})


@app.post("/api/local/workspace/scan")
def post_workspace_scan() -> JSONResponse:
    cfg = _load_workspace_config()
    wroot = str(cfg.get("default_workspace_root") or "").strip()
    if not wroot:
        wroot = str(DEFAULT_LOCAL_WORKSPACE)
    wpath = _validate_allowed_root_path(wroot)
    summary = _scan_workspace_path(wpath)
    return JSONResponse({"workspace_root": str(wpath), **summary})


@app.get("/api/local/files")
def local_files(
    workspace_root: str | None = None,
    query: str = "",
    limit: int = 60,
) -> JSONResponse:
    workspace = _workspace_context(workspace_root)
    files = _list_workspace_files(workspace.root, query=query, limit=limit)
    return JSONResponse({"workspace_root": str(workspace.root), "files": files})


@app.get("/api/local/file")
def local_file(workspace_root: str | None = None, path: str = "") -> JSONResponse:
    if not path:
        raise HTTPException(status_code=400, detail="path is required.")
    workspace = _workspace_context(workspace_root)
    payload = _read_workspace_file(workspace.root, path)
    return JSONResponse({"workspace_root": str(workspace.root), **payload})


@app.get("/api/local/workbook/inspect")
def local_workbook_inspect(workspace_root: str | None = None, path: str = "") -> JSONResponse:
    if not path:
        raise HTTPException(status_code=400, detail="path is required.")
    workspace = _workspace_context(workspace_root)
    payload = _workbook_metadata(workspace.root, path)
    return JSONResponse({"workspace_root": str(workspace.root), **payload})


@app.get("/api/local/workbook/report-sheets")
def local_workbook_report_sheets(workspace_root: str | None = None, path: str = "") -> JSONResponse:
    if not path:
        raise HTTPException(status_code=400, detail="path is required.")
    workspace = _workspace_context(workspace_root)
    payload = _find_report_sheets(workspace.root, path)
    return JSONResponse({"workspace_root": str(workspace.root), **payload})


@app.get("/api/local/workbook/preview")
def local_workbook_preview(
    workspace_root: str | None = None,
    path: str = "",
    sheet_name: str | None = None,
    max_rows: int = WORKBOOK_PREVIEW_MAX_ROWS,
    max_cols: int = WORKBOOK_PREVIEW_MAX_COLS,
) -> JSONResponse:
    if not path:
        raise HTTPException(status_code=400, detail="path is required.")
    workspace = _workspace_context(workspace_root)
    payload = _preview_workbook_sheet(
        workspace.root,
        workbook_path=path,
        sheet_name=sheet_name,
        max_rows=max_rows,
        max_cols=max_cols,
    )
    return JSONResponse({"workspace_root": str(workspace.root), **payload})


@app.get("/api/local/financial/219128_feb_mar_mom")
def api_financial_219128_feb_mar_mom() -> JSONResponse:
    """Serves the same deterministic MOM report as `mom_219128_feb_mar` in financial signals (219128)."""
    return JSONResponse(_mom_219128_feb_mar_payload())


def _assistant_task_envelope(
    *,
    query: str,
    workspace_root: str,
    contract: str | None,
    status: str,
    answer: str,
    result: dict[str, Any],
    trace: list[dict[str, Any]],
    policy: dict[str, Any] | None = None,
    approval: dict[str, Any] | None = None,
    result_contract: str | None = None,
) -> JSONResponse:
    body: dict[str, Any] = {
        "task": {"query": query, "workspace_root": workspace_root, "contract": contract or ""},
        "status": status,
        "answer": answer,
        "result_contract": (result_contract if result_contract is not None else contract) or "",
        "result": result,
        "trace": trace,
        "policy": policy or {},
    }
    if approval is not None:
        body["approval"] = approval
    return JSONResponse(body, status_code=200)


def _friendly_assistant_error(detail: Any) -> str:
    s = str(detail)
    if s.strip().lower() in {"not found", "file not found"} or "file not found" in s.lower():
        return "Something that was required for this task is missing. Run a compare first, or set a valid workspace, then try again."
    if "no completed run" in s.lower() or "not available" in s.lower() and "run" in s.lower():
        return "Run a compare first, then run this command again."
    return s


@app.post("/api/local/assistant/task")
def local_assistant_task(payload: dict[str, Any] = Body(...)) -> JSONResponse:
    query = str(payload.get("query") or "").strip()
    report_builder = bool(payload.get("report_builder")) and str(payload.get("contract") or "").strip() in {
        "generate_financial_signals",
    }
    if not query and not report_builder:
        return _assistant_task_envelope(
            query="",
            workspace_root="",
            contract=None,
            status="failed",
            answer="Use the Report Builder to select a project, reports, and an analysis, then Run analysis.",
            result={"next_steps": [f"Example: {SUGGESTED_FIRST_COMPARE}"]},
            trace=[{"action": "preflight", "outcome": "empty_query"}],
        )
    try:
        return _local_assistant_task_core(payload, query)
    except HTTPException as he:
        d = he.detail
        d_str = d if isinstance(d, str) else str(d)
        return _assistant_task_envelope(
            query=query,
            workspace_root="",
            contract=None,
            status="failed",
            answer=_friendly_assistant_error(d_str),
            result={
                "next_steps": ["If this persists, use Advanced to verify workspace root, then try again."],
            },
            trace=[{"action": "exception", "code": he.status_code, "detail": d_str}],
        )


def _local_assistant_task_core(payload: dict[str, Any], query: str) -> JSONResponse:
    workspace, wbc, werr = _open_workspace_for_assistant(payload.get("workspace_root"))
    wroot = str(workspace.root) if workspace else str(
        werr.get("readiness", {}).get("resolved_root", DEFAULT_LOCAL_WORKSPACE) if werr else DEFAULT_LOCAL_WORKSPACE
    )
    if werr:
        return _assistant_task_envelope(
            query=query,
            workspace_root=wroot,
            contract=None,
            status=werr["status"],
            answer=werr["message"],
            result={
                "readiness": werr.get("readiness"),
                "next_steps": [
                    "In Advanced, set workspace root to an existing, allowed folder, or use the default shown when the page loads.",
                ],
            },
            trace=werr.get("trace", []),
        )
    assert workspace is not None
    wb_count = max(
        wbc,
        _index_workbook_count_for_root(str(workspace.root)),
    )
    if bool(payload.get("report_builder")) and str(payload.get("contract") or "").strip() == "generate_financial_signals":
        if wb_count == 0:
            idx_n = _index_workbook_count_for_root(str(workspace.root))
            looked = _workbook_looked_paths_message(workspace, idx_n > 0)
            snap = _readiness_snapshot(workspace.root)
            return _assistant_task_envelope(
                query=query or "Report Builder",
                workspace_root=str(workspace.root),
                contract="generate_financial_signals",
                status="no_candidates",
                answer="No Excel workbooks were found for this operator workspace, so the Report Builder cannot read files yet. "
                + looked,
                result={
                    "readiness": snap,
                    "next_steps": [
                        "Add approved monthly workbooks under the workspace root, then run Scan workspace.",
                    ],
                },
                trace=[{"action": "preflight.workbooks", "outcome": "none_found"}],
            )
        return _run_generate_financial_signals(workspace, payload)
    context: dict[str, Any] = dict(payload.get("context") or {}) if isinstance(payload.get("context"), dict) else {}
    cfg_mem = _load_workspace_config()
    lcp = cfg_mem.get("last_confirmed_pair")
    if not context.get("last_confirmed_pair") and isinstance(lcp, dict) and lcp:
        context["last_confirmed_pair"] = lcp
    pfr = str(cfg_mem.get("preferred_report_family") or "").strip()
    if pfr and not str(context.get("preferred_report_family") or "").strip():
        context["preferred_report_family"] = pfr
    routed = _resolve_task_contract(query, context, workspace)
    contract = str(routed["contract"])
    if contract not in TASK_CONTRACTS:
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="failed",
            answer="That phrasing is not mapped to a supported operator action yet. Try a shorter command or one of the shortcuts.",
            result={"next_steps": [f"Try: {SUGGESTED_FIRST_COMPARE}"]},
            trace=list(routed.get("trace", [])),
        )
    trace = list(routed.get("trace", []))
    task_payload = routed.get("payload", {}) if isinstance(routed.get("payload"), dict) else {}
    if str(task_payload.get("no_reports_for_project") or "").strip():
        pid = str(task_payload.get("no_reports_for_project") or "")
        has_idx = _index_workbook_count_for_root(str(workspace.root)) > 0
        line = _workspace_source_line_for_transcript(index_backed=has_idx)
        pairing = task_payload.get("pairing") if isinstance(task_payload.get("pairing"), dict) else {}
        n_cand = int(pairing.get("filtered_candidate_count", -1) or -1)
        if n_cand <= 0:
            answer = (
                f"No indexed workbooks for project {pid} under this workspace. "
                f"Confirm paths include {pid} and run scan workspace."
            )
        elif n_cand == 1:
            answer = (
                f"Only one indexed workbook for project {pid}. "
                f"Compare needs at least two in the same project (and pairing prefers two in the same report family)."
            )
        else:
            answer = (
                f"Found {n_cand} workbook(s) for project {pid}, but no before/after pair could be built "
                f"(each report family needs at least two files). Scoped commands do not fall back to other projects."
            )
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="no_candidates",
            answer=answer,
            result={
                "source_line": line,
                "provenance": line,
                "project_filter": pid,
                "pairing": pairing,
                "next_steps": [
                    f"Run: show reports for {pid} to list indexed paths and families.",
                    "If the index is empty or stale, run: scan workspace.",
                ],
            },
            trace=trace,
        )
    if contract == "scan_workspace":
        summ = _scan_workspace_path(workspace.root)
        ok = bool(summ.get("ok"))
        extra = ""
        if ok and summ.get("cap_reached"):
            ex = int(summ.get("files_examined") or 0)
            extra = f" Showing first {ex} Excel files (scan limit)."
        src_line = ""
        if ok:
            idx_ts = str(summ.get("indexed_at") or "").strip()
            src_line = f"Source: live scan (indexed {idx_ts}) after this pass." if idx_ts else "Source: live scan (this pass updated the index)."
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="completed" if ok else "failed",
            answer=(
                f"Indexed {int(summ.get('rows_indexed', 0) or 0)} workbook(s) under {summ.get('root', '')}.{extra}"
                if ok
                else f"Scan failed: {summ.get('error', 'unknown error')}"
            ),
            result={
                "scan": summ,
                "source_line": src_line,
                "next_steps": ["Reload the page or run list projects to see the Project Library update."],
            },
            trace=trace,
        )
    if contract == "list_projects":
        projs = _project_index_summary_for_ui(workspace.root)
        has_idx = _index_workbook_count_for_root(str(workspace.root)) > 0
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="completed",
            answer=f"Found {len(projs)} project group(s) in the workspace index.",
            result={
                "project_index": {"projects": projs},
                "source_line": _workspace_source_line_for_transcript(index_backed=has_idx),
                "next_steps": ["Use compare latest report for <id> to pair within a project."],
            },
            trace=trace,
        )
    if contract == "show_project_reports":
        raw_pid = str(task_payload.get("project_id") or "")
        is_unl = raw_pid in (UNLABELED_GROUP_ID, "(unlabeled)", "unlabeled")
        filter_id = UNLABELED_GROUP_ID if is_unl else raw_pid
        label = (
            "unlabeled workbooks (no project ID in paths)" if is_unl else f"project {raw_pid}"
        )
        has_index = _index_workbook_count_for_root(str(workspace.root)) > 0
        if has_index:
            rows = _index_candidate_models(workspace.root, "", filter_id)
        else:
            rows = [
                m
                for m in _candidate_models_from_rglob(workspace.root, "", {}, context)
                if (is_unl and not str(m.get("project_id") or "").strip())
                or (not is_unl and str(m.get("project_id") or "") == filter_id)
            ]
        n = len(rows)
        looked = _workbook_looked_paths_message(workspace, has_index)
        source = "workspace index" if has_index else "live workspace scan (index empty)"
        if n == 0:
            answer = f"No workbooks under {label} in the {source}. " + looked
        else:
            answer = f"{label}: {n} workbook file(s) in the {source}."
        src_line = _workspace_source_line_for_transcript(index_backed=has_index)
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="completed",
            answer=answer,
            result={
                "project_id": raw_pid,
                "reports": rows[:40],
                "source": source,
                "source_line": src_line,
                "evidence_looked": looked if n == 0 else "",
            },
            trace=trace,
        )
    if contract == "trend_project_reports":
        pid = str(task_payload.get("project_id") or "")
        trend = _build_project_trend_artifact(workspace.root, pid)
        report_count = int(trend.get("report_count") or 0)
        status = "completed" if report_count >= 3 else "insufficient_data"
        if status == "completed":
            answer = (
                f"Trend ready for project {pid}: {report_count} indexed report(s) ordered "
                "oldest-to-newest from workspace index metadata."
            )
        else:
            answer = (
                f"Insufficient indexed reports for project {pid}: found {report_count}, need at least 3. "
                "Scoped trend does not fall back to other projects."
            )
        suggestions = [
            f"Run: show reports for {pid} to inspect indexed paths and report families.",
            "Run: scan workspace if files were added after the last index pass.",
        ]
        trend_summary = trend.get("trend_summary", {}) if isinstance(trend.get("trend_summary"), dict) else {}
        family_counts = trend_summary.get("family_counts", {}) if isinstance(trend_summary.get("family_counts"), dict) else {}
        family_line = ", ".join(f"{family}: {count}" for family, count in sorted(family_counts.items()))
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "workspace_root": str(workspace.root), "project_id": pid},
            "did": {
                "method": "read workspace index rows for the scoped project only; no global fallback",
            },
            "found": trend,
            "project_filter": pid,
            "report_count": report_count,
            "reports_used": trend.get("reports_used", []),
            "period_start": trend.get("period_start", ""),
            "period_end": trend.get("period_end", ""),
            "trend_summary": trend.get("trend_summary", {}),
            "owner_trend_summary": trend.get("owner_trend_summary", {}),
            "source_line": trend.get("source_line", ""),
            "provenance": trend.get("provenance", ""),
            "next_steps": suggestions,
            "summary_card": {
                "title": "Project Trend" if status == "completed" else "Project Trend Needs More Reports",
                "summary": answer,
                "paths": [str(r.get("path")) for r in list(trend.get("reports_used") or [])[:6] if isinstance(r, dict)],
                "key_outputs": [
                    f"Project filter: {pid}",
                    f"Reports indexed for project: {report_count}",
                    f"Period: {trend.get('period_start', '')} to {trend.get('period_end', '')}",
                    f"Basis: {trend_summary.get('basis', 'workspace_index_metadata')}",
                    f"Ordering: {trend_summary.get('ordering', 'oldest-to-newest by indexed report metadata')}",
                    f"Report families: {family_line or 'none'}",
                ],
                "suggestions": suggestions,
            },
        }
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status=status,
            answer=answer,
            result=result_payload,
            trace=trace,
        )
    if contract == "compare_multi_reports":
        pid = str(task_payload.get("project_id") or "")
        requested_n = int(task_payload.get("requested_report_count") or 0)
        multi = _build_multi_report_compare_artifact(workspace.root, pid, requested_n)
        status = str(multi.get("status") or "insufficient_data")
        if status == "completed":
            multi_period_delta = _build_multi_period_delta(workspace.root, multi)
        else:
            multi_period_delta = {
                "pair_results": [],
                "repeated_movers": [],
                "largest_cumulative_movers": [],
                "latest_period_movers": [],
                "action_items": [],
                "movement_categories": _movement_category_summary([]),
                "classification_summary": _classification_summary([]),
                "uncategorized_count": 0,
                "latest_period_watchlist": [],
                "repeated_risk_items": [],
                "action_view": {
                    "top_issues": [],
                    "new_this_period": [],
                    "ongoing_risks": [],
                    "watchlist": [],
                },
                "cost_type_drilldown": _build_cost_type_drilldown([]),
                "owner_lines": list(multi.get("owner_lines") or []),
                "limitations": [
                    f"Need {requested_n} indexed report(s) for project {pid}; found {int(multi.get('available_report_count') or 0)}.",
                    "No global workspace fallback was used.",
                ],
            }
        multi["multi_period_delta"] = multi_period_delta
        if status == "completed":
            answer = (
                f"Multi-period comparison ready for project {pid}: "
                f"{requested_n} indexed report(s), {int(multi.get('pair_count') or 0)} adjacent pair(s)."
            )
        else:
            answer = (
                f"Insufficient indexed reports for project {pid}: found {int(multi.get('available_report_count') or 0)}, "
                f"need {requested_n}. Scoped multi-period compare does not fall back to other projects."
            )
        suggestions = [
            f"Run: show reports for {pid} to inspect indexed paths and report families.",
            "Run: scan workspace if files were added after the last index pass.",
        ]
        result_payload = {
            "contract": contract,
            "requested": {
                "query": query,
                "workspace_root": str(workspace.root),
                "project_id": pid,
                "requested_report_count": requested_n,
            },
            "did": {
                "method": "read workspace index rows for the scoped project only; built adjacent metadata pairs; no global fallback",
            },
            "found": multi,
            "status": status,
            "project_filter": pid,
            "requested_report_count": requested_n,
            "report_count": int(multi.get("report_count") or 0),
            "reports_used": multi.get("reports_used", []),
            "pair_count": int(multi.get("pair_count") or 0),
            "comparison_pairs": multi.get("comparison_pairs", []),
            "multi_period_delta": multi_period_delta,
            "period_start": multi.get("period_start", ""),
            "period_end": multi.get("period_end", ""),
            "owner_lines": multi_period_delta.get("owner_lines") or multi.get("owner_lines", []),
            "source_line": multi.get("source_line", ""),
            "provenance": multi.get("provenance", ""),
            "next_steps": suggestions,
            "summary_card": {
                "title": "Multi-Period Compare" if status == "completed" else "Multi-Period Compare Needs More Reports",
                "summary": answer,
                "paths": [str(r.get("path")) for r in list(multi.get("reports_used") or [])[:6] if isinstance(r, dict)],
                "key_outputs": [
                    f"Project filter: {pid}",
                    f"Requested reports: {requested_n}",
                    f"Reports selected: {int(multi.get('report_count') or 0)}",
                    f"Adjacent pairs: {int(multi.get('pair_count') or 0)}",
                    f"Period: {multi.get('period_start', '')} to {multi.get('period_end', '')}",
                ],
                "suggestions": suggestions,
            },
        }
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status=status,
            answer=answer,
            result=result_payload,
            trace=trace,
        )
    if contract in CONTRACTS_NEEDING_LATEST_PASS_RUN:
        rid = str(task_payload.get("run_id") or context.get("run_id") or "").strip() or _latest_completed_run_id() or ""
        if not rid or not _run_has_operator_outputs(rid):
            return _assistant_task_envelope(
                query=query,
                workspace_root=str(workspace.root),
                contract=contract,
                status="needs_run",
                answer="Run a compare first. There is no finished compare in this environment to use yet.",
                result={
                    "readiness": {"has_completed_run": False, "suggested_query": SUGGESTED_FIRST_COMPARE},
                    "next_steps": [f"Run: {SUGGESTED_FIRST_COMPARE} and confirm, wait for it to complete, then try this again."],
                    "suggested_query": SUGGESTED_FIRST_COMPARE,
                },
                trace=trace,
            )
    if contract in CONTRACTS_NEEDING_WORKBOOK_FILES and wb_count == 0:
        snap = _readiness_snapshot(workspace.root)
        idx_n = _index_workbook_count_for_root(str(workspace.root))
        looked = _workbook_looked_paths_message(workspace, idx_n > 0)
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="no_candidates",
            answer="No Excel workbooks were found for this operator workspace, so the operator cannot pick files yet. "
            + looked,
            result={
                "readiness": snap,
                "next_steps": [
                    "Set an approved workspace root in Advanced (it is remembered on this server).",
                    "Run the command: scan workspace (or use Scan in Project Library) to build the project index.",
                    "If you have no local tree yet, use Advanced → upload as a fallback.",
                    f"Then run: {SUGGESTED_FIRST_COMPARE}.",
                ],
            },
            trace=trace
            + [
                {
                    "action": "preflight.workbooks",
                    "outcome": "none_found",
                    "index_workbook_rows": idx_n,
                }
            ],
        )
    confirm = bool(payload.get("confirm", False))
    compare_contract = contract in {
        "compare_latest_report",
        "compare_latest_prior_reports",
        "compare_and_show_labor_deltas",
        "run_weekly_review",
    }
    if compare_contract:
        pairing = task_payload.get("pairing") if isinstance(task_payload.get("pairing"), dict) else {}
        plan = task_payload.get("plan") if isinstance(task_payload.get("plan"), dict) else {}
        selected_pair = pairing.get("selected_pair") if isinstance(pairing.get("selected_pair"), dict) else {}
        pair_candidates = list(pairing.get("pair_candidates") or [])
        requires_operator_selection = bool(pairing.get("requires_operator_selection", False))
        has_operator_selected_pair = bool(
            _pairing_overrides(query, context).get("selected_pair_id")
            and any(
                str(p.get("pair_id")) == str(_pairing_overrides(query, context).get("selected_pair_id"))
                for p in pair_candidates
            )
        )
        if not plan.get("prior_path") or not plan.get("current_path"):
            return _assistant_task_envelope(
                query=query,
                workspace_root=str(workspace.root),
                contract=contract,
                status="no_candidates",
                answer="Could not select a before/after workbook pair. Add or rename at least two related Excel files so prior and current can be told apart.",
                result={
                    "contract": contract,
                    "requested": {"query": query},
                    "did": {"pair_selection": "attempted"},
                    "found": {
                        "pair_ready": False,
                        "candidate_pairs": pair_candidates,
                        "pairing_confidence": pairing.get("pairing_confidence", 0),
                    },
                    "readiness": _readiness_snapshot(workspace.root) | {"workbook_count": wb_count},
                    "next_steps": [
                        f"Add a clearer prior/current pair, or try: {SUGGESTED_FIRST_COMPARE} after files are in place.",
                    ],
                },
                trace=trace
                + [
                    {
                        "action": "pair.resolve",
                        "outcome": "no_plan",
                    }
                ],
            )
        if confirm and requires_operator_selection and not has_operator_selected_pair:
            approval_block = {
                "files": {"prior": plan["prior_path"], "current": plan["current_path"]},
                "compare_mode": {
                    "requested_mode": plan.get("requested_mode"),
                    "selected_mode": plan.get("selected_mode"),
                    "compare_path": plan.get("compare_path"),
                    "selection_reason": plan.get("compare_path_reason"),
                },
                "pairing": {
                    "pairing_confidence": pairing.get("pairing_confidence", 0),
                    "requires_operator_selection": True,
                    "disambiguation_required": bool(pairing.get("disambiguation_required", False)),
                    "candidate_pairs": pair_candidates,
                    "overrides_applied": pairing.get("overrides_applied", {}),
                },
                "expected_outputs": list(plan.get("expected_artifacts", [])),
            }
            return _assistant_task_envelope(
                query=query,
                workspace_root=str(workspace.root),
                contract=contract,
                status="needs_confirmation",
                answer="Pair confidence is below threshold. Select one candidate pair before confirming compare.",
                result={
                    "contract": contract,
                    "source_line": _workspace_source_line_for_transcript(
                        index_backed=_index_workbook_count_for_root(str(workspace.root)) > 0
                    ),
                    "requested": {"query": query},
                    "did": {"pair_selection": "prepared candidate set; waiting for operator pick"},
                    "found": {
                        "pair_ready": True,
                        "selected_pair": selected_pair,
                        "candidate_pairs": pair_candidates,
                        "pairing_confidence": pairing.get("pairing_confidence", 0),
                    },
                    "next_steps": ["Pick a pair in the confirmation bar above the transcript, then confirm."],
                },
                trace=trace,
                policy={"requires_confirmation": True, "reason": "low-confidence pair requires explicit selection"},
                approval=approval_block,
            )
        if not confirm:
            approval_block2 = {
                "files": {"prior": plan["prior_path"], "current": plan["current_path"]},
                "compare_mode": {
                    "requested_mode": plan.get("requested_mode"),
                    "selected_mode": plan.get("selected_mode"),
                    "compare_path": plan.get("compare_path"),
                    "selection_reason": plan.get("compare_path_reason"),
                },
                "pairing": {
                    "pairing_confidence": pairing.get("pairing_confidence", 0),
                    "requires_operator_selection": requires_operator_selection,
                    "disambiguation_required": bool(pairing.get("disambiguation_required", False)),
                    "candidate_pairs": pair_candidates,
                    "overrides_applied": pairing.get("overrides_applied", {}),
                },
                "expected_outputs": list(plan.get("expected_artifacts", [])),
            }
            return _assistant_task_envelope(
                query=query,
                workspace_root=str(workspace.root),
                contract=contract,
                status="needs_confirmation",
                answer="Compare is staged and ready. Confirm to execute the deterministic compare workflow.",
                result={
                    "contract": contract,
                    "source_line": _workspace_source_line_for_transcript(
                        index_backed=_index_workbook_count_for_root(str(workspace.root)) > 0
                    ),
                    "requested": {"query": query},
                    "did": {
                        "pair_selection": "latest/prior workbook resolver",
                        "compare_plan": "deterministic local compare plan prepared",
                    },
                    "found": {
                        "pair_ready": True,
                        "latest_report": plan["current_path"],
                        "prior_report": plan["prior_path"],
                        "compare_path": plan.get("compare_path"),
                        "selected_pair": selected_pair,
                        "candidate_pairs": pair_candidates,
                        "pairing_confidence": pairing.get("pairing_confidence", 0),
                    },
                    "next_steps": [
                        "Review the staged files in the transcript entry.",
                        "Confirm in the bar above the transcript to run the compare.",
                    ],
                },
                trace=trace,
                policy={"requires_confirmation": True, "reason": "compare wrapper execution"},
                approval=approval_block2,
            )

    result_payload: dict[str, Any]
    answer: str
    suggestions: list[str]
    paths: list[str]
    key_outputs: list[str]
    if contract == "find_latest_prior_reports":
        pair = task_payload.get("pair", {}) if isinstance(task_payload.get("pair"), dict) else {}
        pairing = task_payload.get("pairing", {}) if isinstance(task_payload.get("pairing"), dict) else {}
        latest = pair.get("latest")
        prior = pair.get("prior")
        latest_path = str(latest.get("path")) if isinstance(latest, dict) else ""
        prior_path = str(prior.get("path")) if isinstance(prior, dict) else ""
        pair_ready = bool(latest_path and prior_path)
        answer = "Resolved the most likely latest/prior workbook pair for this workspace."
        suggestions = (
            ["Run `Compare latest vs prior` to execute a deterministic compare."]
            if pair_ready
            else ["No usable pair found. Add or rename workbook versions, then retry."]
        )
        paths = [p for p in [latest_path, prior_path] if p]
        key_outputs = [str(pair.get("selection_reason", ""))]
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "workspace_root": str(workspace.root)},
            "did": {
                "resolver": "ranked workbook family selector",
                "selection_reason": pair.get("selection_reason", ""),
            },
            "found": {
                "pair_ready": pair_ready,
                "latest_report": latest,
                "prior_report": prior,
                "candidate_pairs": pairing.get("pair_candidates", []),
                "pairing_confidence": pairing.get("pairing_confidence", 0),
                "disambiguation_required": pairing.get("disambiguation_required", False),
                "requires_operator_selection": pairing.get("requires_operator_selection", False),
            },
            "next_steps": suggestions,
            "summary_card": {
                "title": "Latest + Prior Reports",
                "summary": "Ready to compare." if pair_ready else "Need a stronger workbook pair.",
                "paths": paths,
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
        }
    elif contract in {
        "compare_latest_report",
        "compare_latest_prior_reports",
        "compare_and_show_labor_deltas",
        "run_weekly_review",
    }:
        plan = task_payload["plan"]
        pairing = task_payload.get("pairing", {}) if isinstance(task_payload.get("pairing"), dict) else {}
        selected_pair = pairing.get("selected_pair") if isinstance(pairing.get("selected_pair"), dict) else {}
        compare = _execute_local_compare(
            workspace.root,
            prior_path=str(plan["prior_path"]),
            current_path=str(plan["current_path"]),
            workflow_mode=str(plan.get("selected_mode") or "auto"),
        )
        so = compare.get("structured_output", {}) if isinstance(compare.get("structured_output"), dict) else {}
        labor_deltas: list[dict[str, Any]] = []
        if contract == "compare_and_show_labor_deltas":
            labor_deltas = _derive_labor_deltas(str(compare.get("run_id", "")), so)
        workflow_output = _operator_compare_output(compare, workspace.root, plan)
        weekly_driver_split: dict[str, Any] | None = None
        if contract == "run_weekly_review" and isinstance(so, dict):
            weekly_driver_split = merge_and_split_driver_rows(so)
            prim = list(weekly_driver_split.get("primary") or [])
            aud = list(weekly_driver_split.get("audit") or [])
            from_primary = bool(prim)
            rows_for_wo = (prim[:6] if from_primary else aud[:6]) if (prim or aud) else []
            out_td: list[dict[str, Any]] = []
            for r in rows_for_wo:
                if not isinstance(r, dict):
                    continue
                d = dict(r)
                d["display_label"] = _format_owner_driver_label(r)
                out_td.append(d)
            workflow_output["top_drivers"] = out_td
            workflow_output["top_drivers_from_smaller_moves_only"] = bool(
                (not from_primary) and (weekly_driver_split.get("total", 0) or 0) > 0
            )
            workflow_output["driver_table_primary_count"] = int(weekly_driver_split.get("primaryCount") or 0)
            workflow_output["driver_table_smaller_moves_count"] = int(weekly_driver_split.get("auditCount") or 0)
        compare_summary = list(compare.get("summary") or [])
        run_id = str(compare.get("run_id", ""))
        artifact_paths = [str(p) for p in list(compare.get("artifacts") or [])[:12]]
        key_outputs = [f"Compare path: {compare.get('compare_path', 'unknown')}"]
        if contract == "run_weekly_review" and isinstance(so, dict):
            sdx = so.get("summary_deltas", {})
            if isinstance(sdx, dict):
                key_outputs.append(
                    "P/L (rolled): "
                    f"profit={float(sdx.get('profit', 0) or 0):,.0f} | "
                    f"revenue={float(sdx.get('revenue', 0) or 0):,.0f} | "
                    f"cost={float(sdx.get('cost', 0) or 0):,.0f}"
                )
        if contract == "run_weekly_review" and weekly_driver_split is not None:
            prim = list(weekly_driver_split.get("primary") or [])
            aud = list(weekly_driver_split.get("audit") or [])
            from_primary = bool(prim)
            n_top = int(weekly_driver_split.get("primaryCount") or 0)
            n_sm = int(weekly_driver_split.get("auditCount") or 0)
            src = prim[:3] if from_primary else aud[:3]
            if src:
                toks = []
                for d in src:
                    if not isinstance(d, dict):
                        continue
                    lab = _format_owner_driver_label(d)
                    toks.append(f"{lab} ({float(d.get('delta', 0) or 0):+,.0f})")
                if from_primary:
                    key_outputs.append("Top drivers by impact: " + " · ".join(toks))
                else:
                    key_outputs.append(
                        "Smaller moves (no line met the Top drivers bar by |Δ| or threshold): " + " · ".join(toks)
                    )
            if (weekly_driver_split.get("total") or 0) == 0:
                key_outputs.append("Driver table: no merged material diff lines (same split as owner summary and export).")
            else:
                key_outputs.append(
                    f"Driver table: {n_top} Top drivers, {n_sm} Smaller moves (same split as owner summary and export)."
                )
            key_outputs.append(
                "Top drivers use the same ranked driver table as the owner summary and export."
            )
        if labor_deltas:
            key_outputs.append(f"Labor delta rows: {len(labor_deltas)}")
        suggestions = [
            "Open `outputs/operator_envelope.json` for the deterministic narrative.",
            "Use `Cost vs revenue` to identify top pressure direction.",
        ]
        if contract == "compare_and_show_labor_deltas":
            suggestions.insert(0, "Review the largest labor deltas first, then confirm with workbook sheet preview.")
        if contract == "compare_latest_report":
            suggestions.insert(0, "Use the full operator output block to brief owners and reviewers.")
        if contract == "run_weekly_review":
            suggestions.insert(0, "This weekly review used the same pairing as `Compare` (no model); owner summary and export are bundled below.")
        owner_summary: dict[str, Any] | None = None
        if contract == "run_weekly_review" and run_id:
            try:
                owner_summary = _summarize_owner_from_run(run_id)
            except Exception:  # noqa: BLE001
                owner_summary = None
        if contract == "run_weekly_review" and not owner_summary and run_id:
            suggestions.append("Owner summary could not be built from the run; try `Summarize for owner` after the run is visible in history.")
        export_payload = None
        if contract == "run_weekly_review":
            export_payload = _export_top_changes_csv(run_id)
            if export_payload.get("skipped"):
                suggestions.append("Top changes export skipped: no merged material diff lines for this run.")
            else:
                suggestions.append("Top changes CSV (driver table order) was generated for this weekly review.")
            suggestions.append(
                "Next: copy the owner summary if shown, then open the CSV and artifact links for the weekly pack."
            )
        weekly_found: dict[str, Any] = {
            "run_id": run_id,
            "summary": compare_summary,
            "what_to_review": list(compare.get("what_to_review") or []),
            "artifacts": artifact_paths,
            "labor_deltas": labor_deltas,
            "selected_pair": selected_pair,
            "pairing_confidence": pairing.get("pairing_confidence", 0),
            "workflow_output": workflow_output,
        }
        if contract == "run_weekly_review":
            if owner_summary and isinstance(owner_summary, dict):
                weekly_found["owner_summary"] = owner_summary
            sd0 = so.get("summary_deltas", {}) if isinstance(so, dict) else {}
            if isinstance(sd0, dict):
                weekly_found["summary_deltas"] = {
                    "profit": float(sd0.get("profit", 0) or 0),
                    "revenue": float(sd0.get("revenue", 0) or 0),
                    "cost": float(sd0.get("cost", 0) or 0),
                }
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "workspace_root": str(workspace.root)},
            "did": {
                "execution": "ran deterministic compare wrapper"
                + ("; then owner summary + top-changes export" if contract == "run_weekly_review" else ""),
                "files_used": {"prior": plan["prior_path"], "current": plan["current_path"]},
                "compare_mode": {
                    "selected_mode": plan.get("selected_mode"),
                    "compare_path": compare.get("compare_path"),
                    "selection_reason": compare.get("compare_path_reason"),
                },
                "pairing_reasoning": {
                    "pairing_confidence": pairing.get("pairing_confidence", 0),
                    "selected_pair_id": selected_pair.get("pair_id", ""),
                    "ranking_factors": selected_pair.get("ranking_factors", []),
                    "candidate_pairs": pairing.get("pair_candidates", []),
                    "overrides_applied": pairing.get("overrides_applied", {}),
                },
            },
            "found": weekly_found,
            "next_steps": suggestions,
            "summary_card": {
                "title": "Weekly review" if contract == "run_weekly_review" else "Compare Complete",
                "summary": (
                    (compare_summary[0] if compare_summary else "Compare completed with deterministic outputs.")
                    + (
                        " Owner summary and export are in this result (same contracts as the standalone actions)."
                        if contract == "run_weekly_review"
                        else ""
                    )
                ),
                "paths": [plan["prior_path"], plan["current_path"], *artifact_paths[:4]],
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
            "run_payload": compare,
            "weekly_export": export_payload if contract == "run_weekly_review" else None,
            "memory_update": {
                "last_confirmed_pair": {
                    "prior_path": plan["prior_path"],
                    "current_path": plan["current_path"],
                    "pair_id": selected_pair.get("pair_id", ""),
                },
                "preferred_report_family": str(
                    (selected_pair.get("current") or {}).get("report_family")
                    or (selected_pair.get("prior") or {}).get("report_family")
                    or ""
                ),
            },
        }
        if contract != "run_weekly_review":
            result_payload["found"] = {
                "run_id": run_id,
                "summary": compare_summary,
                "what_to_review": list(compare.get("what_to_review") or []),
                "artifacts": artifact_paths,
                "labor_deltas": labor_deltas,
                "selected_pair": selected_pair,
                "pairing_confidence": pairing.get("pairing_confidence", 0),
                "workflow_output": workflow_output,
            }
            result_payload["summary_card"] = {
                "title": "Compare Complete",
                "summary": compare_summary[0] if compare_summary else "Compare completed with deterministic outputs.",
                "paths": [plan["prior_path"], plan["current_path"], *artifact_paths[:4]],
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            }
            result_payload["weekly_export"] = None
        answer = (
            "Compare completed with labor deltas extracted."
            if contract == "compare_and_show_labor_deltas"
            else ("Weekly review completed." if contract == "run_weekly_review" else "Compare completed successfully.")
        )
        paths = [plan["prior_path"], plan["current_path"], *artifact_paths[:4]]
        if export_payload and isinstance(export_payload, dict):
            if export_payload.get("skipped"):
                key_outputs.append(f"Top changes export: {export_payload.get('message', 'skipped')}")
            elif export_payload.get("path"):
                paths.append(str(export_payload.get("path", "")))
                key_outputs.append(
                    f"Weekly export: {export_payload.get('path', '')} "
                    f"({export_payload.get('rows_written', export_payload.get('row_count', 0))} rows; "
                    f"{export_payload.get('top_driver_count', 0)} Top, {export_payload.get('smaller_move_count', 0)} smaller)"
                )
        if workflow_output.get("confidence"):
            key_outputs.append(f"Confidence: {workflow_output.get('confidence')}")
        if workflow_output.get("cost_vs_revenue") and isinstance(workflow_output.get("cost_vs_revenue"), dict):
            key_outputs.append(f"Cost vs revenue: {workflow_output['cost_vs_revenue'].get('signal', 'unknown')}")
        result_payload["summary_card"]["key_outputs"] = key_outputs
        result_payload["source_line"] = _workspace_source_line_for_transcript(
            index_backed=_index_workbook_count_for_root(str(workspace.root)) > 0
        )
        result_payload["provenance"] = result_payload["source_line"]
        result_payload["project_filter"] = str(
            pairing.get("project_filter") or _scoping_project_id_from_query(query) or ""
        )
        _persist_compare_memory_to_config(
            selected_pair if isinstance(selected_pair, dict) else {},
            plan if isinstance(plan, dict) else {},
        )
    elif contract == "summarize_for_owner":
        run_id = str(task_payload.get("run_id") or "")
        if not run_id:
            raise HTTPException(status_code=400, detail="No completed run is available for owner summary.")
        owner = _summarize_owner_from_run(run_id)
        answer = "Prepared deterministic owner summary (copy block in the result card) from the latest completed run."
        suggestions = ["Share the copy-ready block with the owner, and export top changes if you need a CSV attachment."]
        paths = [f"runs/{run_id}/outputs/operator_envelope.json", f"runs/{run_id}/outputs/structured_output.json"]
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "run_id": run_id},
            "did": {
                "method": "Merged material lines, same primary vs Smaller moves split as the driver table; no LLM.",
            },
            "found": {"owner_summary": owner},
            "next_steps": suggestions,
            "summary_card": {
                "title": "Source files",
                "paths": paths,
            },
        }
    elif contract == "export_top_changes":
        run_id = str(task_payload.get("run_id") or "")
        if not run_id:
            raise HTTPException(status_code=400, detail="No completed run is available for top changes export.")
        export_info = _export_top_changes_csv(run_id)
        if export_info.get("skipped"):
            answer = "Export not needed: no merged material diff lines for this run."
            suggestions = ["Run a financial compare with material line output, or confirm structured_output has material diff lines."]
        else:
            answer = f"Export completed: driver-table-ordered CSV for run {run_id}."
            suggestions = ["Open the CSV and share it with weekly review stakeholders."]
        paths = [str(export_info.get("path", ""))] if export_info.get("path") else []
        key_outputs = []
        if export_info.get("skipped"):
            key_outputs.append(export_info.get("message", "No rows to export."))
        else:
            key_outputs.append(
                f"Rows: {export_info.get('rows_written', 0)} written"
                f" (Top drivers: {export_info.get('top_driver_count', 0)}, Smaller moves: {export_info.get('smaller_move_count', 0)})"
            )
            if export_info.get("truncated"):
                key_outputs.append(f"Note: output capped at {export_info.get('rows_written', 0)} rows (max reached).")
        if export_info.get("grounding"):
            key_outputs.append(str(export_info.get("grounding", "")))
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "run_id": run_id},
            "project_filter": str(_extract_project_id_from_query(query) or ""),
            "did": {
                "method": "merge/dedupe material diff + same Top vs Smaller split as driver table; CSV export",
            },
            "found": {"export_csv": export_info},
            "next_steps": suggestions,
            "summary_card": {
                "title": "Top Changes Export",
                "summary": (
                    answer
                    if export_info.get("skipped")
                    else f"CSV ready with {export_info.get('rows_written', 0)} row(s) (Top drivers + Smaller moves order)."
                ),
                "paths": paths,
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
        }
    elif contract == "assess_cost_vs_revenue":
        run_id = str(task_payload.get("run_id") or "")
        if not run_id:
            raise HTTPException(status_code=400, detail="No completed run is available for cost vs revenue assessment.")
        signal = _run_cost_vs_revenue_signal(run_id)
        verdict_map = {
            "cost-led": "Cost movement is currently the bigger issue.",
            "revenue-led": "Revenue movement is currently the bigger issue.",
            "balanced": "Revenue and cost movement are currently balanced.",
        }
        answer = verdict_map.get(str(signal.get("signal")), "Cost/revenue signal calculated.")
        suggestions = ["If cost-led, inspect labor/material deltas.", "If revenue-led, inspect top revenue line categories."]
        paths = [f"runs/{run_id}/outputs/structured_output.json"]
        key_outputs = [
            f"Revenue delta: {signal.get('revenue_delta', 0)}",
            f"Cost delta: {signal.get('cost_delta', 0)}",
            f"Profit delta: {signal.get('profit_delta', 0)}",
        ]
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "run_id": run_id},
            "did": {"method": "summary_deltas then material diff fallback"},
            "found": signal,
            "next_steps": suggestions,
            "summary_card": {
                "title": "Cost vs Revenue Signal",
                "summary": answer,
                "paths": paths,
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
        }
    elif contract == "list_current_run_artifacts":
        run_id = str(task_payload.get("run_id") or "")
        if not run_id:
            raise HTTPException(status_code=400, detail="No completed run is available for artifact listing.")
        artifacts = _artifacts_for_run(run_id)
        answer = f"Listed current run artifacts for {run_id}."
        suggestions = ["Open structured output first, then envelope, then intake/debug files as needed."]
        paths = list(artifacts[:10])
        key_outputs = [f"Artifact count: {len(artifacts)}"]
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "run_id": run_id},
            "did": {"method": "read run directory inputs/outputs file list"},
            "found": {"run_id": run_id, "artifacts": artifacts},
            "next_steps": suggestions,
            "summary_card": {
                "title": "Run Artifacts",
                "summary": f"{len(artifacts)} artifacts available for inspection.",
                "paths": paths,
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
        }
    elif contract == "inspect_workbook":
        workbook_path = str(task_payload.get("workbook_path") or "")
        if not workbook_path:
            raise HTTPException(status_code=400, detail="No workbook selected for inspection.")
        meta = _workbook_metadata(workspace.root, workbook_path)
        answer = f"Workbook inspection complete for {meta.get('name', workbook_path)}."
        suggestions = ["Use `Find report sheets` next, then `Preview report sheet` on a likely target."]
        paths = [str(meta.get("path", workbook_path))]
        key_outputs = [f"Sheet count: {meta.get('sheet_count', 0)}", f"Workbook size bytes: {meta.get('size_bytes', 0)}"]
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "workbook_path": workbook_path},
            "did": {"method": "read-only workbook metadata scan"},
            "found": meta,
            "next_steps": suggestions,
            "summary_card": {
                "title": "Workbook Inspection",
                "summary": f"{meta.get('sheet_count', 0)} sheets discovered.",
                "paths": paths,
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
        }
    elif contract == "preview_report_sheet":
        workbook_path = str(task_payload.get("workbook_path") or "")
        if not workbook_path:
            raise HTTPException(status_code=400, detail="No workbook selected for sheet preview.")
        preview = _preview_workbook_sheet(
            workspace.root,
            workbook_path=workbook_path,
            sheet_name=str(task_payload.get("sheet_name") or ""),
        )
        answer = f"Preview ready for sheet {preview.get('sheet_name', 'unknown')}."
        suggestions = ["Use this preview to confirm row labels before compare interpretation."]
        paths = [str(preview.get("path", workbook_path))]
        key_outputs = [
            f"Sheet: {preview.get('sheet_name', '')}",
            f"Preview rows: {preview.get('preview_row_count', 0)}",
        ]
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "workbook_path": workbook_path, "sheet_name": task_payload.get("sheet_name", "")},
            "did": {"method": "bounded read-only cell preview"},
            "found": preview,
            "next_steps": suggestions,
            "summary_card": {
                "title": "Report Sheet Preview",
                "summary": f"Loaded {preview.get('preview_row_count', 0)} preview rows.",
                "paths": paths,
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
        }
    elif contract == "find_report_sheets":
        workbook_path = str(task_payload.get("workbook_path") or "")
        if not workbook_path:
            raise HTTPException(status_code=400, detail="No workbook selected for report-sheet detection.")
        report_sheets = _find_report_sheets(workspace.root, workbook_path)
        answer = f"Found likely report sheets for {report_sheets.get('path', workbook_path)}."
        suggestions = ["Preview the top scored sheet and verify revenue/cost/profit headers."]
        paths = [str(report_sheets.get("path", workbook_path))]
        key_outputs = [f"Likely report sheets: {len(report_sheets.get('report_sheets', []))}"]
        result_payload = {
            "contract": contract,
            "requested": {"query": query, "workbook_path": workbook_path},
            "did": {"method": "sheet-name and header keyword scoring"},
            "found": report_sheets,
            "next_steps": suggestions,
            "summary_card": {
                "title": "Likely Report Sheets",
                "summary": f"{len(report_sheets.get('report_sheets', []))} likely sheets detected.",
                "paths": paths,
                "key_outputs": key_outputs,
                "suggestions": suggestions,
            },
        }
    else:
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="failed",
            answer="This action is not implemented in the current operator build.",
            result={"next_steps": [f"Try: {SUGGESTED_FIRST_COMPARE}"]},
            trace=trace
            + [
                {
                    "action": "task.route",
                    "outcome": "unimplemented",
                    "contract": contract,
                }
            ],
        )

    if contract == "find_latest_prior_reports" and not result_payload.get("found", {}).get("pair_ready", True):
        return _assistant_task_envelope(
            query=query,
            workspace_root=str(workspace.root),
            contract=contract,
            status="no_candidates",
            answer="No before/after pair could be selected from the workbooks in this workspace.",
            result=result_payload,
            trace=trace,
        )

    return _assistant_task_envelope(
        query=query,
        workspace_root=str(workspace.root),
        contract=contract,
        status="completed",
        answer=answer,
        result=result_payload,
        trace=trace,
        policy={"requires_confirmation": compare_contract},
    )


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "title": APP_TITLE,
            "run_history": list(reversed(_load_history())),
            "product_oneliner": "Financial Report Builder — select projects and monthly workbooks, run deterministic financial signals, and read results in structured panels (traceable to workbook sources).",
            "driver_table_primary_n": OPERATOR_UI_DRIVER_TABLE_PRIMARY_N,
            "driver_table_impact_usd": OPERATOR_UI_DRIVER_TABLE_IMPACT_USD,
        },
    )


@app.post("/runs")
def create_run(workflow_mode: str = Form("auto")) -> JSONResponse:
    run_id = _run_id()
    state = SessionState(
        run_id=run_id,
        status="created",
        workflow_mode=workflow_mode,
        workflow_name=None,
        uploaded_files=[],
        created_at=_utc_now_iso(),
    )
    _save_state(state)
    return JSONResponse({"run_id": run_id, "status": state.status})


@app.post("/runs/{run_id}/files")
def upload_files(
    run_id: str,
    prior_file: UploadFile | None = File(None),
    current_file: UploadFile | None = File(None),
    source_file: UploadFile | None = File(None),
) -> JSONResponse:
    state = _load_state(run_id)
    uploads: list[dict[str, Any]] = []
    if prior_file:
        uploads.append(_store_upload(run_id, "prior", prior_file))
    if current_file:
        uploads.append(_store_upload(run_id, "current", current_file))
    if source_file:
        uploads.append(_store_upload(run_id, "source", source_file))
    if not uploads:
        raise HTTPException(status_code=400, detail="No files uploaded.")
    state.uploaded_files = uploads
    state.status = "files_uploaded"
    _save_state(state)
    return JSONResponse({"run_id": run_id, "uploaded_files": uploads, "status": state.status})


@app.post("/runs/{run_id}/execute")
def execute_run(run_id: str) -> JSONResponse:
    state = _load_state(run_id)
    if not state.uploaded_files:
        raise HTTPException(status_code=400, detail="Upload files before execute.")

    workflow_name = _detect_workflow(state.workflow_mode, state.uploaded_files)
    phase = _workflow_phase(workflow_name)
    run_dir = RUNS_ROOT / run_id
    state.workflow_name = workflow_name
    state.status = "running"
    _save_state(state)

    try:
        _run_checked(
            [
                "python3",
                str(SCRIPT_INIT),
                "--run-id",
                run_id,
                "--runs-root",
                str(RUNS_ROOT),
                "--workflow-name",
                workflow_name,
                "--phase",
                phase,
                "--workflow-registry",
                str(REGISTRY_PATH),
                "--force",
            ]
        )

        workflow_inputs = _prepare_workflow_inputs(run_id, workflow_name, state.uploaded_files)
        cmd = [
            "python3",
            str(SCRIPT_RUN),
            "--workflow",
            workflow_name,
            "--run-dir",
            str(run_dir),
            "--registry",
            str(REGISTRY_PATH),
            "--force",
        ]
        for p in workflow_inputs:
            cmd.extend(["--input", str(p.resolve())])
        _run_checked(cmd)
        _run_checked(
            [
                "python3",
                str(SCRIPT_VALIDATE),
                "--run-dir",
                str(run_dir),
                "--workflow-registry",
                str(REGISTRY_PATH),
            ]
        )
    except Exception as exc:  # noqa: BLE001
        _uvicorn_log.exception(
            "[operator_local_ui] execute_run failed run_id=%s workflow=%s",
            run_id,
            workflow_name,
        )
        state.status = "failed"
        state.error = str(exc)
        state.finished_at = _utc_now_iso()
        state.run_dir = str(run_dir.resolve())
        _save_state(state)
        _append_history(
            {
                "timestamp": state.finished_at,
                "run_id": run_id,
                "workflow": workflow_name,
                "status": "FAIL",
                "error": state.error,
            }
        )
        raise HTTPException(status_code=500, detail=str(exc))

    envelope_path = run_dir / "outputs" / "operator_envelope.json"
    structured_path = run_dir / "outputs" / "structured_output.json"
    if not envelope_path.exists():
        raise HTTPException(status_code=500, detail="Missing operator_envelope.json after run.")
    envelope = _load_json(envelope_path)
    structured_output = _load_json(structured_path) if structured_path.exists() else {}

    state.status = "completed"
    state.finished_at = _utc_now_iso()
    state.run_dir = str(run_dir.resolve())
    state.outputs_dir = str((run_dir / "outputs").resolve())
    state.envelope = envelope
    state.structured_output = structured_output
    _save_state(state)
    _append_history(
        {
            "timestamp": state.finished_at,
            "run_id": run_id,
            "workflow": workflow_name,
            "status": "PASS",
        }
    )
    assistant_view = _assistant_view(
        state.workflow_name,
        envelope,
        structured_output,
        state.uploaded_files,
    )
    return JSONResponse(
        {
            "run_id": run_id,
            "status": state.status,
            "workflow": workflow_name,
            "summary": envelope.get("what_i_found", []),
            "what_to_review": envelope.get("what_needs_review", []),
            "structured_output": structured_output,
            "assistant_view": assistant_view,
            "financial_intake_artifacts": _financial_intake_artifact_entries(run_dir),
            "artifacts": _artifact_relative_paths(run_dir),
        }
    )


@app.get("/api/history")
def api_history() -> JSONResponse:
    return JSONResponse(list(reversed(_load_history())))


@app.get("/runs/{run_id}")
def get_run(run_id: str) -> JSONResponse:
    state = _load_state(run_id)
    artifacts: list[str] = []
    financial_intake: list[dict[str, str]] = []
    assistant_view: dict[str, Any] | None = None
    if state.run_dir:
        rd = Path(state.run_dir)
        artifacts = _artifact_relative_paths(rd)
        financial_intake = _financial_intake_artifact_entries(rd)
    if state.status == "completed" and state.envelope and state.structured_output is not None:
        assistant_view = _assistant_view(
            state.workflow_name,
            state.envelope,
            state.structured_output,
            state.uploaded_files,
        )
    payload: dict[str, Any] = {**state.to_json(), "artifacts": artifacts, "financial_intake_artifacts": financial_intake}
    if assistant_view is not None:
        payload["assistant_view"] = assistant_view
    return JSONResponse(payload)


@app.get("/runs/{run_id}/artifacts")
def list_artifacts(run_id: str) -> JSONResponse:
    state = _load_state(run_id)
    if not state.run_dir:
        return JSONResponse({"run_id": run_id, "artifacts": []})
    artifacts = _artifact_relative_paths(Path(state.run_dir))
    return JSONResponse({"run_id": run_id, "artifacts": artifacts})


@app.get("/runs/{run_id}/artifacts/{artifact_path:path}")
def download_artifact(run_id: str, artifact_path: str) -> FileResponse:
    state = _load_state(run_id)
    if not state.run_dir:
        raise HTTPException(status_code=404, detail="Run has no artifact directory.")
    run_dir = Path(state.run_dir).resolve()
    target = (run_dir / artifact_path).resolve()
    if run_dir not in target.parents and target != run_dir:
        raise HTTPException(status_code=400, detail="Invalid artifact path.")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail=f"Artifact not found: {artifact_path}")
    return FileResponse(str(target), filename=target.name)
