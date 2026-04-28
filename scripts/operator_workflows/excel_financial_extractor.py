from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from operator_workflows.project_id_utils import extract_project_id_from_rel_and_name

SNAPSHOT_SCHEMA = "financial_extract_v1"

# Components summed into workbook `total_projected_profit` (order matches the operational formula).
TOTAL_PROJECTED_PROFIT_FORMULA_TERMS = (
    "cm_fee",
    "buyout_savings_realized",
    "budget_savings_overages",
    "pco_profit",
    "labor_rate_profit_to_date",
    "prior_system_profit",
)

# JTD footer / cost-code extraction (219128 Profit Report family). See unit tests and docstrings.
_ORIGINAL_CONTRACT_HINTS = ("current contract", "payment app")
_REVISED_BUDGET_TOTAL_HINTS = ("revised current budget", "total project costs")
_BUDGET_VARIANCE_HINTS = ("budget overage", "budget savings")
_BUYOUT_REALIZED_HINTS = ("buyout savings realized",)
# Profit Calc footer block (219128): "Actual profit to be claimed in old job" (typos: clamied, claimied).

_LABEL_RE = re.compile(r"\s+")
# Cost line first column, e.g. 01.030 (construction-style cost code)
_CODE_LIKE = re.compile(r"^\d{1,3}\.\d{1,3}$")
# 18-004, 18.004, N18.100 (CO / PCO style not matched by the strict d.d.d above)
_JTD_CO_LOOSE = re.compile(r"^N?(?:[12]\d|30|40)[-.]\d+", re.IGNORECASE)

# Order matters: prefer operational / current period columns for JTD-style sheets.
_AMOUNT_HEADER_ORDER = (
    "update current budget",
    "spent to date",
    "current budget",
    "committed to date",
    "billed to date",
    "actual cost",
    "open commitments",
    "budget less committed",
    "original budget",
    "variance",
    "amount",
    "value",
    "total",
)


def _norm_label(value: str) -> str:
    return _LABEL_RE.sub(" ", value.strip().lower())


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    neg = False
    if text.startswith("(") and text.endswith(")"):
        neg = True
        text = text[1:-1]
    text = text.replace("$", "").replace(",", "").replace("%", "")
    if text.endswith("-"):
        neg = True
        text = text[:-1]
    try:
        num = float(text)
        return -num if neg else num
    except ValueError:
        return None


def _norm_cost_code(cell: Any) -> str:
    s = str(cell or "").strip().upper().replace(" ", "")
    if s.startswith("N") and len(s) > 1 and s[1:2].isdigit():
        s = s[1:]
    return s


def _jtd_norm_family_segment(norm: str) -> str:
    """Leading numeric segment of a normalized cost code (e.g. '30.000' -> '30')."""
    n = str(norm or "").strip().upper().replace(" ", "")
    if not n:
        return ""
    head = n.split(".")[0]
    return "".join(ch for ch in head if ch.isdigit())


def _cost_code_namespace_from_raw(raw_code: str) -> str:
    """
    Namespace follows the raw Excel first-column code, not display_cost_code.
    Raw starting with N + digit => extended (n_prefixed); otherwise numeric (original structure).
    """
    u = str(raw_code or "").strip().upper().replace(" ", "")
    if u.startswith("N") and len(u) > 1 and u[1].isdigit():
        return "n_prefixed"
    return "numeric"


def _jtd_component_role(
    norm: str,
    cost_code_name: str,
    raw_code: str,
    project_id: str,
) -> str | None:
    """Semantic profit / CO role for a JTD line; None when not a classified component row."""
    ncode = str(norm or "").strip().upper().replace(" ", "")
    name_raw = str(cost_code_name or "")
    raw = str(raw_code or "").strip()
    if ncode in ("30.000",):
        if is_prior_system_jtd_cost_name(name_raw):
            return "prior_system_profit"
        if is_cm_fee_jtd_cost_name(name_raw):
            return "cm_fee"
        return None
    if ncode in ("40.000",):
        if is_pco_profit_jtd_cost_name(name_raw):
            return "pco_profit"
        return None
    if is_owner_change_order_code(raw, project_id):
        return "owner_change_order"
    if is_cm_change_order_code(raw, project_id):
        return "cm_change_order"
    return None


def _jtd_row_semantic_fields(raw_code: str, norm: str, cost_code_name: str, project_id: str) -> dict[str, Any]:
    return {
        "cost_code_family": _jtd_norm_family_segment(norm),
        "cost_code_namespace": _cost_code_namespace_from_raw(raw_code),
        "component_role": _jtd_component_role(norm, cost_code_name, raw_code, project_id),
    }


def _jtd_namespace_budget_totals(rows: list[Any] | None) -> tuple[float, float]:
    """Sum Update Current Budget by cost_code_namespace (numeric vs n_prefixed)."""
    orig = 0.0
    ext = 0.0
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        ns = r.get("cost_code_namespace")
        v = r.get("update_current_budget")
        if v is None:
            continue
        try:
            fv = float(v)
        except (TypeError, ValueError):
            continue
        if ns == "numeric":
            orig += fv
        elif ns == "n_prefixed":
            ext += fv
    return orig, ext


def _display_cost_code_workbook(raw_code: str, norm: str, project_id: str) -> str:
    """
    UI / workbench display string. For project 219128 only, plain 18/21/30/40-family codes
    are shown with an N prefix to match CMiC-style labeling; raw Excel text is unchanged
    in raw_cost_code. Other projects: display matches Excel (no synthetic N).
    """
    raw = str(raw_code or "").strip()
    if not _project_id_is_219128(project_id):
        return raw
    u = raw.upper().replace(" ", "")
    if u.startswith("N") and len(u) > 1 and u[1].isdigit():
        return raw
    fam = _jtd_norm_family_segment(norm)
    if fam in ("18", "21", "30", "40"):
        return "N" + str(norm or "").strip().upper().replace(" ", "")
    return raw


def _jtd_row_code_fields(raw_code: str, project_id: str) -> dict[str, str]:
    raw = str(raw_code or "").strip()
    norm = _norm_cost_code(raw_code)
    disp = _display_cost_code_workbook(raw, norm, project_id)
    return {
        "raw_cost_code": raw,
        "normalized_cost_code": norm,
        "display_cost_code": disp,
        "cost_code": disp,
    }


def _project_id_is_219128(project_id: str) -> bool:
    return str(project_id or "").strip() == "219128"


def is_owner_change_order_code(raw_code: str, project_id: str) -> bool:
    """
    Owner CO lines: cost_code starts with 18, or (project 219128 only) N18… legacy codes.
    N-prefix rule must NOT apply to other jobs.
    """
    u = str(raw_code or "").strip().upper().replace(" ", "")
    norm = _norm_cost_code(raw_code)
    legacy_n18 = u.startswith("N18")
    if _project_id_is_219128(project_id):
        return bool(norm.startswith("18") or legacy_n18)
    if legacy_n18:
        return False
    return bool(norm.startswith("18"))


def is_cm_change_order_code(raw_code: str, project_id: str) -> bool:
    """
    CM CO lines: cost_code starts with 21, or (project 219128 only) N21… legacy codes.
    """
    u = str(raw_code or "").strip().upper().replace(" ", "")
    norm = _norm_cost_code(raw_code)
    legacy_n21 = u.startswith("N21")
    if _project_id_is_219128(project_id):
        return bool(norm.startswith("21") or legacy_n21)
    if legacy_n21:
        return False
    return bool(norm.startswith("21"))


def is_prior_system_jtd_cost_name(cost_code_name: str) -> bool:
    """
    Label-aware: prior-system / old job profit in the JTD cost table (e.g. N30.000 + ORIG. PROFIT CLAIMED).
    Must be checked before CM Fee heuristics (some labels contain the substring 'prof' or numeric noise).
    """
    n = _norm_label(str(cost_code_name or ""))
    if not n:
        return False
    if "orig" in n and "profit" in n and "claim" in n:
        return True
    if "original profit" in n and "claim" in n:
        return True
    if "actual profit" in n and "old job" in n:
        return True
    if "old job" in n and "profit" in n:
        return True
    if n in {"orig. profit claimed", "original profit claimed", "orig profit claimed"}:
        return True
    return False


def is_cm_fee_jtd_cost_name(cost_code_name: str) -> bool:
    n = _norm_label(str(cost_code_name or ""))
    if not n:
        return False
    if is_prior_system_jtd_cost_name(cost_code_name):
        return False
    if "pco" in n and "profit" in n:
        return False
    if "n60" in n or "n 60" in n or (n.startswith("n60") and "unreal" in n):
        return False
    if re.search(r"\bcm fee\b", n):
        return True
    if "construction management fee" in n or " const mgmt fee" in n or " const management fee" in n:
        return True
    if re.search(r"\bfee\b", n) and "unreal" not in n and "n60" not in n and "n 60" not in n:
        if "provisional" in n and "profit" in n and "pco" not in n:
            return False
        if "parking fee" in n or "permit" in n:
            return False
        return True
    return False


def is_pco_profit_jtd_cost_name(cost_code_name: str) -> bool:
    n = _norm_label(str(cost_code_name or ""))
    if not n:
        return False
    return "pco" in n and "profit" in n


def infer_project_id_from_workbook_path(path: Path) -> str:
    """Numeric job id from full path and filename (aligned with operator UI index rules)."""
    rel = str(path).replace("\\", "/")
    return extract_project_id_from_rel_and_name(rel, path.name)


def _header_original_budget_col(header_row: list[Any]) -> int:
    for j, cell in enumerate(header_row or []):
        lab = _norm_label(str(cell) if cell is not None else "")
        if "original" in lab and "budget" in lab:
            return j
    return -1


def _jtd_data_row_number(header_i: int, offset_in_data_block: int) -> int:
    """0-based header row + offset within data block → 1-based Excel row index."""
    return header_i + 2 + offset_in_data_block


def _collect_jtd_table_and_change_order_rows(
    rows: list[list[Any]],
    header_i: int,
    f_idx: int,
    project_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    All contiguous JTD cost-code lines until first blank in column A (same boundary as main extract).
    change_order rows are the subset of owner/CM CO cost codes.
    """
    if header_i < 0 or f_idx < 0 or header_i >= len(rows):
        return [], []
    header_row = rows[header_i] if len(rows) > header_i else []
    ob_idx = _header_original_budget_col(header_row)
    code_rows: list[dict[str, Any]] = []
    co_rows: list[dict[str, Any]] = []
    ccode_i, name_i = 0, 1
    for k, row in enumerate(rows[header_i + 1 :]):
        if not row or len(row) <= f_idx:
            break
        if row[0] is None or str(row[0]).strip() == "":
            break
        if not _row_looks_like_jtd_cost_line(row, project_id):
            continue
        excel_row = _jtd_data_row_number(header_i, k)
        raw_code = str(row[ccode_i] or "").strip()
        cname = str(row[name_i] if len(row) > name_i else "").strip()
        ucb = _to_float(row[f_idx] if len(row) > f_idx else None)
        if ucb is None:
            continue
        ob = _to_float(row[ob_idx] if ob_idx >= 0 and len(row) > ob_idx else None) if ob_idx >= 0 else None
        code_fields = _jtd_row_code_fields(raw_code, project_id)
        sem = _jtd_row_semantic_fields(raw_code, code_fields["normalized_cost_code"], cname, project_id)
        code_rows.append(
            {
                **code_fields,
                **sem,
                "cost_code_name": cname,
                "update_current_budget": float(ucb),
                "original_budget": float(ob) if ob is not None else None,
                "excel_row": excel_row,
            }
        )
        if is_owner_change_order_code(raw_code, project_id):
            co_rows.append(
                {
                    "co_type": "Owner",
                    **code_fields,
                    **sem,
                    "cost_code_name": cname,
                    "update_current_budget": float(ucb),
                    "excel_row": excel_row,
                }
            )
        if is_cm_change_order_code(raw_code, project_id):
            co_rows.append(
                {
                    "co_type": "CM",
                    **code_fields,
                    **sem,
                    "cost_code_name": cname,
                    "update_current_budget": float(ucb),
                    "excel_row": excel_row,
                }
            )
    return code_rows, co_rows


def _aggregate_change_order_lines(
    rows: list[list[Any]],
    header_i: int,
    f_idx: int,
    project_id: str,
) -> dict[str, Any]:
    """Sum Update Current Budget (column f_idx) for owner (18) and CM (21) change-order lines."""
    co_notes: list[str] = [
        f"Change order values: sum of 'Update Current Budget' using detected column index {f_idx} (0-based; Excel column F when index is 5).",
    ]
    if _project_id_is_219128(project_id):
        co_notes.append(
            "Project 219128 includes legacy N18/N21 alphanumeric code prefixes. "
            "Other projects use numeric 18/21 prefixes only; N18/N21 raw prefixes are excluded outside 219128."
        )
    else:
        co_notes.append(
            "N18/N21 raw prefixes are excluded for this project; only numeric 18/21-mapped cost codes count."
        )
    owner_co_count = 0
    owner_co_value = 0.0
    cm_co_count = 0
    cm_co_value = 0.0
    ccode_i = 0
    for row in rows[header_i + 1 :]:
        if not row or len(row) <= f_idx:
            continue
        if row[0] is None or str(row[0]).strip() == "":
            break
        if not _row_looks_like_jtd_cost_line(row, project_id):
            continue
        raw = str(row[ccode_i] or "").strip()
        v = _to_float(row[f_idx] if len(row) > f_idx else None)
        if v is None:
            continue
        if is_owner_change_order_code(raw, project_id):
            owner_co_count += 1
            owner_co_value += float(v)
        if is_cm_change_order_code(raw, project_id):
            cm_co_count += 1
            cm_co_value += float(v)
    return {
        "owner_change_orders_count": owner_co_count,
        "owner_change_orders_value": owner_co_value,
        "cm_change_orders_count": cm_co_count,
        "cm_change_orders_value": cm_co_value,
        "change_order_source_notes": co_notes,
    }


def _row_looks_like_jtd_cost_line(row: list[Any], project_id: str = "") -> bool:
    a = row[0] if len(row) > 0 else None
    if a is None:
        return False
    s = str(a).strip()
    if not s or s.lower() in {"total", "grand total", "subtotal"}:
        return False
    if _CODE_LIKE.match(s) or (s.upper().replace(" ", "").startswith("N") and any(c.isdigit() for c in s)):
        return True
    if _JTD_CO_LOOSE.match(s.strip()):
        return True
    if is_owner_change_order_code(s, project_id) or is_cm_change_order_code(s, project_id):
        return True
    return False


def _find_lbr_sheet_title(wb: Any) -> str | None:
    for ws in wb.worksheets:
        nt = _norm_label(ws.title)
        if "lbr" in nt and "jtd" not in nt:
            return ws.title
    return None


def _find_primary_jtd_sheet_title(wb: Any) -> str | None:
    for ws in wb.worksheets:
        nt = _norm_label(ws.title)
        if "jtd" in nt and "static" not in nt and "mom" not in nt:
            return ws.title
    return None


def _scan_lbr_workbook_rows(rows: list[list[Any]], max_line: int = 80) -> list[dict[str, Any]]:
    """
    Expose LBR tab lines for the workbench (no new math; label + numeric values per row).
    """
    out: list[dict[str, Any]] = []
    for i, row in enumerate(rows[:max_line], start=1):
        if not row:
            continue
        cells = [row[j] if j < len(row) else None for j in range(min(len(row), 12))]
        if not any(c not in (None, "") for c in cells):
            continue
        label = str(cells[0] if len(cells) > 0 else "").strip()
        if not label and len(cells) > 1:
            label = str(cells[1] or "").strip()
        if not label:
            continue
        nums: list[float] = []
        for c in cells[1:]:
            v = _to_float(c)
            if v is not None:
                nums.append(v)
        low = _norm_label(label)
        out.append(
            {
                "excel_row": i,
                "label": str(cells[0] or "").strip() or label,
                "values": nums,
                "raw": [cells[j] for j in range(min(8, len(cells)))],
                "keyword_labor": any(
                    k in low
                    for k in (
                        "labor",
                        "billable",
                        "cmic",
                        "billed",
                        "spent to date",
                        "old job",
                        "labor rate",
                    )
                ),
            }
        )
    return out


def _lbr_keyword_detail_rows(lbr_block: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Subset of LBR scan rows whose labels look like the extended breakdown lines."""
    scan = lbr_block.get("lbr_workbook_rows") if isinstance(lbr_block, dict) else None
    if not isinstance(scan, list):
        return []
    out: list[dict[str, Any]] = []
    for row in scan:
        if not isinstance(row, dict):
            continue
        if not row.get("keyword_labor"):
            continue
        out.append(
            {
                "excel_row": row.get("excel_row"),
                "label": row.get("label"),
                "values": row.get("values") or [],
            }
        )
    return out


def _lrp_presentation_block(lbr: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Headline LBR values already used in formulas; deterministic one-line notes."""
    if not lbr:
        return []
    b, a, lrp = lbr.get("lrp_billed_to_date"), lbr.get("lrp_actual_cost"), lbr.get("labor_rate_profit_to_date")
    notes = lbr.get("lrp_mapping_notes") or []
    note_s = " ".join(str(n) for n in notes) if notes else "LBR: mapped billed and actual; see lrp_mapping_notes."
    rows = []
    if b is not None:
        rows.append(
            {
                "line": "Billed to date (or Spent to date as billed proxy, per LBR layout)",
                "value": float(b),
                "derivation": note_s,
            }
        )
    if a is not None:
        rows.append(
            {
                "line": "Actual cost (LAB row when headers use Cost Category + Labor (Actual Cost) or Billed/Actual header)",
                "value": float(a),
                "derivation": note_s,
            }
        )
    if lrp is not None and b is not None and a is not None:
        rows.append(
            {
                "line": "Final labor rate profit to date (billed to date minus actual cost)",
                "value": float(lrp),
                "derivation": "Deterministic: billed_to_date - actual_cost (same as headline LRP).",
            }
        )
    return rows


def build_financial_workbench(
    wps: dict[str, Any],
    jtd_block: dict[str, Any] | None,
    lbr_block: dict[str, Any] | None,
    *,
    jtd_sheet: str | None,
    lbr_sheet: str | None,
    mom_tpp: float | None,
) -> dict[str, Any]:
    """
    Deterministic drill-down package for the UI (no new calculations).
    """
    jtd = jtd_block or {}
    lbr = lbr_block or {}
    jt = str(jtd_sheet or "JTD")
    lt = str(lbr_sheet or "LBR")
    jnotes = list(jtd.get("jtd_mapping_notes") or [])
    lnotes = list(lbr.get("lrp_mapping_notes") or [])

    def jnote_has(sub: str) -> list[str]:
        return [n for n in jnotes if sub.lower() in str(n).lower()]

    comp: list[dict[str, Any]] = []

    def add_comp(
        cid: str,
        label: str,
        wkey: str,
        *,
        source_sheet: str,
        cost_code: str = "",
        cost_code_name: str = "",
        deriv: str = "",
    ) -> None:
        v = wps.get(wkey)
        lim: list[str] = []
        for pl in wps.get("projected_profit_limitations") or []:
            s = str(pl).lower()
            if cid.replace("_", " ") in s or wkey in s or label.lower() in s:
                if pl not in lim:
                    lim.append(str(pl))
        comp.append(
            {
                "id": cid,
                "label": label,
                "value": v,
                "source_sheet": source_sheet,
                "cost_code": cost_code,
                "cost_code_name": cost_code_name,
                "derivation": deriv,
                "limitations": lim,
            }
        )

    # Match N30 / N40 lines from label-aware JTD source rows (preferred) or scan cost table
    cmr = jtd.get("cm_fee_source_row")
    p40r = jtd.get("pco_profit_source_row")
    psr = jtd.get("prior_system_profit_jtd_source_row")
    n30u = n30n = ""
    n40u = n40n = ""
    psp_cc = psp_nm = ""
    if isinstance(cmr, dict) and (cmr.get("display_cost_code") or cmr.get("raw_cost_code")):
        n30u = str(cmr.get("display_cost_code") or cmr.get("cost_code") or "")
        n30n = str(cmr.get("cost_code_name") or "")
    else:
        for r in jtd.get("jtd_cost_code_rows") or []:
            if not isinstance(r, dict):
                continue
            if r.get("component_role") == "cm_fee":
                n30u = str(r.get("display_cost_code") or r.get("cost_code") or "")
                n30n = str(r.get("cost_code_name") or "")
                break
    if isinstance(p40r, dict) and (p40r.get("display_cost_code") or p40r.get("raw_cost_code")):
        n40u = str(p40r.get("display_cost_code") or p40r.get("cost_code") or "")
        n40n = str(p40r.get("cost_code_name") or "")
    else:
        for r in jtd.get("jtd_cost_code_rows") or []:
            if not isinstance(r, dict):
                continue
            if r.get("component_role") == "pco_profit":
                n40u = str(r.get("display_cost_code") or r.get("cost_code") or "")
                n40n = str(r.get("cost_code_name") or "")
                break
    if isinstance(psr, dict) and (psr.get("display_cost_code") or psr.get("raw_cost_code")):
        psp_cc = str(psr.get("display_cost_code") or psr.get("cost_code") or "")
        psp_nm = str(psr.get("cost_code_name") or "")

    add_comp(
        "cm_fee",
        "CM fee",
        "cm_fee",
        source_sheet=jt,
        cost_code=n30u or "",
        cost_code_name=n30n or "CM Fee (when a matching N30.000 / 30.000 label is present)",
        deriv="JTD: row with component_role cm_fee (Fee / CM fee label, not ORIG. profit claimed); value from Update Current Budget; "
        + (" ".join(jnote_has("cm_fee")[:1]) or " ".join(jnote_has("N30.000")[:1]) or "See jtd_mapping_notes."),
    )
    add_comp(
        "buyout_savings_realized",
        "Buyout savings realized",
        "buyout_savings_realized",
        source_sheet=jt,
        cost_code="(footer / Profit Calc line)",
        deriv="JTD: Buyout savings realized from label match in footer; see jtd_mapping_notes."
        if jtd.get("buyout_savings_realized") is not None
        else "Not found in this pass; value treated as null for TPP when missing (see limitations).",
    )
    add_comp(
        "budget_savings_overages",
        "Budget savings / overages",
        "budget_savings_overages",
        source_sheet=jt,
        cost_code="(footer or derived from contract minus budget total)",
        deriv="JTD: from Budget Overage / Budget Savings line, or original_contract - update current budget; see jtd_mapping_notes.",
    )
    add_comp(
        "pco_profit",
        "PCO profit",
        "pco_profit",
        source_sheet=jt,
        cost_code="N40.000 / 40.000" if n40u else "",
        cost_code_name=n40n or "PCO Profit (when a matching N40.000 / 40.000 label is present)",
        deriv="JTD: N40.000 with PCO Profit label, Update Current Budget; "
        + (" ".join([x for x in jnotes if "N40" in str(x) or "pco" in str(x).lower()][:1]) or "See jtd_mapping_notes."),
    )
    add_comp(
        "labor_rate_profit_to_date",
        "Labor rate profit to date",
        "labor_rate_profit_to_date",
        source_sheet=lt,
        cost_code="LAB / billed vs actual",
        deriv="LBR: billed to date (or Spent to date) minus actual cost on LAB row; " + " ".join(str(x) for x in lnotes[:2]),
    )
    add_comp(
        "prior_system_profit",
        "Prior-system profit (actual profit in old job)",
        "prior_system_profit",
        source_sheet=jt,
        cost_code=psp_cc or "Profit Calc / old job",
        cost_code_name=psp_nm
        or ("JTD cost line (N30.000 ORIG. profit style or Profit Calc)" if psp_cc else "Actual profit in old job (Profit Calc)"),
        deriv="JTD: ORIG. profit claimed / old-job JTD line when present, else Profit Calc 'Actual profit … old job'; see jtd_mapping_notes.",
    )

    add_comp(
        "workbook_reported_tpp",
        "Workbook-reported total projected profit",
        "workbook_reported_total_projected_profit",
        source_sheet="MoM Profit",
        deriv="MoM Profit sheet: last data row, column headed Total Projected Profit, when that sheet is present."
        if mom_tpp is not None
        else "Not read in this pass (no MoM total); see limitations.",
    )
    add_comp(
        "variance",
        "Variance (formula vs workbook-reported TPP)",
        "projected_profit_variance",
        source_sheet="(derived)",
        cost_code="",
        deriv="Deterministic: total_projected_profit (formula) minus workbook_reported_total_projected_profit when both are present.",
    )

    co_raw = jtd.get("change_order_rows") or []
    co_rows: list[dict[str, Any]] = []
    for r in co_raw:
        if isinstance(r, dict):
            co_rows.append({**r, "source_sheet": jt})

    tpp = wps.get("total_projected_profit")
    wr = wps.get("workbook_reported_total_projected_profit")
    var = wps.get("projected_profit_variance")
    rec_lines: list[dict[str, Any]] = [
        {"label": "Computed total projected profit (formula)", "value": tpp, "source": "Sum of signed formula terms in workbook_profit_summary."},
        {"label": "Workbook-reported total projected profit (MoM)", "value": wr, "source": "MoM Profit: last row under Total Projected Profit column, when read."},
        {"label": "Variance (formula vs workbook-reported)", "value": var, "source": "Computed TPP minus workbook-reported, when both exist."},
    ]
    rec_notes: list[str] = list(wps.get("projected_profit_limitations") or [])
    if jtd.get("prior_system_profit") is not None and isinstance(jtd.get("prior_system_profit"), (int, float)):
        rec_notes.append(
            "Prior-system (old job) profit is a signed term in the formula when present; differences vs workbook TPP can reflect MoM not matching the same JTD + LBR roll-up."
        )
    lrp_val = wps.get("labor_rate_profit_to_date")
    if lrp_val is not None:
        rec_notes.append(
            "LRP: CMiC/LBR billed vs actual feeds labor_rate_profit_to_date; workbook TPP is read from a different sheet (MoM) and can diverge for that reason."
        )
    if mom_tpp is not None and wr is not None and tpp is not None:
        rec_notes.append(
            "MoM-derived TPP (last cell) is the workbook's stated headline for comparison to the JTD+LBR formula TPP."
        )

    return {
        "jtd_sheet": jtd_sheet,
        "lbr_sheet": lbr_sheet,
        "mom_tpp_read": mom_tpp is not None,
        "total_original_project_costs": wps.get("total_original_project_costs"),
        "total_extended_project_costs": wps.get("total_extended_project_costs"),
        "projected_profit_component_sources": comp,
        "jtd_cost_code_rows": list(jtd.get("jtd_cost_code_rows") or []),
        "change_order_rows": co_rows,
        "lrp_source_rows": _lrp_presentation_block(lbr)
        + [
            {
                "line": f"LBR row (keyword match): {d.get('label')}",
                "value": (d.get("values") or [None])[0] if d.get("values") else None,
                "derivation": f"Scanned LBR line at excel_row={d.get('excel_row')} (all figures on the line: {d.get('values')})",
            }
            for d in _lbr_keyword_detail_rows(lbr)[:20]
        ],
        "lbr_workbook_rows": list(lbr.get("lbr_workbook_rows") or []),
        "reconciliation": {
            "lines": rec_lines,
            "deterministic_explanations": [n for n in rec_notes if n],
        },
    }


def extract_lbr_labor_rate_profit(rows: list[list[Any]]) -> dict[str, Any]:
    """
    LBR sheet: Labor Rate Profit to date = Billed to date - Actual cost.

    Observed 219128 layout (rows 1-based for reference):
    - R5: 'Cost Category' | 'Labor (Actual Cost)' — actual is under the latter column for code rows.
    - R6: 'LAB' | <actual cost>
    - A later row: 'Spent to date:' in col B, billed amount in col C (report title is BILLED LABOR RATES).

    If headers contain both 'billed' and 'actual', the legacy LAB row mapping is used.
    """
    lrp_notes: list[str] = []
    out: dict[str, Any] = {
        "lrp_billed_to_date": None,
        "lrp_actual_cost": None,
        "labor_rate_profit_to_date": None,
        "lrp_mapping_notes": lrp_notes,
        "lbr_workbook_rows": [],
    }
    notes = lrp_notes

    header_i = -1
    bill_c = act_c = -1
    for i, row in enumerate(rows[:20]):
        labs = [_norm_label(str(c) if c is not None else "") for c in row]
        if any("billed" in l for l in labs if l) and any("actual" in l for l in labs if l):
            header_i = i
            for j, c in enumerate(row):
                lab = _norm_label(str(c) if c is not None else "")
                if "billed" in lab:
                    bill_c = j
                if "actual" in lab:
                    act_c = j
            break

    if header_i >= 0 and bill_c >= 0 and act_c >= 0:
        for row in rows[header_i + 1 :]:
            if not row or len(row) <= max(bill_c, act_c):
                continue
            if _norm_label(str(row[0] if len(row) > 0 else "")) != "lab":
                continue
            b, a = _to_float(row[bill_c] if len(row) > bill_c else None), _to_float(
                row[act_c] if len(row) > act_c else None
            )
            if b is not None and a is not None:
                out["lrp_billed_to_date"] = b
                out["lrp_actual_cost"] = a
                out["labor_rate_profit_to_date"] = b - a
                notes.append("LBR: header row with 'billed' and 'actual'; LAB row for values.")
                out["lbr_workbook_rows"] = _scan_lbr_workbook_rows(rows)
                return out

    # Label scan for Billed to date / Spent to date (billed proxy in this report family)
    billed_val: float | None = None
    for i, row in enumerate(rows[:30]):
        for j, c in enumerate(row):
            lab = _norm_label(str(c) if c is not None else "")
            if not lab:
                continue
            if "billed to date" in lab:
                for k in range(j + 1, min(j + 4, len(row))):
                    v = _to_float(row[k])
                    if v is not None:
                        billed_val = v
                        notes.append("LBR: 'Billed to date' label cell, value in a nearby column.")
                        break
            elif lab.startswith("spent to date"):
                for k in range(j + 1, min(j + 4, len(row))):
                    v = _to_float(row[k])
                    if v is not None:
                        billed_val = v
                        notes.append(
                            "LBR: 'Spent to date' used as Billed to date (JOB COST REPORT BILLED LABOR RATES; "
                            "no separate 'Billed to date' column in sampled layout)."
                        )
                        break
        if billed_val is not None:
            break

    # Header: Cost Category | Labor (Actual Cost)
    hdr_row = -1
    act_col = -1
    for i, row in enumerate(rows[:15]):
        labs = [_norm_label(str(c) if c is not None else "") for c in row]
        if "cost category" in " ".join(labs) and any("labor" in l and "actual" in l for l in labs if l):
            hdr_row = i
            for j, c in enumerate(row):
                l2 = _norm_label(str(c) if c is not None else "")
                if "labor" in l2 and "actual" in l2:
                    act_col = j
                    break
            break
    actual_val: float | None = None
    if hdr_row >= 0 and act_col >= 0:
        for row in rows[hdr_row + 1 : min(hdr_row + 12, len(rows))]:
            if not row or len(row) <= act_col:
                continue
            r0 = _norm_label(str(row[0] if len(row) > 0 else ""))
            if r0 == "lab":
                actual_val = _to_float(row[act_col])
                if actual_val is not None:
                    notes.append(
                        "LBR: R5-style header 'Labor (Actual Cost)'; LAB row in column A, actual in that amount column."
                    )
                    break

    if actual_val is None:
        notes.append("LBR: could not determine Actual cost (expected LAB row under 'Labor (Actual Cost)').")
    if billed_val is None:
        notes.append("LBR: could not determine Billed to date (no 'Billed to date' or 'Spent to date' label with amount).")
    if billed_val is not None and actual_val is not None:
        out["lrp_billed_to_date"] = billed_val
        out["lrp_actual_cost"] = actual_val
        out["labor_rate_profit_to_date"] = billed_val - actual_val
    out["lbr_workbook_rows"] = _scan_lbr_workbook_rows(rows)
    return out


def extract_prior_system_profit_from_profit_calc(rows: list[list[Any]]) -> float | None:
    """
    JTD footer / Profit Calc block: row containing a cell 'Profit Calc' and, in the same row,
    label text like 'Actual profit to be claimed in old job' (workbooks may typo 'clamied').
    The amount is typically the first numeric cell strictly to the right of the 'Profit Calc' cell.
    """
    for row in rows:
        if not row:
            continue
        for j, cell in enumerate(row):
            lab = _norm_label(str(cell) if cell is not None else "")
            if lab != "profit calc":
                continue
            row_text = " ".join(_norm_label(str(c) if c is not None else "") for c in row)
            if "old job" not in row_text or "actual" not in row_text:
                continue
            for k in range(j + 1, min(j + 6, len(row))):
                v = _to_float(row[k] if len(row) > k else None)
                if v is not None:
                    return v
    return None


def extract_jtd_profit_inputs(rows: list[list[Any]], project_id: str = "") -> dict[str, Any]:
    """
    Deterministic fields from a JTD-style job cost sheet with:
    - A header row containing COST CODE and Update Current Budget
    - Optional footer block with original contract, revised total, budget variance, buyout realized, profit walk.
    - Cost lines N30.000 (Fee) and N40.000 (PCO Profit) when present.
    - Change order lines: cost codes 18* / 21* (N18/N21 for project 219128 only) summed on Update Current Budget.
    """
    jtd_notes: list[str] = []
    out: dict[str, Any] = {
        "cm_fee": None,
        "pco_profit": None,
        "original_contract_cost": None,
        "update_current_budget_total": None,
        "update_current_budget_sum_column_f": None,
        "budget_savings_overages": None,
        "buyout_savings_realized": None,
        "prior_system_profit": None,
        "owner_change_orders_count": 0,
        "owner_change_orders_value": 0.0,
        "cm_change_orders_count": 0,
        "cm_change_orders_value": 0.0,
        "change_order_source_notes": [],
        "jtd_mapping_notes": jtd_notes,
        "jtd_cost_code_rows": [],
        "change_order_rows": [],
        "owner_change_order_rows": [],
        "cm_change_order_rows": [],
        "jtd_profit_label_warnings": [],
        "cm_fee_source_row": None,
        "pco_profit_source_row": None,
        "prior_system_profit_jtd_source_row": None,
    }
    notes = jtd_notes
    if not rows:
        notes.append("JTD: empty sheet.")
        return out

    header_i = -1
    f_idx = -1
    for i, row in enumerate(rows[:40]):
        cols = [_norm_label(str(c) if c is not None else "") for c in row]
        joined = " ".join(c for c in cols if c)
        if "cost code" in joined and "update" in joined and "budget" in joined:
            header_i = i
            for j, lab in enumerate(cols):
                if "update" in lab and "budget" in lab:
                    f_idx = j
                    break
            if f_idx < 0:
                for j, lab in enumerate(cols):
                    if "update current budget" in lab or lab == "update current budget":
                        f_idx = j
            break
    if header_i < 0 or f_idx < 0:
        notes.append("JTD: header row with 'Update Current Budget' not found.")
        notes.append("JTD: change order totals not computed (requires JTD header + cost code block).")
    else:
        jtd_code_block, jtd_co_block = _collect_jtd_table_and_change_order_rows(
            rows, header_i, f_idx, project_id
        )
        out["jtd_cost_code_rows"] = jtd_code_block
        out["change_order_rows"] = jtd_co_block
        o_rows = [r for r in jtd_co_block if r.get("co_type") == "Owner"]
        c_rows = [r for r in jtd_co_block if r.get("co_type") == "CM"]
        out["owner_change_order_rows"] = o_rows
        out["cm_change_order_rows"] = c_rows
        out["owner_change_orders_count"] = len(o_rows)
        out["owner_change_orders_value"] = sum(float(r.get("update_current_budget") or 0) for r in o_rows)
        out["cm_change_orders_count"] = len(c_rows)
        out["cm_change_orders_value"] = sum(float(r.get("update_current_budget") or 0) for r in c_rows)
        col_f_sum = 0.0
        counted = 0
        for row in rows[header_i + 1 :]:
            if not row or len(row) <= f_idx:
                continue
            if row[0] is None or str(row[0]).strip() == "":
                break
            if not _row_looks_like_jtd_cost_line(row, project_id):
                continue
            v = _to_float(row[f_idx] if len(row) > f_idx else None)
            if v is not None:
                col_f_sum += v
                counted += 1
        if counted > 0:
            out["update_current_budget_sum_column_f"] = col_f_sum
            notes.append(
                f"JTD: sum of Update Current Budget over {counted} JTD data row(s) in the contiguous block (non-data rows are skipped, first blank in column A ends the block)."
            )
        ccode_i = 0
        name_i = 1
        prior_jtd_val: float | None = None
        label_warn: list[str] = []
        for k, row in enumerate(rows[header_i + 1 :]):
            if not row or len(row) <= max(f_idx, name_i):
                continue
            if row[0] is None or str(row[0]).strip() == "":
                break
            if not _row_looks_like_jtd_cost_line(row, project_id):
                continue
            code_raw = str(row[ccode_i] or "").strip()
            if not code_raw:
                break
            name_raw = str(row[name_i] if len(row) > name_i else "")
            ncode = _norm_cost_code(row[ccode_i])
            vf = _to_float(row[f_idx] if len(row) > f_idx else None)
            if vf is None:
                continue
            exr = _jtd_data_row_number(header_i, k)
            cf = _jtd_row_code_fields(code_raw, project_id)
            sem = _jtd_row_semantic_fields(code_raw, cf["normalized_cost_code"], name_raw, project_id)
            if ncode in ("30.000",):
                if is_prior_system_jtd_cost_name(name_raw):
                    prev_sr = out.get("prior_system_profit_jtd_source_row")
                    prev_ucb = float(prev_sr.get("update_current_budget") or 0) if isinstance(prev_sr, dict) else 0.0
                    cand_sr = {
                        **cf,
                        **sem,
                        "cost_code_name": name_raw,
                        "update_current_budget": float(vf),
                        "excel_row": exr,
                    }
                    if prev_sr is None or abs(float(vf)) > abs(prev_ucb) + 0.02:
                        out["prior_system_profit_jtd_source_row"] = cand_sr
                    if prior_jtd_val is None and abs(float(vf)) > 0.02:
                        prior_jtd_val = float(vf)
                        out["prior_system_profit"] = float(vf)
                        notes.append(
                            "JTD: prior_system_profit from N30.000 / 30.000 (ORIG. profit claimed / old-job style label) on cost line."
                        )
                elif is_cm_fee_jtd_cost_name(name_raw):
                    if out.get("cm_fee") is None:
                        out["cm_fee"] = float(vf)
                        out["cm_fee_source_row"] = {
                            **cf,
                            **sem,
                            "cost_code_name": name_raw,
                            "update_current_budget": float(vf),
                            "excel_row": exr,
                        }
                        notes.append("JTD: cm_fee from N30.000 / 30.000 with fee / CM fee label (Update Current Budget).")
                else:
                    label_warn.append(
                        f"N30.000 ({code_raw} / {name_raw}): cost code matched, but label did not match expected CM Fee or prior-system profit."
                    )
            if ncode in ("40.000",):
                if is_pco_profit_jtd_cost_name(name_raw):
                    if out.get("pco_profit") is None:
                        out["pco_profit"] = float(vf)
                        out["pco_profit_source_row"] = {
                            **cf,
                            **sem,
                            "cost_code_name": name_raw,
                            "update_current_budget": float(vf),
                            "excel_row": exr,
                        }
                        notes.append("JTD: pco_profit from N40.000 / 40.000 with PCO Profit label (Update Current Budget).")
                else:
                    label_warn.append(
                        f"N40.000 ({code_raw} / {name_raw}): cost code matched, but label did not match expected PCO Profit."
                    )
        out["jtd_profit_label_warnings"] = label_warn
        for w in label_warn:
            notes.append("JTD warning: " + w)

        co_agg = _aggregate_change_order_lines(rows, header_i, f_idx, project_id)
        out["change_order_source_notes"] = list(co_agg.get("change_order_source_notes") or [])
        out["change_order_source_notes"].append(
            f"JTD: {len(o_rows)} owner CO row(s) (18 / N18) and {len(c_rows)} CM CO row(s) (21 / N21) in extracted block; counts/values are derived from those lists."
        )

    g_idx = 6
    f_idx2 = 5
    for row in rows:
        if len(row) <= f_idx2:
            continue
        gtext = str(row[g_idx] if len(row) > g_idx else "") or ""
        g_lower = gtext.lower()
        fval = _to_float(row[f_idx2] if len(row) > f_idx2 else None)
        if any(h in g_lower for h in _ORIGINAL_CONTRACT_HINTS) and "revised" not in g_lower:
            if fval is not None and "pending" not in g_lower:
                if out["original_contract_cost"] is None:
                    out["original_contract_cost"] = fval
                    notes.append("JTD: original contract from footer (column G: Current contract / Payment App).")
        if any(h in g_lower for h in _REVISED_BUDGET_TOTAL_HINTS) and fval is not None:
            out["update_current_budget_total"] = fval
            notes.append("JTD: update current budget total from footer (Revised current budget / Total project costs).")
        if any(h in g_lower for h in _BUDGET_VARIANCE_HINTS) and fval is not None:
            out["budget_savings_overages"] = fval
            notes.append("JTD: budget savings/overage from footer (Budget Overage / Budget Savings).")
        if any(h in g_lower for h in _BUYOUT_REALIZED_HINTS) and fval is not None:
            out["buyout_savings_realized"] = fval
            notes.append("JTD: Buyout savings realized from Profit Calc / footer (column G label).")

    if out["budget_savings_overages"] is None:
        o = out["original_contract_cost"]
        u = out.get("update_current_budget_total")
        if u is None:
            u = out.get("update_current_budget_sum_column_f")
        if isinstance(o, (int, float)) and isinstance(u, (int, float)):
            out["budget_savings_overages"] = float(o) - float(u)
            notes.append("JTD: budget_savings_overages = original_contract_cost - update_current_budget_total (or sum F).")
    if out.get("update_current_budget_total") is None and out.get("update_current_budget_sum_column_f") is not None:
        out["update_current_budget_total"] = out["update_current_budget_sum_column_f"]
        notes.append("JTD: using summed column F as update current budget total (footer total not found).")
    if out["buyout_savings_realized"] is None:
        notes.append("JTD: buyout_savings_realized not found in footer (no 'Buyout savings realized' line with value).")
    psp_calc = extract_prior_system_profit_from_profit_calc(rows)
    if out.get("prior_system_profit") is None and psp_calc is not None:
        out["prior_system_profit"] = float(psp_calc)
        notes.append(
            "JTD: prior_system_profit from Profit Calc row (label e.g. 'Actual profit to be claimed in old job')."
        )
    elif psp_calc is not None and out.get("prior_system_profit") is not None:
        try:
            if abs(float(psp_calc) - float(out["prior_system_profit"])) > 0.02:
                notes.append(
                    f"JTD: Profit Calc old-job amount ({float(psp_calc):.2f}) vs cost-line prior value ({float(out['prior_system_profit']):.2f}); "
                    "formula uses cost-line JTD value when set."
                )
        except (TypeError, ValueError):
            pass
    return out


def _build_workbook_profit_summary(
    lbr_block: dict[str, Any] | None,
    jtd_block: dict[str, Any] | None,
    lbr_limitation: str | None,
    *,
    workbook_reported_total_projected_profit: float | None = None,
) -> dict[str, Any]:
    lbr = lbr_block or {}
    jtd = jtd_block or {}
    limitations: list[str] = []
    if lbr_limitation:
        limitations.append(lbr_limitation)
    if not lbr_limitation and lbr.get("labor_rate_profit_to_date") is None:
        n = lbr.get("lrp_mapping_notes") or []
        if not n or any("could not" in str(x).lower() for x in n):
            limitations.append("LBR: labor rate profit not computed (missing billed/actual on LBR sheet).")
    for n in lbr.get("lrp_mapping_notes") or []:
        if "could not" in str(n).lower():
            limitations.append(f"LBR: {n}")
    if not jtd_block:
        limitations.append("JTD sheet: not found (expected a sheet with 'JTD' in the name, excluding JTD_Static).")
    for n in jtd.get("jtd_mapping_notes") or []:
        if "header row" in str(n).lower() and "not found" in str(n).lower():
            limitations.append(f"JTD: {n}")
    for w in jtd.get("jtd_profit_label_warnings") or []:
        limitations.append("Cost code matched, but label did not match expected component: " + str(w))
    if jtd.get("buyout_savings_realized") is None:
        limitations.append("Buyout savings realized: source not yet mapped in this workbook (no footer value).")
    if jtd_block and jtd.get("prior_system_profit") is None:
        limitations.append(
            "Prior-system profit (Actual profit to be claimed in old job / Profit Calc): not found; using 0."
        )
    psp_in = jtd.get("prior_system_profit")
    prior_system_profit: float = float(psp_in) if isinstance(psp_in, (int, float)) else 0.0
    lrp = lbr.get("labor_rate_profit_to_date")
    cm = jtd.get("cm_fee")
    pco = jtd.get("pco_profit")
    bso = jtd.get("buyout_savings_realized")
    bo = bso
    tpp: float | None = None
    tpp_blockers: list[str] = []
    if lrp is None:
        tpp_blockers.append("labor_rate_profit_to_date")
    if cm is None:
        tpp_blockers.append("cm_fee")
    if pco is None:
        tpp_blockers.append("pco_profit")
    if bso is None:
        tpp_blockers.append("buyout_savings_realized")
    if jtd.get("budget_savings_overages") is None:
        tpp_blockers.append("budget_savings_overages")
    if not tpp_blockers and isinstance(lrp, (int, float)) and jtd.get("budget_savings_overages") is not None:
        tpp = (  # type: ignore[operator]
            float(cm) + float(bo) + float(jtd["budget_savings_overages"]) + float(pco) + float(lrp) + prior_system_profit
        )
    elif tpp_blockers:
        limitations.append("Total Projected Profit: " + (", ".join(tpp_blockers) + " missing (see snapshot fields)."))

    w_reported = workbook_reported_total_projected_profit
    pp_var: float | None = None
    if w_reported is not None and isinstance(tpp, (int, float)):
        pp_var = float(tpp) - float(w_reported)
    if w_reported is None:
        limitations.append("Workbook-reported total projected profit: not read from a MoM Profit sheet in this pass.")

    co_notes = list(jtd.get("change_order_source_notes") or [])
    jtd_rows_for_totals = jtd.get("jtd_cost_code_rows") or []
    total_original_project_costs, total_extended_project_costs = _jtd_namespace_budget_totals(
        jtd_rows_for_totals if isinstance(jtd_rows_for_totals, list) else []
    )

    return {
        "labor_rate_profit_to_date": lbr.get("labor_rate_profit_to_date"),
        "lrp_billed_to_date": lbr.get("lrp_billed_to_date"),
        "lrp_actual_cost": lbr.get("lrp_actual_cost"),
        "owner_change_order_rows": list(jtd.get("owner_change_order_rows") or []),
        "cm_change_order_rows": list(jtd.get("cm_change_order_rows") or []),
        "cm_fee": jtd.get("cm_fee"),
        "budget_savings_overages": jtd.get("budget_savings_overages"),
        "original_contract_cost": jtd.get("original_contract_cost"),
        "update_current_budget_total": jtd.get("update_current_budget_total"),
        "pco_profit": jtd.get("pco_profit"),
        "buyout_savings_realized": jtd.get("buyout_savings_realized"),
        "prior_system_profit": prior_system_profit,
        "owner_change_orders_count": int(jtd.get("owner_change_orders_count") or 0),
        "owner_change_orders_value": float(jtd.get("owner_change_orders_value") or 0.0),
        "cm_change_orders_count": int(jtd.get("cm_change_orders_count") or 0),
        "cm_change_orders_value": float(jtd.get("cm_change_orders_value") or 0.0),
        "change_order_source_notes": co_notes,
        "workbook_reported_total_projected_profit": w_reported,
        "projected_profit_variance": pp_var,
        "total_projected_profit": tpp,
        "total_projected_profit_formula_terms": list(TOTAL_PROJECTED_PROFIT_FORMULA_TERMS),
        "total_original_project_costs": total_original_project_costs,
        "total_extended_project_costs": total_extended_project_costs,
        "projected_profit_limitations": sorted(set(limitations)) if limitations else [],
        "jtd_profit_extraction": jtd,
        "lbr_profit_extraction": lbr,
    }


def _label_looks_like_category_header(lab: str) -> bool:
    """True if a normalized header cell plausibly names a category / line bucket column."""
    if not lab:
        return False
    if "category" in lab:
        return True
    if "trade" in lab or "division" in lab or "csi" in lab:
        return True
    if any(lab == x or lab.startswith(x + " ") for x in ("item", "description")):
        return True
    return False


def _best_amount_col(labels: list[str]) -> int:
    """Choose a numeric value column, preferring current / spent columns over static budget."""
    for hint in _AMOUNT_HEADER_ORDER:
        for j, lab in enumerate(labels):
            if lab and hint in lab:
                return j
    for j, lab in enumerate(labels):
        if not lab:
            continue
        if any(
            k in lab
            for k in (
                "amount",
                "cost",
                "value",
                "budget",
                "total",
                "profit",
                "revenue",
            )
        ):
            return j
    return -1


def _best_cat_col(labels: list[str]) -> int:
    for j, lab in enumerate(labels):
        if "category" in lab:
            return j
    for j, lab in enumerate(labels):
        if any(k in lab for k in ("trade", "division", "csi", "item", "description")):
            return j
    for j, lab in enumerate(labels):
        if "name" in lab and "code" not in lab and "code name" not in lab:
            return j
    return 0


def _row_category_label(row: list[Any], cat_col: int) -> str:
    """
    One label per line item. Use cost code + description + category when the sheet
    is cost-code shaped so drivers do not collapse many rows to 'LAB'/'MAT'.
    """
    raw = str(row[cat_col]).strip() if row[cat_col] is not None else ""
    c0 = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
    if c0 and _CODE_LIKE.match(c0) and len(row) > max(2, cat_col):
        c1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        c2 = raw
        return " | ".join(p for p in (c0, c1, c2) if p)
    return raw or c0


def _extract_mom_last_total_projected_profit(rows: list[list[Any]]) -> float | None:
    """'MoM Profit' sheet: last data row, 'Total Projected Profit' column."""
    header_i = -1
    for i, row in enumerate(rows[:30]):
        labs = [_norm_label(str(c) if c is not None else "") for c in row]
        if any("total projected profit" in lab for lab in labs if lab):
            header_i = i
            break
    if header_i < 0:
        return None
    pcol = -1
    for j, c in enumerate(rows[header_i]):
        lab = _norm_label(str(c) if c is not None else "")
        if "total projected profit" in lab:
            pcol = j
            break
    if pcol < 0:
        return None
    last: float | None = None
    for row in rows[header_i + 1:]:
        if len(row) <= pcol:
            continue
        v = _to_float(row[pcol])
        if v is not None:
            last = v
    return last


def _extract_lbr_lab_billed_actual(rows: list[list[Any]]) -> dict[str, float]:
    """
    '219128 LBR' style sheet: map billed -> revenue, actual -> cost
    (stable keys for `summary_deltas` in handlers).
    """
    d = extract_lbr_labor_rate_profit(rows)
    b, a = d.get("lrp_billed_to_date"), d.get("lrp_actual_cost")
    if b is not None and a is not None:
        return {"revenue": float(b), "cost": float(a)}
    return {}


def _extract_profit_report_family_summary(sheet_title: str, rows: list[list[Any]]) -> dict[str, float]:
    """This workbook family: 'MoM Profit' + '… LBR …' (not JTD) carry the headline numbers."""
    out: dict[str, float] = {}
    nt = _norm_label(sheet_title)
    if "jtd" in nt:
        return out
    if "mom" in nt and "profit" in nt:
        tpp = _extract_mom_last_total_projected_profit(rows)
        if tpp is not None:
            out["profit"] = tpp
    if "lbr" in nt:
        out.update(_extract_lbr_lab_billed_actual(rows))
    return out


def _extract_summary(rows: list[list[Any]]) -> dict[str, float]:
    """Legacy: label-in-column-A style lines. Prefer `_extract_profit_report_family_summary` for real profit reports."""
    summary: dict[str, float] = {}
    keys = {
        "revenue": ("revenue", "sales", "income"),
        "cost": ("cost", "expense", "total cost"),
        "profit": ("profit", "margin", "ebitda", "net income"),
    }
    for row in rows:
        if not row:
            continue
        first = _norm_label(str(row[0]))
        if not first:
            continue
        nums = [n for n in (_to_float(v) for v in row[1:]) if n is not None]
        if not nums:
            continue
        candidate = nums[-1]
        for field, aliases in keys.items():
            if any(alias in first for alias in aliases):
                summary[field] = candidate
                break
    if "profit" not in summary and "revenue" in summary and "cost" in summary:
        summary["profit"] = summary["revenue"] - summary["cost"]
    return summary


def _detect_category_table(rows: list[list[Any]]) -> tuple[int, int] | None:
    for idx, row in enumerate(rows[:60]):
        labels = [_norm_label(str(c)) for c in row]
        if not labels:
            continue
        if not any(_label_looks_like_category_header(lab) for lab in labels if lab):
            continue
        amount_col = _best_amount_col(labels)
        if amount_col < 0:
            continue
        cat_col = _best_cat_col(labels)
        if cat_col == amount_col:
            return None
        return cat_col, amount_col
    return None


def _build_extraction_block(
    scanned_sheets: list[str],
    summary: dict[str, float],
    categories: list[dict[str, float | str]],
    evidence: list[str],
) -> dict[str, Any]:
    sk = set(summary.keys())
    have = {k: k in sk for k in ("revenue", "cost", "profit")}
    complete = all(have.values())
    ncat = len(categories)
    if complete and ncat >= 2:
        conf = "high"
    elif complete or (len(sk) >= 2 and ncat >= 1):
        conf = "medium"
    else:
        conf = "low"
    notes: list[str] = []
    if not complete:
        miss = [k for k in ("revenue", "cost", "profit") if not have[k]]
        if miss:
            notes.append(
                "Summary line detection incomplete (expected labels for): " + ", ".join(miss) + "."
            )
    if ncat == 0:
        notes.append("No category table detected in scanned sheets; line-item drivers may be empty.")
    elif ncat < 2:
        notes.append("Few categories extracted; verify the sheet and column headers used for line items.")
    return {
        "confidence": conf,
        "sheets_scanned": scanned_sheets,
        "category_count": ncat,
        "summary_keys_found": sorted(sk),
        "summary_complete": complete,
        "evidence": evidence[:20],
        "notes": notes,
    }


def _extract_categories(rows: list[list[Any]]) -> list[dict[str, float | str]]:
    header = _detect_category_table(rows)
    if header is None:
        return []
    cat_col, amount_col = header
    categories: list[dict[str, float | str]] = []
    for row in rows:
        if len(row) <= max(cat_col, amount_col):
            continue
        name = _row_category_label(row, cat_col)
        if not name or name.lower() in {"total", "grand total"}:
            continue
        if _norm_label(name) in ("cost category", "cost code", "category"):
            continue
        value = _to_float(row[amount_col])
        if value is None:
            continue
        categories.append({"name": name, "value": value})
    return categories


def extract_financial_snapshot_from_workbook(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    project_id = infer_project_id_from_workbook_path(path)
    summary: dict[str, float] = {}
    categories: list[dict[str, float | str]] = []
    sheet_evidence: list[str] = []
    scanned_sheets: list[str] = []

    lbr_block: dict[str, Any] | None = None
    jtd_block: dict[str, Any] | None = None
    lbr_limitation: str | None = None
    t_lbr = _find_lbr_sheet_title(wb)
    if t_lbr is not None:
        ws_l = wb[t_lbr]
        lbr_rows: list[list[Any]] = []
        for row in ws_l.iter_rows(min_row=1, max_row=80, max_col=16, values_only=True):
            lbr_rows.append(list(row))
        lbr_block = extract_lbr_labor_rate_profit(lbr_rows)
    else:
        lbr_limitation = "LBR sheet not found (no worksheet with 'LBR' in the tab name; labor rate profit skipped)."
    t_jtd = _find_primary_jtd_sheet_title(wb)
    if t_jtd is not None:
        ws_j = wb[t_jtd]
        jtd_rows: list[list[Any]] = []
        jtd_max = min(int(ws_j.max_row or 2000), 2000)
        for row in ws_j.iter_rows(min_row=1, max_row=jtd_max, min_col=1, max_col=12, values_only=True):
            jtd_rows.append(list(row))
        jtd_block = extract_jtd_profit_inputs(jtd_rows, project_id=project_id)
    else:
        jtd_block = None
    mom_tpp: float | None = None
    for ws in wb.worksheets[:8]:
        nt = _norm_label(ws.title)
        if "mom" in nt and "profit" in nt:
            rows_m: list[list[Any]] = []
            for row in ws.iter_rows(min_row=1, max_row=160, max_col=24, values_only=True):
                vals = list(row)
                if any(v not in (None, "") for v in vals):
                    rows_m.append(vals)
            if rows_m:
                tpp_m = _extract_mom_last_total_projected_profit(rows_m)
                if tpp_m is not None:
                    mom_tpp = tpp_m
    workbook_profit_summary = _build_workbook_profit_summary(
        lbr_block, jtd_block, lbr_limitation, workbook_reported_total_projected_profit=mom_tpp
    )
    wps = workbook_profit_summary
    lbr_t = t_lbr or "absent"
    jtd_t = t_jtd or "absent"
    sheet_evidence.append(
        f"workbook_profit: LBR_tab={lbr_t} JTD_tab={jtd_t} lrp={wps.get('labor_rate_profit_to_date')!s} tpp={wps.get('total_projected_profit')!s}"
    )

    for ws in wb.worksheets[:8]:
        scanned_sheets.append(ws.title)
        rows: list[list[Any]] = []
        for row in ws.iter_rows(min_row=1, max_row=160, max_col=24, values_only=True):
            vals = list(row)
            if any(v not in (None, "") for v in vals):
                rows.append(vals)
        if not rows:
            continue
        family = _extract_profit_report_family_summary(ws.title, rows)
        for k, v in family.items():
            summary[k] = v
        s = _extract_summary(rows)
        for k, v in s.items():
            if k not in summary:
                summary[k] = v
        cat = _extract_categories(rows)
        if cat and len(cat) > len(categories):
            categories = cat
            sheet_evidence.append(f"{ws.title}: category_rows={len(cat)}")
        elif s:
            sheet_evidence.append(f"{ws.title}: summary_keys={sorted(s.keys())}")
        if family:
            sheet_evidence.append(
                f"{ws.title}: profit_report_family={sorted(family.keys())} values={{{', '.join(f'{k}={v}' for k, v in family.items())}}}"
            )

    if "profit" not in summary and "revenue" in summary and "cost" in summary:
        summary["profit"] = summary["revenue"] - summary["cost"]

    extraction = _build_extraction_block(scanned_sheets, summary, categories[:300], sheet_evidence)
    financial_workbench = build_financial_workbench(
        wps,
        jtd_block,
        lbr_block,
        jtd_sheet=t_jtd,
        lbr_sheet=t_lbr,
        mom_tpp=mom_tpp,
    )
    return {
        "schema": SNAPSHOT_SCHEMA,
        "source_file": str(path.resolve()),
        "summary": summary,
        "categories": categories[:300],
        "evidence": sheet_evidence[:20],
        "extraction": extraction,
        "workbook_profit_summary": workbook_profit_summary,
        "financial_workbench": financial_workbench,
    }


def write_snapshot_json(path: Path, snapshot: dict[str, Any]) -> None:
    path.write_text(json.dumps(snapshot, indent=2) + "\n", encoding="utf-8")


def load_snapshot_json(path: Path) -> dict[str, Any] | None:
    if path.suffix.lower() != ".json":
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if payload.get("schema") != SNAPSHOT_SCHEMA:
        return None
    return payload


def snapshot_to_markdown(snapshot: dict[str, Any]) -> str:
    lines = ["# Financial Intake Debug", ""]
    lines.append(f"- Schema: `{snapshot.get('schema', 'unknown')}`")
    lines.append(f"- Source: `{snapshot.get('source_file', 'n/a')}`")
    lines.append("")
    lines.append("## Summary")
    summary = snapshot.get("summary", {})
    if isinstance(summary, dict) and summary:
        for k in ("revenue", "cost", "profit"):
            if k in summary:
                lines.append(f"- {k}: {summary[k]}")
    else:
        lines.append("- (none)")
    lines.append("")
    lines.append("## Category sample")
    cats = snapshot.get("categories", [])
    if isinstance(cats, list) and cats:
        for row in cats[:20]:
            lines.append(f"- {row.get('name')}: {row.get('value')}")
    else:
        lines.append("- (none)")
    wps = snapshot.get("workbook_profit_summary")
    lines.append("")
    lines.append("## Workbook profit summary (deterministic)")
    if isinstance(wps, dict) and wps:
        for k in (
            "labor_rate_profit_to_date",
            "total_projected_profit",
            "workbook_reported_total_projected_profit",
            "projected_profit_variance",
            "cm_fee",
            "buyout_savings_realized",
            "budget_savings_overages",
            "prior_system_profit",
            "owner_change_orders_count",
            "owner_change_orders_value",
            "cm_change_orders_count",
            "cm_change_orders_value",
        ):
            if k in wps and wps[k] is not None:
                lines.append(f"- {k}: {wps[k]}")
        plim = wps.get("projected_profit_limitations") or []
        if plim:
            lines.append(f"- limitations: {plim!s}")
    else:
        lines.append("- (none)")
    lines.append("")
    lines.append("## Evidence")
    evidence = snapshot.get("evidence", [])
    if isinstance(evidence, list) and evidence:
        for item in evidence:
            lines.append(f"- {item}")
    else:
        lines.append("- (none)")
    ext = snapshot.get("extraction")
    lines.append("")
    lines.append("## Extraction confidence")
    if isinstance(ext, dict) and ext:
        lines.append(f"- confidence: {ext.get('confidence', 'n/a')}")
        lines.append(f"- sheets_scanned: {ext.get('sheets_scanned', [])}")
        lines.append(f"- category_count: {ext.get('category_count', 0)}")
        for note in ext.get("notes", []) or []:
            lines.append(f"- {note}")
    else:
        lines.append("- (none — legacy snapshot)")
    return "\n".join(lines).strip() + "\n"
